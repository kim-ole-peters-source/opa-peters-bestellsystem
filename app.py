#!/usr/bin/env python3
# Internes Bestellsystem ohne Preise
# Start: python app.py, dann http://localhost:8000 öffnen

import os
import re
import csv
import io
import html
import uuid
import json
import hmac
import base64
import hashlib
import sqlite3
import smtplib
import calendar
from datetime import datetime, date, timedelta
from email.message import EmailMessage
from http import cookies
from urllib.parse import parse_qs, urlparse, quote_plus
from email.parser import BytesParser
from email.policy import default as email_policy
from http.server import HTTPServer, BaseHTTPRequestHandler
try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Vercel ist serverless: Schreibzugriff ist dort nur unter /tmp zuverlässig möglich.
# Lokal bleibt alles im Projektordner, damit die Mac-Version weiter funktioniert.
IS_VERCEL = bool(os.environ.get("VERCEL") or os.environ.get("VERCEL_ENV"))
DATA_DIR = os.environ.get("DATA_DIR") or ("/tmp/opa-peters-bestellsystem" if IS_VERCEL else BASE_DIR)
os.makedirs(DATA_DIR, exist_ok=True)

DB_PATH = os.path.join(DATA_DIR, "bestellsystem.db")
UPLOAD_DIR = os.path.join(DATA_DIR, "uploads")
ORDER_DIR = os.path.join(DATA_DIR, "orders")
ORDER_IMAGE_DIR = os.path.join(DATA_DIR, "order_images")
TIME_EXPORT_DIR = os.path.join(DATA_DIR, "time_exports")
SETTINGS_PATH = os.path.join(DATA_DIR, "settings.json")

# Startdaten beim ersten Serverless-Start nach /tmp kopieren.
for _name in ["bestellsystem.db", "settings.json"]:
    _src = os.path.join(BASE_DIR, _name)
    _dst = os.path.join(DATA_DIR, _name)
    if not os.path.exists(_dst) and os.path.exists(_src):
        try:
            import shutil
            shutil.copy2(_src, _dst)
        except Exception:
            pass
ADMIN_PIN = os.environ.get("ADMIN_PIN", "1234")  # Bitte vor Onlinebetrieb ändern!
APP_SECRET = os.environ.get("APP_SECRET", ADMIN_PIN + "-lokal-test")
HOST = os.environ.get("HOST", "127.0.0.1")
PORT = int(os.environ.get("PORT", "8000"))

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(ORDER_DIR, exist_ok=True)
os.makedirs(ORDER_IMAGE_DIR, exist_ok=True)
os.makedirs(TIME_EXPORT_DIR, exist_ok=True)

BUYER_ACCOUNTS = [
    {
        "key": "filialisten_eis",
        "label": "Filialisten Eis",
        "pin": os.environ.get("BUYER_PIN_FILIALISTEN_EIS", "1111"),
    },
    {
        "key": "filialisten_feine_kost",
        "label": "Filialisten Feine Kost",
        "pin": os.environ.get("BUYER_PIN_FILIALISTEN_FEINE_KOST", "2222"),
    },
    {
        "key": "produzenten",
        "label": "Produzenten",
        "pin": os.environ.get("BUYER_PIN_PRODUZENTEN", "3333"),
    },
    {
        "key": "boss",
        "label": "Boss",
        "pin": os.environ.get("BUYER_PIN_BOSS", "4444"),
    },
]
BUYER_BY_KEY = {x["key"]: x for x in BUYER_ACCOUNTS}
ALL_LOCATIONS_KEY = "all_locations"
NO_LOCATIONS_KEY = "no_locations"
DEFAULT_VISIBLE_TO = ALL_LOCATIONS_KEY
DEFAULT_CATEGORIES = ["Allgemein", "Eis", "Feine Kost", "Reinigung", "Verbrauchsmaterial"]
ACCESS_ROLES = [
    ("standard", "Standort"),
    ("b2b", "B2B Kunde"),
    ("order_only", "Nur Bestellungen"),
    ("time_only", "Nur Zeiterfassung"),
    ("manager", "Filialleitung"),
]
ACCESS_ROLE_KEYS = {key for key, _label in ACCESS_ROLES}

DEFAULT_SETTINGS = {
    "locations": [
        {"id": "filiale-1", "name": "Filiale 1", "password": ""},
        {"id": "filiale-2", "name": "Filiale 2", "password": ""},
        {"id": "filiale-3", "name": "Filiale 3", "password": ""},
        {"id": "zentrale", "name": "Zentrale", "password": ""},
    ],
    "order_email_to": "",
    "whatsapp_number": "",
    "central_email_hint": "zentrale@example.com",
    "admin_pin_hash": "",
    "time_export_email": "info@opapeters",
    "last_auto_time_export_month": "",
    "last_auto_time_export_status": "",
    "time_employees": [],
}

APP_NAME = "Opa Peters Bestellung"
APP_SHORT_NAME = "OP Bestellung"
THEME_COLOR = "#1e3a8a"
ASSET_VERSION = "2026-08-27-task-progress-order-cleanup"
BACKGROUND_COLOR = "#f6f7fb"
MAX_FORM_BYTES = 12 * 1024 * 1024
MAX_CART_DRAFT_BYTES = 220 * 1024
MAX_PUSH_SUBSCRIPTION_BYTES = 24 * 1024
MAX_IMAGE_BYTES = 6 * 1024 * 1024
MAX_ORDER_IMAGE_BYTES = 10 * 1024 * 1024
MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_ORDER_QUANTITY = 9999
BERLIN_TZ = ZoneInfo("Europe/Berlin") if ZoneInfo else None
COCKPIT_API_TOKEN = os.environ.get("COCKPIT_API_TOKEN", "").strip()
PRODUCTION_SECTIONS = {
    "backen": "Backen",
    "eisvitrinen": "Eisvitrinen",
    "eistorten": "Eistorten",
}


class UploadedFile:
    def __init__(self, filename, content):
        self.filename = filename or ""
        self.content = content or b""


class RequestTooLarge(Exception):
    pass


def berlin_now():
    return datetime.now(BERLIN_TZ) if BERLIN_TZ else datetime.now()


def hash_pin(pin):
    raw = str(pin or "").encode("utf-8")
    digest = hashlib.sha256(b"opa-peters-pin:" + raw).hexdigest()
    return f"sha256:{digest}"


def verify_pin(pin, stored_hash):
    stored_hash = stored_hash or hash_pin(ADMIN_PIN)
    return hmac.compare_digest(hash_pin(pin), stored_hash)


def is_valid_hhmm(value):
    if not re.match(r"^\d{2}:\d{2}$", value or ""):
        return False
    try:
        hour, minute = [int(part) for part in value.split(":", 1)]
    except ValueError:
        return False
    return 0 <= hour <= 23 and minute in [0, 15, 30, 45]


def is_valid_clock_time(value):
    if not re.match(r"^\d{2}:\d{2}$", value or ""):
        return False
    try:
        hour, minute = [int(part) for part in value.split(":", 1)]
    except ValueError:
        return False
    return 0 <= hour <= 23 and 0 <= minute <= 59


def minutes_from_hhmm(value):
    if not is_valid_hhmm(value):
        return None
    hour, minute = [int(part) for part in value.split(":", 1)]
    return hour * 60 + minute


def minutes_from_clock_time(value):
    if not is_valid_clock_time(value):
        return None
    hour, minute = [int(part) for part in value.split(":", 1)]
    return hour * 60 + minute


def hhmm_from_minutes(minutes):
    minutes = max(0, min(23 * 60 + 45, int(minutes)))
    return f"{minutes // 60:02d}:{minutes % 60:02d}"


def load_settings():
    if not os.path.exists(SETTINGS_PATH):
        save_settings(DEFAULT_SETTINGS)
    with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)
    changed = False
    for key, value in DEFAULT_SETTINGS.items():
        if key not in data:
            data[key] = value
            changed = True
    if not data.get("admin_pin_hash"):
        data["admin_pin_hash"] = hash_pin(ADMIN_PIN)
        changed = True
    if changed:
        save_settings(data)
    return data


def save_settings(data):
    with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def parse_decimal(value, default=0.0):
    text = str(value or "").strip().replace("€", "").replace(" ", "").replace(",", ".")
    if not text:
        return default
    try:
        return max(0.0, float(text))
    except ValueError:
        return default


def format_decimal(value, digits=2):
    number = float(value or 0)
    return f"{number:.{digits}f}".replace(".", ",")


def format_money(value):
    return f"{format_decimal(value, 2)} €"


def employee_settings_map():
    return {normalize_text_key(employee["name"]): employee for employee in get_time_employees(active_only=False)}


def employee_payroll_values(employee_name, minutes, employees_by_key=None):
    employees_by_key = employees_by_key if employees_by_key is not None else employee_settings_map()
    employee = employees_by_key.get(normalize_text_key(employee_name), {})
    hours = (minutes or 0) / 60
    max_hours = parse_decimal(employee.get("monthly_hours", 0))
    hourly_wage = parse_decimal(employee.get("hourly_wage", 0))
    overtime = max(0.0, hours - max_hours) if max_hours > 0 else 0.0
    salary = hours * hourly_wage
    return hours, overtime, salary


def normalize_time_employees(raw_employees):
    normalized = []
    seen = set()
    for item in raw_employees or []:
        name = str(item.get("name") if isinstance(item, dict) else item or "").strip()
        if not name:
            continue
        key = normalize_text_key(name)
        if key in seen:
            continue
        seen.add(key)
        active = bool(item.get("active", True)) if isinstance(item, dict) else True
        hourly_wage = parse_decimal(item.get("hourly_wage", 0) if isinstance(item, dict) else 0)
        monthly_hours = parse_decimal(item.get("monthly_hours", 0) if isinstance(item, dict) else 0)
        normalized.append({
            "name": name,
            "active": active,
            "hourly_wage": hourly_wage,
            "monthly_hours": monthly_hours,
        })
    return normalized


def get_time_employees(active_only=True):
    employees = normalize_time_employees(load_settings().get("time_employees", []))
    return [employee for employee in employees if employee.get("active") or not active_only]


def get_time_employee_names(active_only=True):
    return [employee["name"] for employee in get_time_employees(active_only=active_only)]


def save_time_employees(employees):
    settings = load_settings()
    settings["time_employees"] = normalize_time_employees(employees)
    save_settings(settings)


def make_location_id(name):
    base = normalize_text_key(name) or "standort"
    return base[:40] or "standort"


def normalize_min_stock_items(raw_items):
    normalized = []
    seen = set()
    for item in raw_items or []:
        if not isinstance(item, dict):
            continue
        product_id = str(item.get("product_id") or "").strip()
        if not product_id.isdigit() or product_id in seen:
            continue
        try:
            quantity = int(item.get("quantity") or 0)
        except (TypeError, ValueError):
            quantity = 0
        quantity = max(0, min(quantity, MAX_ORDER_QUANTITY))
        if quantity <= 0:
            continue
        seen.add(product_id)
        normalized.append({"product_id": product_id, "quantity": quantity})
    return normalized


def make_cockpit_item_id(prefix="item"):
    return f"{prefix}-{uuid.uuid4().hex[:10]}"


def normalize_cockpit_tasks(raw_tasks):
    normalized = []
    seen = set()
    for item in raw_tasks or []:
        if not isinstance(item, dict):
            continue
        title = str(item.get("title") or "").strip()
        if not title:
            continue
        item_id = str(item.get("id") or "").strip() or make_cockpit_item_id("task")
        if item_id in seen:
            item_id = make_cockpit_item_id("task")
        seen.add(item_id)
        normalized.append({
            "id": item_id,
            "title": title[:180],
            "template": bool(item.get("template")),
            "active": bool(item.get("active", True)),
            "image_filename": os.path.basename(str(item.get("image_filename") or "").strip()),
        })
    return normalized


def normalize_cockpit_orders(raw_orders):
    normalized = []
    seen = set()
    for item in raw_orders or []:
        if not isinstance(item, dict):
            continue
        title = str(item.get("title") or "").strip()
        note = str(item.get("note") or "").strip()
        if not title and not note:
            continue
        item_id = str(item.get("id") or "").strip() or make_cockpit_item_id("order")
        if item_id in seen:
            item_id = make_cockpit_item_id("order")
        seen.add(item_id)
        normalized.append({
            "id": item_id,
            "title": title[:180] or "Bestellung",
            "note": note[:500],
            "active": bool(item.get("active", True)),
            "image_filename": os.path.basename(str(item.get("image_filename") or "").strip()),
        })
    return normalized


def is_valid_iso_date(value):
    try:
        datetime.strptime(value or "", "%Y-%m-%d")
        return True
    except ValueError:
        return False


def normalize_production_jobs(raw_jobs):
    normalized = []
    seen = set()
    for item in raw_jobs or []:
        if not isinstance(item, dict):
            continue
        customer_name = str(item.get("customer_name") or item.get("name") or item.get("title") or "").strip()
        order_text = str(item.get("order_text") or item.get("what") or item.get("title") or "").strip()
        title = customer_name or order_text
        due_date = str(item.get("due_date") or "").strip()
        section = str(item.get("section") or "").strip().lower()
        note = str(item.get("note") or "").strip()
        if not title or section not in PRODUCTION_SECTIONS or not is_valid_iso_date(due_date):
            continue
        item_id = str(item.get("id") or "").strip() or make_cockpit_item_id("production")
        if item_id in seen:
            item_id = make_cockpit_item_id("production")
        seen.add(item_id)
        normalized.append({
            "id": item_id,
            "section": section,
            "title": title[:180],
            "customer_name": customer_name[:180],
            "order_text": order_text[:500],
            "due_date": due_date,
            "note": note[:500],
            "active": bool(item.get("active", True)),
            "source": str(item.get("source") or "admin").strip()[:40] or "admin",
            "image_filename": os.path.basename(str(item.get("image_filename") or "").strip()),
        })
    return sorted(normalized, key=lambda item: (item["due_date"], item["section"], item["title"].lower()))


def normalize_locations(raw_locations):
    normalized = []
    used_ids = set()
    for index, item in enumerate(raw_locations or []):
        if isinstance(item, dict):
            name = str(item.get("name") or "").strip()
            password = str(item.get("password") or "")
            contact_name = str(item.get("contact_name") or "")
            role = str(item.get("role") or item.get("access_role") or "standard").strip()
            time_tracking_enabled = bool(item.get("time_tracking_enabled"))
            time_tracking_max_end = str(item.get("time_tracking_max_end") or "")
            min_stock_enabled = bool(item.get("min_stock_enabled"))
            min_stock_items = normalize_min_stock_items(item.get("min_stock_items", []))
            cockpit_tasks = normalize_cockpit_tasks(item.get("cockpit_tasks", []))
            cockpit_orders = normalize_cockpit_orders(item.get("cockpit_orders", []))
            production_jobs = normalize_production_jobs(item.get("production_jobs", []))
            raw_id = str(item.get("id") or "").strip()
        else:
            name = str(item or "").strip()
            password = ""
            contact_name = ""
            role = "standard"
            time_tracking_enabled = False
            time_tracking_max_end = ""
            min_stock_enabled = False
            min_stock_items = []
            cockpit_tasks = []
            cockpit_orders = []
            production_jobs = []
            raw_id = ""
        if not name:
            continue
        base_id = make_location_id(raw_id or name)
        location_id = base_id
        counter = 2
        while location_id in used_ids:
            location_id = f"{base_id}-{counter}"
            counter += 1
        used_ids.add(location_id)
        normalized.append({
            "id": location_id,
            "name": name,
            "password": password,
            "contact_name": contact_name,
            "role": role if role in ACCESS_ROLE_KEYS else "standard",
            "time_tracking_enabled": time_tracking_enabled,
            "time_tracking_max_end": time_tracking_max_end if is_valid_hhmm(time_tracking_max_end) else "",
            "min_stock_enabled": min_stock_enabled,
            "min_stock_items": min_stock_items,
            "cockpit_tasks": cockpit_tasks,
            "cockpit_orders": cockpit_orders,
            "production_jobs": production_jobs,
        })
    return normalized


def get_locations():
    settings = load_settings()
    locations = normalize_locations(settings.get("locations", []))
    if not locations:
        locations = normalize_locations(DEFAULT_SETTINGS["locations"])
    return locations


def save_locations(locations):
    settings = load_settings()
    settings["locations"] = normalize_locations(locations)
    save_settings(settings)


def find_location(location_id):
    for location in get_locations():
        if location["id"] == location_id:
            return location
    return None


def location_ids():
    return [location["id"] for location in get_locations()]


def location_label(location_id):
    location = find_location(location_id)
    return location["name"] if location else (location_id or "")


def is_production_location(location):
    key = normalize_text_key((location or {}).get("id") or "")
    name = normalize_text_key((location or {}).get("name") or "")
    return key == "produktion" or name == "produktion"


def is_schwarzenbek_location(location):
    key = normalize_text_key((location or {}).get("id") or "")
    name = normalize_text_key((location or {}).get("name") or "")
    variants = {"schwarzenbek", "schwarzenbeck", "schwareznbek", "schwareznbeck"}
    return key in variants or name in variants


def current_workweek(offset=0):
    today = berlin_now().date() + timedelta(days=offset * 7)
    monday = today - timedelta(days=today.weekday())
    return [monday + timedelta(days=offset) for offset in range(5)]


def production_jobs_from_all_locations():
    jobs = []
    for location in get_locations():
        jobs.extend(normalize_production_jobs(location.get("production_jobs", [])))
    return jobs


def find_production_job_owner(job_id):
    job_id = str(job_id or "").strip()
    if not job_id:
        return None
    for location in get_locations():
        if any(job["id"] == job_id for job in normalize_production_jobs(location.get("production_jobs", []))):
            return location
    return None


def location_labels_from_keys(keys):
    ids = location_ids()
    if not keys or ALL_LOCATIONS_KEY in keys:
        return "Alle Standorte"
    known = [location_label(key) for key in keys if key in ids]
    return ", ".join(known) if known else "Alle Standorte"


def db():
    con = sqlite3.connect(DB_PATH, timeout=15)
    con.row_factory = sqlite3.Row
    return con


def table_columns(con, table_name):
    return [row[1] for row in con.execute(f"PRAGMA table_info({table_name})").fetchall()]


def add_column_if_missing(con, table_name, column_name, sql_definition):
    if column_name not in table_columns(con, table_name):
        con.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {sql_definition}")


def category_key(name):
    return normalize_text_key(name or "Allgemein") or "allgemein"


def category_exists_in_products(con, category_name):
    rows = con.execute("SELECT category FROM products WHERE deleted_at IS NULL").fetchall()
    return any(product_has_category(row["category"], category_name) for row in rows)


def sync_category_metadata(con):
    now = datetime.now().isoformat(timespec="seconds")
    rows = con.execute("SELECT id, name, category_key, deleted_at FROM categories ORDER BY id").fetchall()
    seen_live = {}
    for row in rows:
        key = category_key(row["name"])
        if row["category_key"] != key:
            con.execute("UPDATE categories SET category_key=? WHERE id=?", (key, row["id"]))
        if row["deleted_at"]:
            continue
        if key in seen_live:
            con.execute("UPDATE categories SET active=0, deleted_at=? WHERE id=?", (now, row["id"]))
            continue
        seen_live[key] = row["name"]


def ensure_category_in_connection(con, name, active=1):
    clean_name = (name or "Allgemein").strip() or "Allgemein"
    key = category_key(clean_name)
    now = datetime.now().isoformat(timespec="seconds")
    row = con.execute("SELECT * FROM categories WHERE category_key=? ORDER BY deleted_at IS NULL DESC, id LIMIT 1", (key,)).fetchone()
    if row:
        con.execute(
            "UPDATE categories SET name=?, active=?, deleted_at=NULL WHERE id=?",
            (row["name"] or clean_name, active, row["id"]),
        )
        return row["name"] or clean_name
    con.execute(
        "INSERT INTO categories (name, category_key, active, created_at, deleted_at) VALUES (?, ?, ?, ?, NULL)",
        (clean_name, key, active, now),
    )
    return clean_name


def category_was_deleted(con, name):
    row = con.execute("SELECT deleted_at FROM categories WHERE category_key=? ORDER BY id LIMIT 1", (category_key(name),)).fetchone()
    return bool(row and row["deleted_at"])


def init_db():
    con = db()
    cur = con.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            package_size TEXT NOT NULL,
            category TEXT NOT NULL DEFAULT 'Allgemein',
            source TEXT NOT NULL,
            visible_to TEXT NOT NULL DEFAULT '',
            image_filename TEXT,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL
        )
    """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            category_key TEXT,
            active INTEGER NOT NULL DEFAULT 1,
            deleted_at TEXT,
            created_at TEXT NOT NULL
        )
    """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_number TEXT NOT NULL,
            location TEXT NOT NULL,
            ordered_by TEXT NOT NULL,
            buyer_group TEXT,
            note TEXT,
            pdf_filename TEXT NOT NULL,
            pdf_data BLOB,
            order_image_filename TEXT,
            completed_at TEXT,
            created_at TEXT NOT NULL
        )
    """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS order_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER NOT NULL,
            product_id INTEGER,
            product_name TEXT NOT NULL,
            package_size TEXT NOT NULL,
            category TEXT NOT NULL DEFAULT 'Allgemein',
            source TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            completed_at TEXT,
            deleted_at TEXT,
            FOREIGN KEY(order_id) REFERENCES orders(id)
        )
    """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS cart_drafts (
            location_id TEXT PRIMARY KEY,
            cart_json TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
    """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS cockpit_item_states (
            location_id TEXT NOT NULL,
            item_type TEXT NOT NULL,
            item_id TEXT NOT NULL,
            state TEXT NOT NULL,
            completed_by TEXT,
            updated_at TEXT NOT NULL,
            PRIMARY KEY(location_id, item_type, item_id)
        )
    """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS time_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            location_id TEXT NOT NULL,
            location_name TEXT NOT NULL,
            employee_name TEXT NOT NULL,
            work_location TEXT NOT NULL,
            work_date TEXT NOT NULL,
            start_time TEXT NOT NULL,
            end_time TEXT NOT NULL,
            duration_minutes INTEGER NOT NULL,
            note TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT,
            edited INTEGER NOT NULL DEFAULT 0
        )
    """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS time_export_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            export_month TEXT NOT NULL,
            filename TEXT,
            sent_to TEXT,
            status TEXT NOT NULL,
            message TEXT,
            auto INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        )
    """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS push_subscriptions (
            endpoint TEXT PRIMARY KEY,
            subscription_json TEXT NOT NULL,
            enabled INTEGER NOT NULL DEFAULT 1,
            user_agent TEXT,
            owner TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
    """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS pdf_files (
            filename TEXT PRIMARY KEY,
            pdf_data BLOB NOT NULL,
            created_at TEXT NOT NULL
        )
    """
    )

    # Migration für ältere ZIP-Versionen / bestehende lokale Datenbanken
    add_column_if_missing(con, "products", "category", "TEXT NOT NULL DEFAULT 'Allgemein'")
    add_column_if_missing(con, "products", "visible_to", f"TEXT NOT NULL DEFAULT '{DEFAULT_VISIBLE_TO}'")
    add_column_if_missing(con, "products", "deleted_at", "TEXT")
    add_column_if_missing(con, "categories", "category_key", "TEXT")
    add_column_if_missing(con, "categories", "deleted_at", "TEXT")
    add_column_if_missing(con, "orders", "buyer_group", "TEXT")
    add_column_if_missing(con, "orders", "order_image_filename", "TEXT")
    add_column_if_missing(con, "orders", "pdf_data", "BLOB")
    add_column_if_missing(con, "orders", "completed_at", "TEXT")
    add_column_if_missing(con, "order_items", "category", "TEXT NOT NULL DEFAULT 'Allgemein'")
    add_column_if_missing(con, "order_items", "completed_at", "TEXT")
    add_column_if_missing(con, "order_items", "deleted_at", "TEXT")
    add_column_if_missing(con, "cockpit_item_states", "completed_by", "TEXT")
    add_column_if_missing(con, "time_entries", "updated_at", "TEXT")
    add_column_if_missing(con, "time_entries", "edited", "INTEGER NOT NULL DEFAULT 0")

    cur.execute("UPDATE products SET category='Allgemein' WHERE category IS NULL OR TRIM(category)=''")
    cur.execute("UPDATE products SET visible_to=? WHERE visible_to IS NULL OR TRIM(visible_to)=''", (DEFAULT_VISIBLE_TO,))
    cur.execute("UPDATE order_items SET category='Allgemein' WHERE category IS NULL OR TRIM(category)=''")
    sync_category_metadata(con)

    now = datetime.now().isoformat(timespec="seconds")
    for category_name in DEFAULT_CATEGORIES:
        if category_key(category_name) == category_key("Allgemein") or not category_was_deleted(con, category_name):
            ensure_category_in_connection(con, category_name, active=1)
    for row in cur.execute("SELECT DISTINCT category FROM products WHERE deleted_at IS NULL AND category IS NOT NULL AND TRIM(category) != ''").fetchall():
        for category_name in split_categories(row[0]):
            if not category_was_deleted(con, category_name):
                ensure_category_in_connection(con, category_name, active=1)
    sync_category_metadata(con)
    try:
        con.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_categories_live_key ON categories(category_key) WHERE deleted_at IS NULL")
    except sqlite3.IntegrityError:
        sync_category_metadata(con)
        con.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_categories_live_key ON categories(category_key) WHERE deleted_at IS NULL")
    con.commit()
    con.close()


def seed_demo_products():
    con = db()
    count = con.execute("SELECT COUNT(*) FROM products WHERE deleted_at IS NULL").fetchone()[0]
    if count == 0:
        now = datetime.now().isoformat(timespec="seconds")
        demo = [
            ("Kaffeebecher", "Karton à 100 Stück", "Verbrauchsmaterial", "Großhandel A", DEFAULT_VISIBLE_TO, None, now),
            ("Servietten", "Packung à 250 Stück", "Verbrauchsmaterial", "Großhandel A", DEFAULT_VISIBLE_TO, None, now),
            ("Reinigungsmittel", "Kanister à 5 Liter", "Reinigung", "Lieferant B", DEFAULT_VISIBLE_TO, None, now),
        ]
        con.executemany(
            """
            INSERT INTO products (name, package_size, category, source, visible_to, image_filename, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            demo,
        )
        con.commit()
    con.close()


def esc(s):
    return html.escape(str(s or ""))


def pdf_text(value, style):
    return Paragraph(esc(value), style)


def pdf_table_row(values, style):
    return [pdf_text(value, style) for value in values]


def pdf_viewer_href(pdf_filename):
    return "/pdf-viewer?file=" + quote_plus(os.path.basename(pdf_filename or ""))


def pdf_direct_href(pdf_filename):
    return "/orders/" + quote_plus(os.path.basename(pdf_filename or ""))


def pdf_download_href(pdf_filename):
    return "/orders-download/" + quote_plus(os.path.basename(pdf_filename or ""))


def slug_filename(filename):
    filename = os.path.basename(filename or "bild")
    name, ext = os.path.splitext(filename)
    ext = ext.lower() if ext.lower() in [".jpg", ".jpeg", ".png", ".gif", ".webp", ".heic", ".heif"] else ".jpg"
    safe = re.sub(r"[^a-zA-Z0-9_-]+", "-", name).strip("-") or "produktbild"
    return f"{safe}-{uuid.uuid4().hex[:10]}{ext}"


def buyer_label(key):
    return location_label(key)


def visible_keys_from_text(text):
    return [x.strip() for x in (text or "").split(",") if x.strip()]


def product_visible_location_keys(text):
    keys = visible_keys_from_text(text)
    ids = location_ids()
    if NO_LOCATIONS_KEY in keys:
        return []
    if not keys or ALL_LOCATIONS_KEY in keys:
        return ids
    matching = [key for key in keys if key in ids]
    return matching or ids


def visible_labels(text):
    return location_labels_from_keys(visible_keys_from_text(text))


def role_label(role):
    labels = dict(ACCESS_ROLES)
    return labels.get(role or "standard", labels["standard"])


def role_options(selected="standard"):
    return "".join(
        f'<option value="{esc(key)}" {"selected" if key == selected else ""}>{esc(label)}</option>'
        for key, label in ACCESS_ROLES
    )


def location_can_order(location):
    return (location or {}).get("role", "standard") in ["standard", "b2b", "order_only", "manager"]


def location_can_time(location):
    return bool((location or {}).get("time_tracking_enabled")) and (location or {}).get("role", "standard") in ["standard", "time_only", "manager"]


def store_visible_locations(keys):
    ids = location_ids()
    unique = [key for key in ids if key in set(keys)]
    if not unique:
        return NO_LOCATIONS_KEY
    return ALL_LOCATIONS_KEY if set(unique) == set(ids) else ",".join(unique)


def apply_category_visibility_for_location(location_id, categories):
    allowed = {normalize_text_key(category) for category in categories if category}
    all_locations = location_ids()
    con = db()
    rows = con.execute("SELECT id, category, visible_to FROM products WHERE deleted_at IS NULL").fetchall()
    for row in rows:
        visible = product_visible_location_keys(row["visible_to"])
        category_keys = {normalize_text_key(category) for category in product_categories(row)}
        if category_keys & allowed:
            if location_id not in visible:
                visible.append(location_id)
        else:
            visible = [key for key in visible if key != location_id]
        visible = [key for key in all_locations if key in visible]
        con.execute("UPDATE products SET visible_to=? WHERE id=?", (store_visible_locations(visible), row["id"]))
    con.commit()
    con.close()


def time_greeting():
    today = berlin_now().date()
    if (today.month, today.day) in [(1, 1), (12, 25), (12, 26)]:
        return "Einen schönen Feiertag"
    hour = berlin_now().hour
    if hour < 11:
        return "Guten Morgen"
    if hour < 18:
        return "Guten Tag"
    return "Guten Abend"


def farewell_message():
    today = berlin_now().date()
    if today.weekday() >= 5:
        return "Hab ein schönes Wochenende und vielen Dank für deine Unterstützung."
    hour = berlin_now().hour
    if hour >= 18:
        return "Genieß deinen Abend und danke für deine Arbeit."
    if hour < 12:
        return "Danke, dass du heute schon mit angepackt hast."
    return "Danke für deinen Einsatz heute."


def normalize_text_key(text):
    text = (text or "").strip().lower()
    replacements = {
        "ä": "ae", "ö": "oe", "ü": "ue", "ß": "ss",
        "é": "e", "è": "e", "á": "a", "à": "a",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return re.sub(r"[^a-z0-9]+", "", text)


def parse_visible_to(value):
    raw = (value or "").strip()
    if not raw or normalize_text_key(raw) in ["alle", "all", "jeder"]:
        return location_ids()
    parts = [p.strip() for p in re.split(r"[,;|/\n]+", raw) if p.strip()]
    keys = []
    for part in parts:
        norm = normalize_text_key(part)
        for location in get_locations():
            if norm == normalize_text_key(location["id"]) or norm == normalize_text_key(location["name"]):
                keys.append(location["id"])
                break
        else:
            for account in BUYER_ACCOUNTS:
                if norm == normalize_text_key(account["key"]) or norm == normalize_text_key(account["label"]):
                    keys.extend(location_ids())
                    break
    unique = []
    for key in keys:
        if key in location_ids() and key not in unique:
            unique.append(key)
    return unique or location_ids()


def make_buyer_token(key):
    sig = hmac.new(APP_SECRET.encode("utf-8"), key.encode("utf-8"), hashlib.sha256).hexdigest()[:24]
    return f"{key}:{sig}"


def parse_buyer_token(token):
    if not token or ":" not in token:
        return None
    key, sig = token.split(":", 1)
    if not find_location(key):
        return None
    expected = make_buyer_token(key).split(":", 1)[1]
    if hmac.compare_digest(sig, expected):
        return key
    return None


def make_admin_token():
    sig = hmac.new(APP_SECRET.encode("utf-8"), b"admin", hashlib.sha256).hexdigest()[:32]
    return f"admin:{sig}"


def is_valid_admin_token(token):
    return hmac.compare_digest(token or "", make_admin_token())


def get_products(active_only=True, buyer_key=None, sort="name_az", category_filter=""):
    con = db()
    q = "SELECT * FROM products"
    params = []
    conditions = ["deleted_at IS NULL"]
    if active_only:
        conditions.append("active=1")
    if conditions:
        q += " WHERE " + " AND ".join(conditions)
    rows = list(con.execute(q, params).fetchall())
    con.close()

    if buyer_key:
        rows = [p for p in rows if buyer_key in product_visible_location_keys(p["visible_to"])]

    if category_filter:
        rows = [p for p in rows if product_has_category(p, category_filter)]

    def name_key(p):
        return (p["name"] or "").lower()

    if sort == "name_za":
        rows.sort(key=name_key, reverse=True)
    elif sort == "category":
        rows.sort(key=lambda p: (product_categories(p)[0].lower(), name_key(p)))
    elif sort == "package":
        rows.sort(key=lambda p: ((p["package_size"] or "").lower(), name_key(p)))
    elif sort == "newest":
        rows.sort(key=lambda p: p["created_at"] or "", reverse=True)
    else:
        rows.sort(key=name_key)
    return rows


def get_source_names():
    con = db()
    rows = con.execute(
        "SELECT DISTINCT source FROM products WHERE deleted_at IS NULL AND TRIM(source) != '' ORDER BY LOWER(source)"
    ).fetchall()
    con.close()
    return [row["source"] for row in rows if row["source"]]


def get_min_stock_items_for_location(location, buyer_key=None):
    if not location or not location.get("min_stock_enabled"):
        return []
    products_by_id = {
        str(product["id"]): product
        for product in get_products(False, sort="category")
    }
    items = []
    for item in normalize_min_stock_items(location.get("min_stock_items", [])):
        product = products_by_id.get(str(item.get("product_id")))
        if not product:
            continue
        items.append({"product": product, "quantity": int(item.get("quantity") or 0)})
    return items


def get_all_categories(active_only=True, include_deleted=False):
    con = db()
    q = "SELECT * FROM categories"
    conditions = []
    if active_only:
        conditions.append("active=1")
    if not include_deleted:
        conditions.append("deleted_at IS NULL")
    if conditions:
        q += " WHERE " + " AND ".join(conditions)
    q += " ORDER BY LOWER(name)"
    rows = con.execute(q).fetchall()
    con.close()
    return rows


def get_category_names(active_only=True):
    names = [row["name"] for row in get_all_categories(active_only=active_only)]
    if "Allgemein" not in names:
        names.insert(0, "Allgemein")
    return names


def split_categories(value):
    raw = str(value or "").strip()
    if not raw:
        return ["Allgemein"]
    parts = [part.strip() for part in re.split(r"[,;|/\n]+", raw) if part.strip()]
    unique = []
    seen = set()
    for part in parts or ["Allgemein"]:
        key = normalize_text_key(part)
        if key and key not in seen:
            unique.append(part)
            seen.add(key)
    return unique or ["Allgemein"]


def category_text(categories):
    return ", ".join(categories or ["Allgemein"])


def product_categories(product_or_value):
    if isinstance(product_or_value, sqlite3.Row) or isinstance(product_or_value, dict):
        return split_categories(product_or_value["category"] if product_or_value["category"] else "Allgemein")
    return split_categories(product_or_value)


def product_has_category(product_or_value, category):
    wanted = normalize_text_key(category or "Allgemein")
    return any(normalize_text_key(item) == wanted for item in product_categories(product_or_value))


def category_input_name(prefix, category):
    return f"{prefix}_{normalize_text_key(category)}"


def category_checkboxes(name_prefix, selected_categories=None):
    selected_keys = {normalize_text_key(category) for category in (selected_categories or [])}
    return "".join(
        f'<label class="visibility-option"><input type="checkbox" name="{esc(category_input_name(name_prefix, category))}" value="1" {"checked" if normalize_text_key(category) in selected_keys else ""}><span>{esc(category)}</span></label>'
        for category in get_category_names(True)
    )


def form_categories(form, prefix, fallback=None):
    names = get_category_names(True)
    selected = [category for category in names if str(form.get(category_input_name(prefix, category), "")).strip()]
    return selected or (fallback or ["Allgemein"])


def ensure_category(name):
    con = db()
    name = ensure_category_in_connection(con, name, active=1)
    sync_category_metadata(con)
    con.commit()
    con.close()
    return name


def ensure_categories(categories):
    cleaned = []
    for category in categories or ["Allgemein"]:
        ensured = ensure_category(category)
        if normalize_text_key(ensured) not in {normalize_text_key(item) for item in cleaned}:
            cleaned.append(ensured)
    return cleaned or ["Allgemein"]


def get_product(product_id):
    con = db()
    row = con.execute("SELECT * FROM products WHERE id=? AND deleted_at IS NULL", (product_id,)).fetchone()
    con.close()
    return row


def get_categories_for_buyer(buyer_key=None):
    products = get_products(True, buyer_key=buyer_key, sort="category")
    categories = sorted({category for p in products for category in product_categories(p)}, key=lambda x: x.lower())
    return categories



def filter_products_by_search(products, search_text, include_source=False):
    term = (search_text or "").strip().lower()
    if not term:
        return products
    words = [w for w in re.split(r"\s+", term) if w]
    filtered = []
    for p in products:
        values = [p["name"], category_text(product_categories(p)), p["package_size"]]
        if include_source:
            values.extend([p["source"], visible_labels(p["visible_to"])])
        haystack = " ".join(str(v or "") for v in values).lower()
        if all(word in haystack for word in words):
            filtered.append(p)
    return filtered


def normalize_whatsapp_number(number):
    # WhatsApp wa.me erwartet internationale Nummern ohne +, Leerzeichen oder Sonderzeichen.
    return re.sub(r"\D+", "", number or "")


def whatsapp_order_link(order):
    settings = load_settings()
    number = normalize_whatsapp_number(settings.get("whatsapp_number", ""))
    if not number:
        return ""
    if order.get("pdf_filename") and order.get("base_url"):
        pdf_hint = f"PDF: {str(order.get('base_url')).rstrip('/')}/orders/{order.get('pdf_filename')}"
    elif order.get("pdf_filename"):
        pdf_hint = f"PDF im Adminbereich: /orders/{order.get('pdf_filename', '')}"
    else:
        pdf_hint = "PDF ist im Adminbereich abrufbar."
    text = (
        f"Neue interne Bestellung {order.get('order_number', '')}\n"
        f"Standortzugang: {order.get('buyer_group', '')}\n"
        f"Standort: {order.get('location', '')}\n"
        f"Besteller: {order.get('ordered_by', '')}\n"
        f"Datum: {order.get('created_at', '')}\n"
        f"{pdf_hint}"
    )
    return f"https://wa.me/{number}?text={quote_plus(text)}"


HEADER_ALIASES = {
    "name": ["produktname", "produkt", "name", "artikel", "artikelname"],
    "package_size": ["gebindegroesse", "gebindegrosse", "gebinde", "verpackung", "package_size", "packaging", "einheit"],
    "category": ["kategorie", "category", "gruppe", "warengruppe"],
    "source": ["bezugsquelle", "source", "lieferant", "lieferantname", "supplier"],
    "visible_to": ["sichtbarfuer", "sichtbarkeit", "visibleto", "visible_to", "bestellerbereich", "zielgruppe"],
}


def canonical_header(header):
    norm = normalize_text_key(header)
    for key, aliases in HEADER_ALIASES.items():
        if norm in aliases:
            return key
    return norm


def parse_csv_products(uploaded_file):
    raw = uploaded_file.content or b""
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1", errors="replace")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t,")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    return normalize_import_rows(reader)


def parse_xlsx_products(uploaded_file):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("Für Excel-Import bitte zuerst installieren: python -m pip install openpyxl") from exc
    wb = load_workbook(io.BytesIO(uploaded_file.content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(x or "").strip() for x in rows[0]]
    data = []
    for row in rows[1:]:
        if not row or not any(str(x or "").strip() for x in row):
            continue
        data.append({headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))})
    return normalize_import_rows(data)


def normalize_import_rows(rows):
    normalized = []
    for raw_row in rows:
        item = {}
        for header, value in dict(raw_row).items():
            key = canonical_header(header)
            if key in ["name", "package_size", "category", "source", "visible_to"]:
                item[key] = str(value or "").strip()
        if any(item.get(k) for k in ["name", "package_size", "source"]):
            normalized.append(item)
    return normalized


def parse_import_file(uploaded_file):
    filename = (uploaded_file.filename or "").lower()
    if filename.endswith(".xlsx"):
        return parse_xlsx_products(uploaded_file)
    if filename.endswith(".csv"):
        return parse_csv_products(uploaded_file)
    raise RuntimeError("Bitte eine CSV- oder Excel-Datei auswählen.")


def parse_order_created_at(value):
    text = str(value or "").strip()
    if not text:
        return datetime.min
    for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return datetime.min


def order_created_sort_key(order):
    try:
        order_id = int(order["id"])
    except (TypeError, ValueError, KeyError):
        order_id = 0
    return (parse_order_created_at(order["created_at"] if "created_at" in order.keys() else ""), order_id)


def get_orders():
    con = db()
    rows = con.execute("SELECT * FROM orders").fetchall()
    con.close()
    rows = sorted(rows, key=order_created_sort_key, reverse=True)
    return rows


def order_day_key(order):
    dt = parse_order_created_at(order["created_at"] if "created_at" in order.keys() else "")
    if dt == datetime.min:
        return "ohne-datum"
    return dt.strftime("%Y-%m-%d")


def order_day_label(day_key):
    try:
        return datetime.strptime(day_key, "%Y-%m-%d").strftime("%d.%m.%Y")
    except ValueError:
        return "Ohne Datum"


def format_iso_date(value):
    try:
        return datetime.strptime(value or "", "%Y-%m-%d").strftime("%d.%m.%Y")
    except ValueError:
        return value or ""


def order_day_sort_key(day_key):
    try:
        return datetime.strptime(day_key, "%Y-%m-%d")
    except ValueError:
        return datetime.min


def get_open_order_product_ids(location_name):
    return set(get_open_order_product_quantities(location_name).keys())


def get_open_order_product_quantities(location_name):
    con = db()
    rows = con.execute(
        """
        SELECT oi.product_id, SUM(oi.quantity) AS total_quantity
        FROM order_items oi
        JOIN orders o ON o.id = oi.order_id
        WHERE o.completed_at IS NULL
          AND oi.product_id IS NOT NULL
          AND oi.quantity > 0
          AND oi.completed_at IS NULL
          AND oi.deleted_at IS NULL
          AND o.location = ?
        GROUP BY oi.product_id
        """,
        (location_name or "",),
    ).fetchall()
    con.close()
    return {
        str(row["product_id"]): int(row["total_quantity"] or 0)
        for row in rows
        if row["product_id"] is not None and int(row["total_quantity"] or 0) > 0
    }


def set_orders_completed(order_ids, completed=True):
    order_ids = [str(order_id) for order_id in order_ids if str(order_id).isdigit()]
    if not order_ids:
        return 0
    placeholders = ",".join("?" for _ in order_ids)
    completed_at = berlin_now().strftime("%d.%m.%Y %H:%M") if completed else None
    con = db()
    con.execute(f"UPDATE orders SET completed_at=? WHERE id IN ({placeholders})", [completed_at] + order_ids)
    con.commit()
    changed = con.total_changes
    con.close()
    return changed


def set_order_items_completed(item_ids, completed=True):
    item_ids = [str(item_id) for item_id in item_ids if str(item_id).isdigit()]
    if not item_ids:
        return 0
    placeholders = ",".join("?" for _ in item_ids)
    completed_at = berlin_now().strftime("%d.%m.%Y %H:%M") if completed else None
    con = db()
    con.execute(
        f"UPDATE order_items SET completed_at=? WHERE id IN ({placeholders}) AND deleted_at IS NULL",
        [completed_at] + item_ids,
    )
    con.commit()
    changed = con.total_changes
    con.close()
    return changed


def delete_order_items_by_ids(item_ids):
    item_ids = [str(item_id) for item_id in item_ids if str(item_id).isdigit()]
    if not item_ids:
        return 0
    placeholders = ",".join("?" for _ in item_ids)
    con = db()
    con.execute(
        f"UPDATE order_items SET deleted_at=? WHERE id IN ({placeholders}) AND deleted_at IS NULL",
        [berlin_now().isoformat(timespec="seconds")] + item_ids,
    )
    con.commit()
    changed = con.total_changes
    con.close()
    return changed


def get_order_items(order_id):
    con = db()
    rows = con.execute(
        "SELECT * FROM order_items WHERE order_id=? AND deleted_at IS NULL ORDER BY category, product_name",
        (order_id,),
    ).fetchall()
    con.close()
    return rows


def empty_cart_state():
    return {"items": {}, "details": {"ordered_by": "", "note": ""}}


def sanitize_cart_state(raw_state):
    if not isinstance(raw_state, dict):
        return empty_cart_state()
    raw_items = raw_state.get("items")
    if not isinstance(raw_items, dict):
        raw_items = raw_state
    items = {}
    for product_id, qty in raw_items.items():
        product_id = str(product_id or "").strip()
        if not product_id.isdigit():
            continue
        try:
            number = int(qty)
        except (TypeError, ValueError):
            number = 0
        number = max(0, min(number, MAX_ORDER_QUANTITY))
        if number > 0:
            items[product_id] = number
    raw_details = raw_state.get("details") if isinstance(raw_state.get("details"), dict) else {}
    details = {
        "ordered_by": str(raw_details.get("ordered_by", ""))[:120].strip(),
        "note": str(raw_details.get("note", ""))[:2000].strip(),
    }
    return {"items": items, "details": details}


def cart_state_has_content(state):
    details = state.get("details") or {}
    return bool(state.get("items") or details.get("ordered_by") or details.get("note"))


def get_cart_draft(location_id):
    if not location_id:
        return empty_cart_state(), ""
    con = db()
    row = con.execute("SELECT cart_json, updated_at FROM cart_drafts WHERE location_id=?", (location_id,)).fetchone()
    con.close()
    if not row:
        return empty_cart_state(), ""
    try:
        state = sanitize_cart_state(json.loads(row["cart_json"] or "{}"))
    except Exception:
        state = empty_cart_state()
    return state, row["updated_at"] or ""


def save_cart_draft(location_id, state):
    if not location_id:
        return
    state = sanitize_cart_state(state)
    con = db()
    if cart_state_has_content(state):
        updated_at = berlin_now().isoformat(timespec="seconds")
        con.execute(
            """
            INSERT INTO cart_drafts (location_id, cart_json, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(location_id) DO UPDATE SET cart_json=excluded.cart_json, updated_at=excluded.updated_at
            """,
            (location_id, json.dumps(state, ensure_ascii=False), updated_at),
        )
    else:
        con.execute("DELETE FROM cart_drafts WHERE location_id=?", (location_id,))
    con.commit()
    con.close()


def delete_cart_draft(location_id):
    if not location_id:
        return
    con = db()
    con.execute("DELETE FROM cart_drafts WHERE location_id=?", (location_id,))
    con.commit()
    con.close()


def get_cockpit_states(location_id):
    if not location_id:
        return {}
    con = db()
    rows = con.execute(
        "SELECT item_type, item_id, state FROM cockpit_item_states WHERE location_id=?",
        (location_id,),
    ).fetchall()
    con.close()
    return {(row["item_type"], row["item_id"]): row["state"] for row in rows}


def get_cockpit_state_details(location_id):
    if not location_id:
        return {}
    con = db()
    rows = con.execute(
        "SELECT item_type, item_id, state, completed_by, updated_at FROM cockpit_item_states WHERE location_id=?",
        (location_id,),
    ).fetchall()
    con.close()
    return {
        (row["item_type"], row["item_id"]): {
            "state": row["state"],
            "completed_by": row["completed_by"] or "",
            "updated_at": row["updated_at"] or "",
        }
        for row in rows
    }


def cockpit_state_is_from_today(state_info):
    updated_at = (state_info or {}).get("updated_at", "")
    try:
        return datetime.strptime(updated_at, "%d.%m.%Y %H:%M").date() == berlin_now().date()
    except ValueError:
        return False


def set_cockpit_state(location_id, item_type, item_id, state, completed_by=""):
    if not location_id or item_type not in ["task", "order", "production_job"] or not item_id:
        return False
    state = str(state or "").strip()
    valid_states = {
        "task": ["open", "done"],
        "order": ["open", "paid_picked", "picked_invoice"],
        "production_job": ["open", "started", "done"],
    }
    if state not in valid_states[item_type]:
        state = "open"
    con = db()
    if state == "open":
        con.execute(
            "DELETE FROM cockpit_item_states WHERE location_id=? AND item_type=? AND item_id=?",
            (location_id, item_type, item_id),
        )
    else:
        con.execute(
            """
            INSERT INTO cockpit_item_states (location_id, item_type, item_id, state, completed_by, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(location_id, item_type, item_id) DO UPDATE SET state=excluded.state, completed_by=excluded.completed_by, updated_at=excluded.updated_at
            """,
            (location_id, item_type, item_id, state, completed_by, berlin_now().strftime("%d.%m.%Y %H:%M")),
        )
    con.commit()
    con.close()
    return True


def get_order_by_pdf_filename(pdf_filename):
    con = db()
    row = con.execute("SELECT * FROM orders WHERE pdf_filename=?", (pdf_filename,)).fetchone()
    con.close()
    return row


def store_pdf_file(filename, data):
    con = db()
    con.execute(
        """
        INSERT INTO pdf_files (filename, pdf_data, created_at)
        VALUES (?, ?, ?)
        ON CONFLICT(filename) DO UPDATE SET pdf_data=excluded.pdf_data, created_at=excluded.created_at
        """,
        (filename, data, berlin_now().strftime("%d.%m.%Y %H:%M")),
    )
    con.commit()
    con.close()


def get_cached_pdf_file(filename):
    con = db()
    row = con.execute("SELECT pdf_data FROM pdf_files WHERE filename=?", (filename,)).fetchone()
    con.close()
    return bytes(row["pdf_data"]) if row and row["pdf_data"] else None


def get_orders_by_ids(order_ids):
    order_ids = [str(order_id) for order_id in order_ids if str(order_id).isdigit()]
    if not order_ids:
        return []
    placeholders = ",".join("?" for _ in order_ids)
    con = db()
    rows = con.execute(f"SELECT * FROM orders WHERE id IN ({placeholders}) ORDER BY created_at, id", order_ids).fetchall()
    con.close()
    return rows


def delete_orders_by_ids(order_ids):
    order_ids = [str(order_id) for order_id in order_ids if str(order_id).isdigit()]
    if not order_ids:
        return 0
    orders = get_orders_by_ids(order_ids)
    if not orders:
        return 0
    placeholders = ",".join("?" for _ in order_ids)
    con = db()
    try:
        con.execute(f"DELETE FROM order_items WHERE order_id IN ({placeholders})", order_ids)
        con.execute(f"DELETE FROM orders WHERE id IN ({placeholders})", order_ids)
        pdf_names = [order["pdf_filename"] for order in orders if order["pdf_filename"]]
        if pdf_names:
            pdf_placeholders = ",".join("?" for _ in pdf_names)
            con.execute(f"DELETE FROM pdf_files WHERE filename IN ({pdf_placeholders})", pdf_names)
        con.commit()
    finally:
        con.close()

    for order in orders:
        for directory, filename in (
            (ORDER_DIR, order["pdf_filename"]),
            (ORDER_IMAGE_DIR, order["order_image_filename"]),
        ):
            if not filename:
                continue
            path = os.path.join(directory, os.path.basename(filename))
            if os.path.exists(path):
                try:
                    os.remove(path)
                except OSError:
                    pass
    return len(orders)


def today_iso():
    return berlin_now().date().isoformat()


def current_month():
    return berlin_now().strftime("%Y-%m")


def floor_now_to_quarter():
    now = berlin_now()
    return now.hour * 60 + (now.minute // 15) * 15


def time_options(limit_minutes=None):
    if limit_minutes is None:
        limit_minutes = 23 * 60 + 45
    limit_minutes = max(0, min(limit_minutes, 23 * 60 + 45))
    return [hhmm_from_minutes(minutes) for minutes in range(0, limit_minutes + 1, 15)]


def time_options_between(start_minutes, end_minutes):
    start_minutes = max(0, min(start_minutes, 23 * 60 + 45))
    end_minutes = max(0, min(end_minutes, 23 * 60 + 45))
    start_minutes = (start_minutes // 15) * 15
    end_minutes = (end_minutes // 15) * 15
    if start_minutes > end_minutes:
        start_minutes = end_minutes
    return [hhmm_from_minutes(minutes) for minutes in range(start_minutes, end_minutes + 1, 15)]


def recent_end_time_options(limit_minutes, window_minutes=120):
    return time_options_between(limit_minutes - window_minutes, limit_minutes)


def option_html(options, selected=""):
    return "".join(f'<option value="{esc(value)}" {"selected" if value == selected else ""}>{esc(value)}</option>' for value in options)


def location_time_limit_minutes(location):
    limits = [floor_now_to_quarter()]
    max_end = (location or {}).get("time_tracking_max_end", "")
    if is_valid_hhmm(max_end):
        limits.append(minutes_from_hhmm(max_end))
    return min(value for value in limits if value is not None)


def is_time_tracking_enabled(location_id):
    location = find_location(location_id)
    return bool(location and location.get("time_tracking_enabled"))


def format_duration(minutes):
    minutes = int(minutes or 0)
    hours = minutes // 60
    rest = minutes % 60
    return f"{hours}:{rest:02d} Std."


def validate_time_entry(location, employee_name, work_location, start_time, end_time, admin=False):
    employee_name = (employee_name or "").strip()
    work_location = (work_location or "").strip()
    start_time = (start_time or "").strip()
    end_time = (end_time or "").strip()
    if not employee_name:
        return None, "Bitte den Namen des Mitarbeiters eintragen."
    if not work_location:
        return None, "Bitte Einsatzort oder Filiale eintragen."
    if not is_valid_clock_time(start_time):
        return None, "Bitte eine gültige Anfangszeit auswählen."
    if not is_valid_hhmm(end_time):
        return None, "Bitte die Endzeit in 15-Minuten-Schritten auswählen."
    start_minutes = minutes_from_clock_time(start_time)
    end_minutes = minutes_from_hhmm(end_time)
    if end_minutes <= start_minutes:
        return None, "Die Endzeit muss nach der Anfangszeit liegen."
    if not admin:
        if end_minutes > floor_now_to_quarter():
            return None, "Die Endzeit darf nicht in der Zukunft liegen."
        max_end = (location or {}).get("time_tracking_max_end", "")
        if is_valid_hhmm(max_end) and end_minutes > minutes_from_hhmm(max_end):
            return None, f"Die Endzeit darf für diesen Standort nicht nach {max_end} liegen."
        window_start = max(0, location_time_limit_minutes(location) - 120)
        if end_minutes < window_start:
            return None, "Die Endzeit muss im aktuellen 2-Stunden-Fenster liegen."
    return end_minutes - start_minutes, ""


def get_time_entries(month=None, employee="", work_location=""):
    month = (month or current_month()).strip()
    con = db()
    q = "SELECT * FROM time_entries WHERE work_date LIKE ?"
    params = [f"{month}%"]
    if employee:
        q += " AND employee_name LIKE ?"
        params.append(f"%{employee}%")
    if work_location:
        q += " AND work_location LIKE ?"
        params.append(f"%{work_location}%")
    q += " ORDER BY work_date DESC, employee_name COLLATE NOCASE, start_time"
    rows = con.execute(q, params).fetchall()
    con.close()
    return rows


def summarize_time_entries(entries):
    by_employee = {}
    by_location = {}
    matrix = {}
    for entry in entries:
        minutes = entry["duration_minutes"] or 0
        employee = entry["employee_name"] or "Ohne Namen"
        work_location = entry["work_location"] or "Ohne Einsatzort"
        by_employee[employee] = by_employee.get(employee, 0) + minutes
        by_location[work_location] = by_location.get(work_location, 0) + minutes
        matrix.setdefault(employee, {})
        matrix[employee][work_location] = matrix[employee].get(work_location, 0) + minutes
    return by_employee, by_location, matrix


def time_entry_by_id(entry_id):
    con = db()
    row = con.execute("SELECT * FROM time_entries WHERE id=?", (entry_id,)).fetchone()
    con.close()
    return row


def find_duplicate_time_entry(employee_name, work_date, start_time, end_time, exclude_id=None):
    start_minutes = minutes_from_clock_time(start_time)
    end_minutes = minutes_from_hhmm(end_time)
    if start_minutes is None or end_minutes is None:
        return None
    q = """
        SELECT * FROM time_entries
        WHERE work_date=?
          AND employee_name = ? COLLATE NOCASE
    """
    params = [work_date, employee_name]
    if exclude_id:
        q += " AND id<>?"
        params.append(exclude_id)
    con = db()
    rows = con.execute(q, params).fetchall()
    con.close()
    for row in rows:
        existing_start = minutes_from_clock_time(row["start_time"])
        existing_end = minutes_from_hhmm(row["end_time"])
        if existing_start is None or existing_end is None:
            continue
        if existing_start < end_minutes and start_minutes < existing_end:
            return row
    return None


def build_time_export(month):
    entries = get_time_entries(month)
    by_employee, by_location, matrix = summarize_time_entries(entries)
    employees_by_key = employee_settings_map()
    entries_by_employee = {}
    for entry in entries:
        entries_by_employee.setdefault(entry["employee_name"] or "Ohne Namen", []).append(entry)

    filename = f"zeiterfassung_{month}.pdf"
    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=A4, rightMargin=12 * mm, leftMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TimeExportTitle", parent=styles["Title"], fontSize=18, leading=22, spaceAfter=8)
    section_style = ParagraphStyle("TimeExportSection", parent=styles["Heading2"], fontSize=12.5, leading=15, spaceBefore=10, spaceAfter=6)
    small_style = ParagraphStyle("TimeExportSmall", parent=styles["BodyText"], fontSize=7.5, leading=9)
    normal_style = ParagraphStyle("TimeExportNormal", parent=styles["BodyText"], fontSize=9, leading=11)
    story = [
        Paragraph(f"Zeiterfassung {esc(month)}", title_style),
        Paragraph(f"Erstellt am {esc(berlin_now().strftime('%d.%m.%Y %H:%M'))}", normal_style),
        Spacer(1, 4 * mm),
    ]

    def p(value, style=small_style):
        return Paragraph(esc(str(value or "")), style)

    def styled_table(rows, widths):
        table = Table(rows, colWidths=widths, repeatRows=1, hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef3fb")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#24346b")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("LEADING", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d7deea")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fbfcff")]),
        ]))
        return table

    story.append(Paragraph("Summe je Mitarbeiter", section_style))
    employee_rows = [[p("Mitarbeiter"), p("Gesamtstunden"), p("Überstunden"), p("Gehalt")]]
    for name, minutes in sorted(by_employee.items(), key=lambda item: item[0].lower()):
        hours, overtime, salary = employee_payroll_values(name, minutes, employees_by_key)
        employee_rows.append([p(name), p(format_decimal(hours)), p(format_decimal(overtime)), p(format_money(salary))])
    if len(employee_rows) == 1:
        employee_rows.append([p("Keine Daten"), p(""), p(""), p("")])
    story.append(styled_table(employee_rows, [70 * mm, 35 * mm, 35 * mm, 38 * mm]))

    story.append(Paragraph("Summe je Einsatzort", section_style))
    location_rows = [[p("Einsatzort"), p("Stunden")]]
    for name, minutes in sorted(by_location.items(), key=lambda item: item[0].lower()):
        location_rows.append([p(name), p(format_decimal((minutes or 0) / 60))])
    if len(location_rows) == 1:
        location_rows.append([p("Keine Daten"), p("")])
    story.append(styled_table(location_rows, [120 * mm, 58 * mm]))

    story.append(Paragraph("Schichten je Mitarbeiter", section_style))
    for employee, employee_entries in sorted(entries_by_employee.items(), key=lambda item: item[0].lower()):
        employee_minutes = sum(entry["duration_minutes"] or 0 for entry in employee_entries)
        hours, overtime, salary = employee_payroll_values(employee, employee_minutes, employees_by_key)
        shift_rows = [[p("Datum"), p("Einsatzort"), p("Anfang"), p("Ende"), p("Stunden"), p("Vorkommnisse"), p("Bearbeitet")]]
        for entry in sorted(employee_entries, key=lambda item: (item["work_date"], item["start_time"], item["end_time"])):
            shift_rows.append([
                p(entry["work_date"]),
                p(entry["work_location"]),
                p(entry["start_time"]),
                p(entry["end_time"]),
                p(format_decimal((entry["duration_minutes"] or 0) / 60)),
                p(entry["note"] or ""),
                p("ja" if entry["edited"] else "nein"),
            ])
        employee_block = [
            Paragraph(
                f"{esc(employee)} - Gesamtstunden: {esc(format_decimal(hours))} - Überstunden: {esc(format_decimal(overtime))} - Gehalt: {esc(format_money(salary))}",
                section_style,
            ),
            styled_table(shift_rows, [22 * mm, 38 * mm, 18 * mm, 18 * mm, 20 * mm, 47 * mm, 15 * mm]),
        ]
        story.append(KeepTogether(employee_block))
    if not entries_by_employee:
        story.append(Paragraph("Keine Schichten für diesen Monat vorhanden.", normal_style))

    doc.build(story)
    return filename, out.getvalue(), "application/pdf"


def log_time_export(month, filename, sent_to, status, message, auto=False):
    con = db()
    con.execute(
        """
        INSERT INTO time_export_logs (export_month, filename, sent_to, status, message, auto, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (month, filename, sent_to, status, message, 1 if auto else 0, berlin_now().strftime("%d.%m.%Y %H:%M")),
    )
    con.commit()
    con.close()


def latest_time_export_log():
    con = db()
    row = con.execute("SELECT * FROM time_export_logs ORDER BY id DESC LIMIT 1").fetchone()
    con.close()
    return row


def auto_time_export_attempted(month):
    con = db()
    row = con.execute("SELECT id FROM time_export_logs WHERE export_month=? AND auto=1 LIMIT 1", (month,)).fetchone()
    con.close()
    return bool(row)


def send_time_export_email(month, filename, data, content_type):
    settings = load_settings()
    smtp_host = os.environ.get("SMTP_HOST")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    smtp_user = os.environ.get("SMTP_USER")
    smtp_password = os.environ.get("SMTP_PASSWORD")
    recipient = (settings.get("time_export_email") or "info@opapeters").strip()
    sender = os.environ.get("SMTP_FROM", smtp_user or recipient or "")
    if not (smtp_host and smtp_user and smtp_password and recipient and sender):
        raise RuntimeError("SMTP-Daten oder Empfängeradresse sind nicht vollständig konfiguriert.")
    msg = EmailMessage()
    msg["Subject"] = f"Zeiterfassung {month}"
    msg["From"] = sender
    msg["To"] = recipient
    msg.set_content(f"Der Monats-Export der Zeiterfassung für {month} befindet sich im Anhang.")
    if content_type == "application/pdf":
        msg.add_attachment(data, maintype="application", subtype="pdf", filename=filename)
    elif "spreadsheet" in content_type:
        msg.add_attachment(data, maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=filename)
    else:
        msg.add_attachment(data, maintype="text", subtype="csv", filename=filename)
    with smtplib.SMTP(smtp_host, smtp_port) as smtp:
        smtp.starttls()
        smtp.login(smtp_user, smtp_password)
        smtp.send_message(msg)
    return recipient


def read_text_file(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return f.read().strip()
    except OSError:
        return ""


def push_public_pem():
    return (os.environ.get("PUSH_VAPID_PUBLIC_KEY") or read_text_file(os.environ.get("PUSH_VAPID_PUBLIC_KEY_FILE", "/etc/opa-peters/push_public.pem"))).strip()


def push_private_pem():
    return (os.environ.get("PUSH_VAPID_PRIVATE_KEY") or read_text_file(os.environ.get("PUSH_VAPID_PRIVATE_KEY_FILE", "/etc/opa-peters/push_private.pem"))).strip()


def push_private_key_for_webpush():
    configured_file = os.environ.get("PUSH_VAPID_PRIVATE_KEY_FILE", "/etc/opa-peters/push_private.pem")
    if configured_file and os.path.exists(configured_file):
        return configured_file
    return push_private_pem()


def push_contact_email():
    contact = (os.environ.get("PUSH_CONTACT_EMAIL") or "mailto:info@opapeters.de").strip()
    return contact if ":" in contact else f"mailto:{contact}"


def push_public_application_key():
    public_pem = push_public_pem()
    if not public_pem:
        return ""
    try:
        from cryptography.hazmat.primitives import serialization
        key = serialization.load_pem_public_key(public_pem.encode("utf-8"))
        numbers = key.public_numbers()
        raw = b"\x04" + int(numbers.x).to_bytes(32, "big") + int(numbers.y).to_bytes(32, "big")
        return base64.urlsafe_b64encode(raw).decode("ascii").rstrip("=")
    except Exception:
        return ""


def push_is_configured():
    return bool(push_public_application_key() and push_private_pem())


def push_subscription_rows(enabled_only=True):
    con = db()
    q = "SELECT * FROM push_subscriptions"
    if enabled_only:
        q += " WHERE enabled=1"
    rows = con.execute(q).fetchall()
    con.close()
    return rows


def push_subscription_count(enabled_only=True):
    try:
        return len(push_subscription_rows(enabled_only=enabled_only))
    except sqlite3.Error:
        return 0


def save_push_subscription(subscription, user_agent="", owner=""):
    endpoint = str((subscription or {}).get("endpoint") or "").strip()
    keys = subscription.get("keys") if isinstance(subscription, dict) else None
    if not endpoint or not isinstance(keys, dict) or not keys.get("p256dh") or not keys.get("auth"):
        return False
    now = berlin_now().strftime("%d.%m.%Y %H:%M")
    subscription_json = json.dumps(subscription, ensure_ascii=False, separators=(",", ":"))
    con = db()
    con.execute(
        """
        INSERT INTO push_subscriptions (endpoint, subscription_json, enabled, user_agent, owner, created_at, updated_at)
        VALUES (?, ?, 1, ?, ?, ?, ?)
        ON CONFLICT(endpoint) DO UPDATE SET
            subscription_json=excluded.subscription_json,
            enabled=1,
            user_agent=excluded.user_agent,
            owner=excluded.owner,
            updated_at=excluded.updated_at
        """,
        (endpoint, subscription_json, (user_agent or "")[:500], (owner or "")[:120], now, now),
    )
    con.commit()
    con.close()
    return True


def disable_push_subscription(endpoint):
    endpoint = (endpoint or "").strip()
    if not endpoint:
        return False
    con = db()
    cur = con.execute(
        "UPDATE push_subscriptions SET enabled=0, updated_at=? WHERE endpoint=?",
        (berlin_now().strftime("%d.%m.%Y %H:%M"), endpoint),
    )
    con.commit()
    con.close()
    return cur.rowcount > 0


def send_push_notification(payload):
    if not push_is_configured():
        return {"sent": 0, "failed": 0, "disabled": 0, "configured": False}
    try:
        from pywebpush import WebPushException, webpush
    except Exception:
        return {"sent": 0, "failed": 0, "disabled": 0, "configured": False}
    private_key = push_private_key_for_webpush()
    claims = {"sub": push_contact_email()}
    data = json.dumps(payload, ensure_ascii=False)
    sent = failed = disabled = 0
    for row in push_subscription_rows(enabled_only=True):
        try:
            webpush(
                subscription_info=json.loads(row["subscription_json"]),
                data=data,
                vapid_private_key=private_key,
                vapid_claims=claims,
            )
            sent += 1
        except WebPushException as exc:
            status = getattr(getattr(exc, "response", None), "status_code", None)
            if status in (404, 410):
                disable_push_subscription(row["endpoint"])
                disabled += 1
            else:
                failed += 1
        except Exception:
            failed += 1
    return {"sent": sent, "failed": failed, "disabled": disabled, "configured": True}


def send_time_entry_push(employee_name, work_location, work_date, start_time, end_time, duration_minutes, note="", created_by_admin=False):
    payload = {
        "title": "Neue Zeiterfassung",
        "body": f"{employee_name} · {work_location} · {start_time}-{end_time} · {format_decimal((duration_minutes or 0) / 60)} Std.",
        "url": "/admin/time?month=" + quote_plus((work_date or current_month())[:7]),
        "tag": f"time-{normalize_text_key(employee_name)}-{work_date}-{start_time}-{end_time}",
        "icon": "/static/icons/icon-192.png",
        "badge": "/static/icons/icon-192.png",
        "details": {
            "employee": employee_name,
            "location": work_location,
            "date": work_date,
            "start": start_time,
            "end": end_time,
            "duration": format_duration(duration_minutes),
            "note": note or "",
            "source": "Admin" if created_by_admin else "Mitarbeiter",
        },
    }
    return send_push_notification(payload)


def send_push_test_notification():
    return send_push_notification({
        "title": "Push-Test",
        "body": "Push-Benachrichtigung funktioniert auf diesem Gerät.",
        "url": "/admin/settings",
        "tag": "opa-peters-push-test",
        "icon": "/static/icons/icon-192.png",
        "badge": "/static/icons/icon-192.png",
    })


def maybe_run_auto_time_export():
    now = berlin_now()
    last_day = calendar.monthrange(now.year, now.month)[1]
    if now.day != last_day or now.hour < 23 or (now.hour == 23 and now.minute < 55):
        return
    month = now.strftime("%Y-%m")
    if auto_time_export_attempted(month):
        return
    filename, data, content_type = build_time_export(month)
    try:
        recipient = send_time_export_email(month, filename, data, content_type)
        log_time_export(month, filename, recipient, "gesendet", "Automatischer Export wurde versendet.", auto=True)
        settings = load_settings()
        settings["last_auto_time_export_month"] = month
        settings["last_auto_time_export_status"] = f"gesendet am {berlin_now().strftime('%d.%m.%Y %H:%M')}"
        save_settings(settings)
    except Exception as exc:
        log_time_export(month, filename, load_settings().get("time_export_email", ""), "fehlgeschlagen", str(exc), auto=True)
        settings = load_settings()
        settings["last_auto_time_export_month"] = month
        settings["last_auto_time_export_status"] = f"fehlgeschlagen am {berlin_now().strftime('%d.%m.%Y %H:%M')}: {exc}"
        save_settings(settings)


def send_pdf_email_if_configured(order, pdf_filename):
    """Versendet die PDF optional per SMTP, wenn SMTP-Daten und Empfänger-Adresse gesetzt sind."""
    settings = load_settings()
    smtp_host = os.environ.get("SMTP_HOST")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    smtp_user = os.environ.get("SMTP_USER")
    smtp_password = os.environ.get("SMTP_PASSWORD")
    central_email = (settings.get("order_email_to") or os.environ.get("CENTRAL_EMAIL") or "").strip()
    sender = os.environ.get("SMTP_FROM", smtp_user or central_email or "")
    if not (smtp_host and smtp_user and smtp_password and central_email and sender):
        return False

    msg = EmailMessage()
    msg["Subject"] = f"Interne Bestellung {order['order_number']} - {order['location']}"
    msg["From"] = sender
    msg["To"] = central_email
    msg.set_content(
        f"Eine neue interne Bestellung wurde erstellt.\n\n"
        f"Bestellnummer: {order['order_number']}\n"
        f"Standortzugang: {order.get('buyer_group', '')}\n"
        f"Standort: {order['location']}\n"
        f"Besteller: {order['ordered_by']}\n"
        f"Datum: {order['created_at']}\n\n"
        f"Die PDF-Bestellung befindet sich im Anhang."
    )
    with open(os.path.join(ORDER_DIR, pdf_filename), "rb") as f:
        msg.add_attachment(f.read(), maintype="application", subtype="pdf", filename=pdf_filename)
    if order.get("order_image_filename"):
        image_path = os.path.join(ORDER_IMAGE_DIR, order["order_image_filename"])
        if os.path.exists(image_path):
            ext = os.path.splitext(image_path)[1].lower().lstrip(".") or "jpeg"
            subtype = "jpeg" if ext in ["jpg", "jpeg"] else ext
            with open(image_path, "rb") as f:
                msg.add_attachment(f.read(), maintype="image", subtype=subtype, filename=order["order_image_filename"])

    with smtplib.SMTP(smtp_host, smtp_port) as smtp:
        smtp.starttls()
        smtp.login(smtp_user, smtp_password)
        smtp.send_message(msg)
    return True


def read_order_pdf_data(pdf_filename):
    path = os.path.join(ORDER_DIR, os.path.basename(pdf_filename or ""))
    if os.path.exists(path):
        with open(path, "rb") as f:
            return f.read()
    order = get_order_by_pdf_filename(pdf_filename)
    if order and order["pdf_data"]:
        return bytes(order["pdf_data"])
    cached = get_cached_pdf_file(pdf_filename)
    if cached:
        return cached
    if order:
        items = [dict(item) for item in get_order_items(order["id"])]
        order_dict = dict(order)
        regenerated = create_pdf(order_dict, items)
        regenerated_path = os.path.join(ORDER_DIR, regenerated)
        if os.path.exists(regenerated_path):
            with open(regenerated_path, "rb") as f:
                data = f.read()
            con = db()
            con.execute("UPDATE orders SET pdf_data=? WHERE id=?", (data, order["id"]))
            con.commit()
            con.close()
            return data
    return None


def create_pdf(order, items):
    filename = f"bestellung_{order['order_number']}.pdf"
    path = os.path.join(ORDER_DIR, filename)
    doc = SimpleDocTemplate(path, pagesize=A4, rightMargin=14 * mm, leftMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    table_text = ParagraphStyle("A4TableText", parent=styles["BodyText"], fontSize=8.5, leading=10.5, wordWrap="CJK")
    table_head = ParagraphStyle("A4TableHead", parent=table_text, fontName="Helvetica-Bold")
    story = []

    story.append(Paragraph("Interne Bestellung", styles["Title"]))
    story.append(Spacer(1, 6 * mm))
    meta = [
        ["Bestellnummer", order["order_number"]],
        ["Datum / Uhrzeit", order["created_at"]],
        ["Standortzugang", order.get("buyer_group", "")],
        ["Standort", order["location"]],
        ["Besteller", order["ordered_by"]],
    ]
    if order.get("note"):
        meta.append(["Bemerkung", order["note"]])
    if order.get("order_image_filename"):
        meta.append(["Bild zur Bestellung", "im Adminbereich gespeichert"])
    meta_rows = [[pdf_text(label, table_head), pdf_text(value, table_text)] for label, value in meta]
    t = Table(meta_rows, colWidths=[38 * mm, 136 * mm], repeatRows=0, splitByRow=1)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("PADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(t)
    story.append(Spacer(1, 8 * mm))

    grouped = {}
    for item in items:
        grouped.setdefault(item["source"] or "Ohne Bezugsquelle", []).append(item)

    for source in sorted(grouped.keys(), key=lambda x: x.lower()):
        story.append(Paragraph(f"Bezugsquelle: {esc(source)}", styles["Heading2"]))
        data = [pdf_table_row(["Produkt", "Kategorie", "Gebindegröße", "Menge"], table_head)]
        for item in sorted(grouped[source], key=lambda x: ((x.get("category") or "").lower(), x["product_name"].lower())):
            data.append(pdf_table_row([item["product_name"], item.get("category") or "Allgemein", item["package_size"], str(item["quantity"])], table_text))
        table = Table(data, colWidths=[76 * mm, 38 * mm, 46 * mm, 14 * mm], repeatRows=1, splitByRow=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (3, 1), (3, -1), "RIGHT"),
                    ("PADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 7 * mm))

    doc.build(story)
    return filename


def create_combined_order_pdf(orders):
    combined = {}
    order_rows = []
    note_rows = []
    for order in sorted(orders, key=lambda o: ((o["location"] or "").lower(), o["created_at"] or "", o["order_number"] or "")):
        order_rows.append([order["created_at"], order["order_number"], order["location"], order["ordered_by"]])
        if (order["note"] or "").strip():
            note_rows.append([order["created_at"], order["order_number"], order["location"], order["note"]])
        for item in get_order_items(order["id"]):
            key = (
                order["location"] or "Ohne Standort",
                item["source"] or "Ohne Bezugsquelle",
                item["product_name"] or "",
                item["category"] or "Allgemein",
                item["package_size"] or "",
            )
            combined.setdefault(
                key,
                {
                    "location": key[0],
                    "source": key[1],
                    "product_name": key[2],
                    "category": key[3],
                    "package_size": key[4],
                    "quantity": 0,
                },
            )
            combined[key]["quantity"] += int(item["quantity"] or 0)

        number = berlin_now().strftime("%Y%m%d-%H%M%S") + "-" + uuid.uuid4().hex[:4].upper()
    filename = f"gesamtbestellung_{number}.pdf"
    path = os.path.join(ORDER_DIR, filename)
    doc = SimpleDocTemplate(path, pagesize=A4, rightMargin=14 * mm, leftMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    table_text = ParagraphStyle("A4CombinedTableText", parent=styles["BodyText"], fontSize=7.6, leading=9.2, wordWrap="CJK")
    table_head = ParagraphStyle("A4CombinedTableHead", parent=table_text, fontName="Helvetica-Bold")
    note_text = ParagraphStyle("A4CombinedNoteText", parent=table_text, textColor=colors.red, fontName="Helvetica-Bold")
    note_head = ParagraphStyle("A4CombinedNoteHead", parent=note_text, fontName="Helvetica-Bold")
    story = [
        Paragraph("Gesamtbestellung", styles["Title"]),
        Spacer(1, 5 * mm),
        Paragraph(f"Erstellt am: {berlin_now().strftime('%d.%m.%Y %H:%M')}", styles["Normal"]),
        Paragraph(f"Zusammengefasste Bestellungen: {len(orders)}", styles["Normal"]),
        Spacer(1, 7 * mm),
        Paragraph("Ausgewählte Einzelbestellungen", styles["Heading2"]),
    ]
    order_data = [pdf_table_row(["Datum", "Bestellnr.", "Standort", "Besteller"], table_head)] + [pdf_table_row(row, table_text) for row in order_rows]
    order_table = Table(order_data, colWidths=[34 * mm, 40 * mm, 50 * mm, 50 * mm], repeatRows=1, splitByRow=1)
    order_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("PADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(order_table)
    if note_rows:
        story.extend([Spacer(1, 5 * mm), Paragraph("Bemerkungen aus den Einzelbestellungen", styles["Heading2"])])
        note_data = [pdf_table_row(["Datum", "Bestellnr.", "Standort", "Bemerkung"], note_head)] + [pdf_table_row(row, note_text) for row in note_rows]
        note_table = Table(note_data, colWidths=[34 * mm, 38 * mm, 42 * mm, 60 * mm], repeatRows=1, splitByRow=1)
        note_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.mistyrose),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.red),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("PADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(note_table)
    story.extend([Spacer(1, 7 * mm), Paragraph("Addierte Produktmengen", styles["Heading2"])])
    by_location = {}
    for item in combined.values():
        if item["quantity"] > 0:
            by_location.setdefault(item["location"], {}).setdefault(item["source"], []).append(item)
    for location in sorted(by_location.keys(), key=lambda value: value.lower()):
        story.append(Paragraph(f"Standort: {esc(location)}", styles["Heading2"]))
        for source in sorted(by_location[location].keys(), key=lambda value: value.lower()):
            story.append(Paragraph(f"Bezugsquelle: {esc(source)}", styles["Heading3"]))
            item_rows = [
                [item["category"], item["product_name"], item["package_size"], str(item["quantity"])]
                for item in sorted(by_location[location][source], key=lambda x: (x["category"].lower(), x["product_name"].lower(), x["package_size"].lower()))
            ]
            item_data = [pdf_table_row(["Kategorie", "Produkt", "Gebinde", "Menge"], table_head)] + [pdf_table_row(row, table_text) for row in item_rows]
            item_table = Table(item_data, colWidths=[36 * mm, 82 * mm, 42 * mm, 14 * mm], repeatRows=1, splitByRow=1)
            item_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
                ("PADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(item_table)
            story.append(Spacer(1, 5 * mm))
        story.append(Spacer(1, 3 * mm))
    doc.build(story)
    return filename


def page(title, body, admin=False, buyer_key=None):
    if admin:
        nav = ""
    elif buyer_key:
        location = find_location(buyer_key)
        nav_links = ['<a href="/cockpit">Cockpit</a>']
        if location_can_order(location):
            nav_links.append('<a href="/">Shop</a>')
        if location_can_time(location):
            nav_links.append('<a href="/time">Zeiterfassung</a>')
        nav_links.append('<a href="/logout">Logout</a>')
        nav = "".join(nav_links)
    else:
        nav = '<a href="/login">Standort Login</a><a href="/admin/login">Admin</a>'
    subtitle = "" if title == "Besteller Login" else "<p>Opa Peters · internes Bestellsystem</p>"
    push_control = """
<div id="pushControl" class="push-control" hidden>
  <span id="pushStatus">Push nicht aktiv</span>
  <button type="button" id="pushToggle" class="button">Push aktivieren</button>
</div>
""" if admin else ""
    return f"""<!doctype html>
<html lang="de">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{esc(title)}</title>
<meta name="theme-color" content="{THEME_COLOR}">
<meta name="mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-title" content="{APP_SHORT_NAME}">
<link rel="manifest" href="/manifest.json">
<link rel="icon" href="/static/icons/icon-192.png">
<link rel="apple-touch-icon" href="/static/icons/icon-192.png">
<link rel="stylesheet" href="/static/style.css?v={ASSET_VERSION}">
</head>
<body>
<header class="topbar"><div class="brand"><div class="brand-mark"><img src="/static/icons/brand-logo.png" alt="Opa Peters Logo"></div><div><h1>{esc(title)}</h1>{subtitle}</div></div><nav>{nav}</nav></header>
<main>{push_control}{body}</main>
<button id="installApp" class="install-app" hidden>App installieren</button>
<div id="iosInstallHelp" class="install-help" hidden>
  <div class="install-help-card">
    <button type="button" class="install-help-close" aria-label="Schließen">×</button>
    <h2>App zum Home-Bildschirm hinzufügen</h2>
    <p>Auf iPhone und iPad öffnest du unten das Teilen-Symbol und wählst danach <strong>Zum Home-Bildschirm</strong>.</p>
    <p class="muted">Danach erscheint das Bestellsystem wie eine App auf dem Home-Bildschirm.</p>
  </div>
</div>
<script src="/static/app.js?v={ASSET_VERSION}" defer></script>
</body>
</html>""".encode("utf-8")


def admin_menu():
    links = """
        <a class="button" href="/admin">Produkte</a>
        <a class="button" href="/admin/orders">Bestellungen</a>
        <a class="button" href="/admin/time">Zeiterfassung</a>
        <a class="button" href="/admin/employees">Personen</a>
        <a class="button" href="/admin/task-report">Aufgabenreport</a>
        <a class="button" href="/admin/categories">Kategorien</a>
        <a class="button" href="/admin/import">Import</a>
        <a class="button" href="/admin/settings">Einstellungen</a>
        <a class="button" href="/admin/locations">Standorte</a>
        <a class="button" href="/admin/visibility">Sichtbarkeit</a>
        <a class="button logout-button" href="/admin/logout">Admin Logout</a>
    """
    return """
    <details class="admin-mobile-menu">
        <summary>Admin-Menü</summary>
        <div class="admin-mobile-links">
            {links}
        </div>
    </details>
    <section class="admin-menu">
        {links}
    </section>
    """.format(links=links)


class App(BaseHTTPRequestHandler):
    def is_admin(self):
        c = cookies.SimpleCookie(self.headers.get("Cookie"))
        return c.get("admin") and is_valid_admin_token(c["admin"].value)

    def current_buyer_key(self):
        c = cookies.SimpleCookie(self.headers.get("Cookie"))
        morsel = c.get("buyer")
        if not morsel:
            return None
        return parse_buyer_token(morsel.value)

    def base_url(self):
        host = self.headers.get("Host") or f"localhost:{PORT}"
        return f"http://{host}"

    def send_html(self, html_bytes, status=200):
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(html_bytes)

    def send_json(self, payload, status=200):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, Authorization, X-API-Key")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.end_headers()
        self.wfile.write(body)

    def redirect(self, path):
        self.send_response(303)
        self.send_header("Cache-Control", "no-store")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        self.send_header("Location", path)
        self.end_headers()

    def read_form(self):
        content_type = self.headers.get("Content-Type", "")
        length = int(self.headers.get("Content-Length", 0))
        if length > MAX_FORM_BYTES:
            raise RequestTooLarge("Die hochgeladenen Daten sind zu groß.")
        body = self.rfile.read(length)

        if content_type.startswith("multipart/form-data"):
            msg = BytesParser(policy=email_policy).parsebytes(
                b"Content-Type: " + content_type.encode("utf-8") + b"\r\n"
                b"MIME-Version: 1.0\r\n\r\n" + body
            )
            data = {}
            for part in msg.iter_parts():
                disposition = part.get("Content-Disposition", "")
                if "form-data" not in disposition:
                    continue
                name = part.get_param("name", header="Content-Disposition")
                filename = part.get_param("filename", header="Content-Disposition")
                payload = part.get_payload(decode=True) or b""
                if not name:
                    continue
                if filename:
                    data[name] = UploadedFile(filename, payload)
                else:
                    charset = part.get_content_charset() or "utf-8"
                    data[name] = payload.decode(charset, errors="replace")
            return data

        raw = body.decode("utf-8", errors="replace")
        return {k: v[0] if v else "" for k, v in parse_qs(raw, keep_blank_values=True).items()}

    def read_json_payload(self, max_bytes=MAX_CART_DRAFT_BYTES):
        length = int(self.headers.get("Content-Length", 0))
        if length > max_bytes:
            raise RequestTooLarge("Die gespeicherten Warenkorbdaten sind zu groß.")
        if length <= 0:
            return {}
        raw = self.rfile.read(length).decode("utf-8", errors="replace")
        try:
            payload = json.loads(raw)
        except json.JSONDecodeError:
            payload = {}
        return payload if isinstance(payload, dict) else {}

    def form_value(self, form, key, default=""):
        item = form.get(key, default)
        if isinstance(item, UploadedFile):
            return default
        return item if item is not None else default

    def do_OPTIONS(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/production-job":
            self.send_response(204)
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Access-Control-Allow-Headers", "Content-Type, Authorization, X-API-Key")
            self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
            self.end_headers()
            return
        self.send_error(404)

    def show_cart_draft(self):
        buyer_key = self.current_buyer_key()
        location = find_location(buyer_key)
        if not buyer_key or not location_can_order(location):
            return self.send_json({"ok": False, "error": "Nicht angemeldet."}, 401)
        state, updated_at = get_cart_draft(buyer_key)
        return self.send_json({"ok": True, "state": state, "updated_at": updated_at})

    def handle_cart_draft(self):
        buyer_key = self.current_buyer_key()
        location = find_location(buyer_key)
        if not buyer_key or not location_can_order(location):
            return self.send_json({"ok": False, "error": "Nicht angemeldet."}, 401)
        payload = self.read_json_payload()
        if payload.get("action") == "clear":
            delete_cart_draft(buyer_key)
            return self.send_json({"ok": True, "cleared": True})
        state = sanitize_cart_state(payload.get("state") if isinstance(payload.get("state"), dict) else payload)
        save_cart_draft(buyer_key, state)
        return self.send_json({"ok": True})

    def current_push_owner(self):
        if self.is_admin():
            return "Admin"
        return ""

    def show_push_config(self):
        owner = self.current_push_owner()
        if not owner:
            return self.send_json({"ok": False, "error": "Nicht angemeldet."}, 401)
        public_key = push_public_application_key()
        return self.send_json({
            "ok": True,
            "enabled": bool(public_key and push_private_pem()),
            "publicKey": public_key,
        })

    def handle_push_subscribe(self):
        owner = self.current_push_owner()
        if not owner:
            return self.send_json({"ok": False, "error": "Nicht angemeldet."}, 401)
        payload = self.read_json_payload(max_bytes=MAX_PUSH_SUBSCRIPTION_BYTES)
        subscription = payload.get("subscription") if isinstance(payload.get("subscription"), dict) else payload
        if not save_push_subscription(subscription, self.headers.get("User-Agent", ""), owner):
            return self.send_json({"ok": False, "error": "Push-Abo konnte nicht gespeichert werden."}, 400)
        return self.send_json({"ok": True})

    def handle_push_unsubscribe(self):
        owner = self.current_push_owner()
        if not owner:
            return self.send_json({"ok": False, "error": "Nicht angemeldet."}, 401)
        payload = self.read_json_payload(max_bytes=MAX_PUSH_SUBSCRIPTION_BYTES)
        endpoint = str(payload.get("endpoint") or "").strip()
        disable_push_subscription(endpoint)
        return self.send_json({"ok": True})

    def handle_admin_push_test(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        result = send_push_test_notification()
        if not result.get("configured"):
            return self.redirect("/admin/settings?error=" + quote_plus("Push ist auf dem Server noch nicht vollständig eingerichtet."))
        if result.get("sent", 0) > 0:
            return self.redirect("/admin/settings?msg=" + quote_plus(f"Push-Test gesendet an {result.get('sent', 0)} Gerät(e)."))
        return self.redirect("/admin/settings?error=" + quote_plus("Push-Test konnte nicht gesendet werden. Bitte Gerät im Adminbereich neu aktivieren."))

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        maybe_run_auto_time_export()
        if path == "/manifest.json":
            return self.serve_manifest()
        if path == "/service-worker.js":
            return self.serve_file(os.path.join(BASE_DIR, "static", "service-worker.js"), "application/javascript")
        if path.startswith("/static/"):
            return self.serve_file(os.path.join(BASE_DIR, path.lstrip("/")), None)
        if path == "/cart-draft":
            return self.show_cart_draft()
        if path == "/push/config":
            return self.show_push_config()
        if path.startswith("/uploads/"):
            upload_name = os.path.basename(path.replace("/uploads/", "", 1))
            return self.serve_file(os.path.join(UPLOAD_DIR, upload_name), None)
        if path.startswith("/order-images/"):
            if not self.is_admin():
                return self.redirect("/admin/login")
            image_name = os.path.basename(path.replace("/order-images/", "", 1))
            return self.serve_file(os.path.join(ORDER_IMAGE_DIR, image_name), None)
        if path.startswith("/orders-download/"):
            if not (self.is_admin() or self.current_buyer_key()):
                return self.redirect("/login")
            pdf_name = os.path.basename(path.replace("/orders-download/", "", 1))
            pdf_data = read_order_pdf_data(pdf_name)
            if not pdf_data:
                self.send_error(404, "PDF wurde nicht gefunden.")
                return
            self.send_response(200)
            self.send_header("Content-Type", "application/pdf")
            self.send_header("Content-Disposition", f'attachment; filename="{pdf_name}"')
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(pdf_data)
            return
        if path.startswith("/orders/"):
            # PDF darf vom Admin und vom eingeloggten Standort direkt geöffnet/gedruckt werden.
            if not (self.is_admin() or self.current_buyer_key()):
                return self.redirect("/login")
            pdf_name = os.path.basename(path.replace("/orders/", "", 1))
            pdf_data = read_order_pdf_data(pdf_name)
            if not pdf_data:
                self.send_error(404, "PDF wurde nicht gefunden.")
                return
            self.send_response(200)
            self.send_header("Content-Type", "application/pdf")
            self.send_header("Content-Disposition", f'inline; filename="{pdf_name}"')
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(pdf_data)
            return
        if path == "/":
            return self.show_order_form(query=parse_qs(parsed.query))
        if path == "/cockpit":
            return self.show_location_cockpit(query=parse_qs(parsed.query))
        if path == "/choose":
            return self.show_buyer_choice()
        if path == "/time":
            return self.show_time_form(query=parse_qs(parsed.query))
        if path == "/login":
            return self.show_buyer_login()
        if path == "/logout":
            self.send_response(303)
            self.send_header("Set-Cookie", "buyer=; Max-Age=0; HttpOnly; SameSite=Lax; Path=/")
            self.send_header("Location", "/login")
            self.end_headers()
            return
        if path == "/admin/login":
            return self.show_admin_login()
        if path == "/admin/logout":
            self.send_response(303)
            self.send_header("Set-Cookie", "admin=; Max-Age=0; HttpOnly; SameSite=Lax; Path=/")
            self.send_header("Location", "/")
            self.end_headers()
            return
        if path == "/admin":
            return self.show_admin(query=parse_qs(parsed.query))
        if path == "/admin/edit-product":
            return self.show_edit_product(query=parse_qs(parsed.query))
        if path == "/admin/orders":
            return self.show_admin_orders(query=parse_qs(parsed.query))
        if path == "/pdf-viewer":
            return self.show_pdf_viewer(query=parse_qs(parsed.query))
        if path == "/admin/time":
            return self.show_admin_time(query=parse_qs(parsed.query))
        if path == "/admin/employees":
            return self.show_employees(query=parse_qs(parsed.query))
        if path == "/admin/task-report":
            return self.show_task_report(query=parse_qs(parsed.query))
        if path == "/admin/time/export":
            return self.serve_time_export(query=parse_qs(parsed.query))
        if path == "/admin/categories":
            return self.show_categories(query=parse_qs(parsed.query))
        if path == "/admin/import":
            return self.show_import(query=parse_qs(parsed.query))
        if path == "/admin/import-template.csv":
            return self.serve_import_template()
        if path == "/admin/settings":
            return self.show_settings(query=parse_qs(parsed.query))
        if path == "/admin/locations":
            return self.show_locations(query=parse_qs(parsed.query))
        if path == "/admin/visibility":
            return self.show_visibility(query=parse_qs(parsed.query))
        self.send_html(page("Nicht gefunden", "<p>Diese Seite gibt es nicht.</p>", buyer_key=self.current_buyer_key()), 404)


    def serve_manifest(self):
        manifest = {
            "name": APP_NAME,
            "short_name": APP_SHORT_NAME,
            "description": "Internes Bestellsystem für Opa Peters.",
            "start_url": "/",
            "scope": "/",
            "display": "standalone",
            "orientation": "portrait",
            "background_color": BACKGROUND_COLOR,
            "theme_color": THEME_COLOR,
            "icons": [
                {"src": "/static/icons/icon-192.png", "sizes": "192x192", "type": "image/png", "purpose": "any maskable"},
                {"src": "/static/icons/icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any maskable"}
            ]
        }
        data = json.dumps(manifest, ensure_ascii=False).encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "application/manifest+json; charset=utf-8")
        self.send_header("Cache-Control", "no-cache")
        self.end_headers()
        self.wfile.write(data)

    def serve_file(self, filepath, content_type):
        abs_path = os.path.abspath(filepath)
        allowed_roots = [os.path.abspath(BASE_DIR), os.path.abspath(DATA_DIR)]
        if not any(abs_path.startswith(root) for root in allowed_roots) or not os.path.exists(abs_path):
            self.send_error(404)
            return
        if content_type is None:
            ext = os.path.splitext(filepath)[1].lower()
            content_type = {
                ".jpg": "image/jpeg",
                ".jpeg": "image/jpeg",
                ".png": "image/png",
                ".gif": "image/gif",
                ".webp": "image/webp",
                ".heic": "image/heic",
                ".heif": "image/heif",
                ".svg": "image/svg+xml",
                ".js": "application/javascript",
                ".json": "application/json",
                ".css": "text/css",
            }.get(ext, "application/octet-stream")
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("X-Content-Type-Options", "nosniff")
        self.end_headers()
        with open(filepath, "rb") as f:
            self.wfile.write(f.read())

    def serve_import_template(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        content = (
            "Produktname;Kategorie;Gebindegröße;Bezugsquelle\n"
            "Kaffeebecher;Verbrauchsmaterial;Karton à 100 Stück;Großhandel A\n"
            "Servietten;Verbrauchsmaterial;Packung à 250 Stück;Großhandel A\n"
        ).encode("utf-8-sig")
        self.send_response(200)
        self.send_header("Content-Type", "text/csv; charset=utf-8")
        self.send_header("Content-Disposition", "attachment; filename=produkt_import_vorlage.csv")
        self.end_headers()
        self.wfile.write(content)

    def serve_time_export(self, query=None):
        if not self.is_admin():
            return self.redirect("/admin/login")
        query = query or {}
        month = (query.get("month", [current_month()])[0] or current_month()).strip()
        if not re.match(r"^\d{4}-\d{2}$", month):
            month = current_month()
        filename, data, content_type = build_time_export(month)
        log_time_export(month, filename, "", "erstellt", "Manueller Export wurde heruntergeladen.", auto=False)
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Disposition", f"attachment; filename={filename}")
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(data)

    def show_buyer_choice(self):
        buyer_key = self.current_buyer_key()
        if not buyer_key:
            return self.redirect("/login")
        location = find_location(buyer_key)
        if not location:
            return self.redirect("/login")
        order_button = '<a class="button primary" href="/">Shop</a>' if location_can_order(location) else ""
        time_button = '<a class="button primary" href="/time">Zeiterfassung</a>' if location_can_time(location) else ""
        body = f"""
        <section class="box narrow choice-box">
            <h2>{esc(location['name'])}</h2>
            <p class="muted">Rolle: {esc(role_label(location.get('role')))}</p>
            <div class="choice-actions">
                {order_button}
                {time_button}
            </div>
            {'' if order_button or time_button else '<p class="error">Für diese Rolle ist aktuell kein Bereich freigeschaltet.</p>'}
        </section>
        """
        self.send_html(page("Auswahl", body, buyer_key=buyer_key))

    def show_location_cockpit(self, query=None):
        buyer_key = self.current_buyer_key()
        if not buyer_key:
            return self.redirect("/login")
        location = find_location(buyer_key)
        if not location:
            return self.redirect("/login")
        query = query or {}
        msg = (query.get("msg", [""])[0] or "").strip()
        open_location = (query.get("open_location", [""])[0] or "").strip()
        try:
            week_offset = int((query.get("week", ["0"])[0] or "0").strip())
        except ValueError:
            week_offset = 0
        states = get_cockpit_states(buyer_key)
        state_details = get_cockpit_state_details(buyer_key)
        employee_names = get_time_employee_names(True)
        employee_options = "".join(f'<option value="{esc(name)}">{esc(name)}</option>' for name in employee_names)
        tasks = []
        for task in normalize_cockpit_tasks(location.get("cockpit_tasks", [])):
            if not task.get("active"):
                continue
            info = state_details.get(("task", task["id"]), {})
            if info.get("state") == "done" and not cockpit_state_is_from_today(info):
                continue
            tasks.append(task)
        cockpit_orders = [] if (is_production_location(location) or is_schwarzenbek_location(location)) else [item for item in normalize_cockpit_orders(location.get("cockpit_orders", [])) if item.get("active")]
        task_cards = []
        for task in tasks:
            info = state_details.get(("task", task["id"]), {})
            done = info.get("state") == "done"
            completed_by = info.get("completed_by", "")
            image = f'<img class="cockpit-thumb" src="/uploads/{esc(task["image_filename"])}" alt="">' if task.get("image_filename") else ""
            task_cards.append(
                f"""
                <form method="post" action="/cockpit/task" class="cockpit-item cockpit-task-form {'is-done' if done else ''}">
                    <input type="hidden" name="task_id" value="{esc(task['id'])}">
                    <input type="hidden" name="state" value="{'open' if done else 'done'}">
                    <input type="hidden" name="completed_by" value="">
                    <button type="submit" class="cockpit-check" aria-label="Aufgabe abhaken">{'✓' if done else ''}</button>
                    <div>{image}<strong>{esc(task['title'])}</strong>{f'<span>erledigt von {esc(completed_by)}</span>' if done and completed_by else ('<span>erledigt</span>' if done else '<span>offen</span>')}</div>
                </form>
                """
            )
        order_section = ""
        order_cards = []
        if is_schwarzenbek_location(location):
            today_iso = berlin_now().date().strftime("%Y-%m-%d")
            production_states = {}
            for source_location in get_locations():
                if normalize_production_jobs(source_location.get("production_jobs", [])):
                    production_states.update(get_cockpit_states(source_location["id"]))
            today_jobs = [
                job for job in production_jobs_from_all_locations()
                if job.get("active") and job["section"] in ["eisvitrinen", "eistorten"] and job["due_date"] == today_iso
            ]
            for job in sorted(today_jobs, key=lambda item: (item["section"], item["title"].lower())):
                job_state = production_states.get(("production_job", job["id"]), "open")
                done = job_state == "done"
                image = f'<img class="cockpit-thumb" src="/uploads/{esc(job["image_filename"])}" alt="">' if job.get("image_filename") else ""
                order_cards.append(
                    f"""
                    <form method="post" action="/cockpit/production-job" class="cockpit-order production-job {'is-done' if done else ''}">
                        <input type="hidden" name="job_id" value="{esc(job['id'])}">
                        <input type="hidden" name="state" value="{'open' if done else 'done'}">
                        <button type="submit" class="cockpit-check" aria-label="Auftrag abhaken">{'✓' if done else ''}</button>
                        <div>
                            {image}
                            <strong>{esc(PRODUCTION_SECTIONS.get(job['section'], job['section']))}: {esc(job.get('customer_name') or job['title'])}</strong>
                            {f'<p>{esc(job.get("order_text", ""))}</p>' if job.get("order_text") else ''}
                            {f'<p>{esc(job["note"])}</p>' if job.get("note") else ''}
                            <span class="cockpit-state">{'fertiggestellt' if done else ('angefangen' if job_state == 'started' else 'heute')}</span>
                        </div>
                    </form>
                    """
                )
            order_section = f"""
            <section class="box cockpit-panel">
                <div class="section-head"><div><h2>Bestellungen</h2><p class="muted">Heute fällige Eisvitrinen und Eistorten aus der Produktion.</p></div><span class="pill">{len(today_jobs)}</span></div>
                <div class="cockpit-list">{''.join(order_cards) if order_cards else '<p class="muted">Für heute sind keine Eisvitrinen oder Eistorten hinterlegt.</p>'}</div>
            </section>
            """
        elif cockpit_orders:
            for item in cockpit_orders:
                state = states.get(("order", item["id"]), "open")
                image = f'<img class="cockpit-thumb" src="/uploads/{esc(item["image_filename"])}" alt="">' if item.get("image_filename") else ""
                state_label = {
                    "paid_picked": "bezahlt & abgeholt",
                    "picked_invoice": "abgeholt & Rechnungsstellung",
                }.get(state, "offen")
                order_cards.append(
                    f"""
                    <article class="cockpit-order {'is-done' if state != 'open' else ''}">
                        <div>
                            {image}
                            <strong>{esc(item['title'])}</strong>
                            {f'<p>{esc(item["note"])}</p>' if item.get("note") else ''}
                            <span class="cockpit-state">{esc(state_label)}</span>
                        </div>
                        <div class="cockpit-order-actions">
                            <form method="post" action="/cockpit/order-state">
                                <input type="hidden" name="order_id" value="{esc(item['id'])}">
                                <input type="hidden" name="state" value="paid_picked">
                                <button class="button" type="submit">bezahlt & abgeholt</button>
                            </form>
                            <form method="post" action="/cockpit/order-state">
                                <input type="hidden" name="order_id" value="{esc(item['id'])}">
                                <input type="hidden" name="state" value="picked_invoice">
                                <button class="button" type="submit">abgeholt & Rechnungsstellung</button>
                            </form>
                            <form method="post" action="/cockpit/order-state">
                                <input type="hidden" name="order_id" value="{esc(item['id'])}">
                                <input type="hidden" name="state" value="open">
                                <button class="button" type="submit">zurücksetzen</button>
                            </form>
                        </div>
                    </article>
                    """
                )
            order_section = f"""
            <section class="box cockpit-panel">
                <div class="section-head"><div><h2>Bestellungen</h2><p class="muted">Status für hinterlegte Bestellinformationen setzen.</p></div><span class="pill">{len(cockpit_orders)}</span></div>
                <div class="cockpit-list">{''.join(order_cards)}</div>
            </section>
            """
        quick_actions = []
        if location_can_order(location):
            quick_actions.append('<a class="button primary" href="/">Shop öffnen</a>')
        if location_can_time(location):
            quick_actions.append('<a class="button primary" href="/time">Zeiterfassung öffnen</a>')
        production_section = ""
        if is_production_location(location):
            weekday_labels = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag"]
            week_days = current_workweek(week_offset)
            week_start = week_days[0]
            week_end = week_days[-1]
            active_jobs = [job for job in normalize_production_jobs(location.get("production_jobs", [])) if job.get("active")]
            jobs_by_section = {key: [] for key in PRODUCTION_SECTIONS}
            for job in active_jobs:
                jobs_by_section.setdefault(job["section"], []).append(job)

            def production_job_card(job):
                job_info = state_details.get(("production_job", job["id"]), {})
                job_state = job_info.get("state", "open")
                done = job_state == "done"
                started_by = job_info.get("completed_by", "")
                image = f'<img class="cockpit-thumb" src="/uploads/{esc(job["image_filename"])}" alt="">' if job.get("image_filename") else ""
                return f"""
                <article class="production-job {'is-done' if done else ('is-started' if job_state == 'started' else '')}">
                    <div>
                        {image}
                        <strong>{esc(job.get('customer_name') or job['title'])}</strong>
                        {f'<p>{esc(job.get("order_text", ""))}</p>' if job.get("order_text") else ''}
                        {f'<p>{esc(job["note"])}</p>' if job.get("note") else ''}
                        <span>{'fertiggestellt' if done else (f'angefangen von {esc(started_by)}' if job_state == 'started' and started_by else esc(format_iso_date(job['due_date'])))}</span>
                    </div>
                    <div class="production-job-actions">
                        <form method="post" action="/cockpit/production-job" class="production-start-form">
                            <input type="hidden" name="job_id" value="{esc(job['id'])}">
                            <input type="hidden" name="state" value="started">
                            <input type="hidden" name="completed_by" value="">
                            <button class="button production-start-button" type="submit">Angefangen</button>
                        </form>
                        <form method="post" action="/cockpit/production-job">
                            <input type="hidden" name="job_id" value="{esc(job['id'])}">
                            <input type="hidden" name="state" value="done">
                            <button class="primary production-done-button" type="submit">Fertiggestellt</button>
                        </form>
                    </div>
                </article>
                """

            def production_week(section_key):
                week_columns = []
                for day, label in zip(week_days, weekday_labels):
                    iso = day.strftime("%Y-%m-%d")
                    day_jobs = [job for job in jobs_by_section.get(section_key, []) if job["due_date"] == iso]
                    week_columns.append(
                        f"""
                        <div class="production-day">
                            <h4>{label}</h4>
                            <span>{day.strftime('%d.%m.')}</span>
                            <div class="production-day-list">{''.join(production_job_card(job) for job in day_jobs) if day_jobs else '<p class="muted">Keine Aufträge</p>'}</div>
                        </div>
                        """
                    )
                return ''.join(week_columns)

            production_section = f"""
            <section class="box production-cockpit">
                <div class="section-head">
                    <div><h2>Produktion</h2><p class="muted">Woche vom {week_start.strftime('%d.%m.%Y')} bis {week_end.strftime('%d.%m.%Y')}</p></div>
                    <div class="table-actions">
                        <a class="button week-arrow" href="/cockpit?week={week_offset - 1}" aria-label="Vorwoche">‹</a>
                        <a class="button" href="/cockpit">Aktuelle Woche</a>
                        <a class="button week-arrow" href="/cockpit?week={week_offset + 1}" aria-label="Folgewoche">›</a>
                    </div>
                </div>
                <details class="category-panel production-panel">
                    <summary>Backen <span>Wochenansicht</span></summary>
                    <div class="production-week">{production_week('backen')}</div>
                </details>
                <details class="category-panel production-panel">
                    <summary>Eisvitrinen <span>Wochenansicht</span></summary>
                    <div class="production-week">{production_week('eisvitrinen')}</div>
                </details>
                <details class="category-panel production-panel">
                    <summary>Eistorten <span>Wochenansicht</span></summary>
                    <div class="production-week">{production_week('eistorten')}</div>
                </details>
            </section>
            """
        body = f"""
        {f'<div class="success box narrow">{esc(msg)}</div>' if msg else ''}
        <div class="task-complete-modal" id="taskCompleteModal" hidden>
            <div class="task-complete-card">
                <h2>Wer hat die Aufgabe erledigt?</h2>
                <label>Personal<select id="taskCompletePerson" size="6">{employee_options}</select></label>
                <div class="table-actions">
                    <button class="button" type="button" id="taskCompleteCancel">Abbrechen</button>
                    <button class="primary" type="button" id="taskCompleteConfirm">Speichern</button>
                </div>
            </div>
        </div>
        <div class="task-complete-modal" id="productionStartModal" hidden>
            <div class="task-complete-card">
                <h2>Wer beginnt den Auftrag?</h2>
                <label>Name<input id="productionStartPerson" placeholder="Name eingeben"></label>
                <div class="table-actions">
                    <button class="button" type="button" id="productionStartCancel">Abbrechen</button>
                    <button class="primary" type="button" id="productionStartConfirm">Speichern</button>
                </div>
            </div>
        </div>
        <section class="box cockpit-hero">
            <div>
                <h2>Cockpit {esc(location['name'])}</h2>
                <p class="muted">Übersicht für Tagesaufgaben und Bestellinformationen.</p>
            </div>
            <div class="table-actions cockpit-actions">{''.join(quick_actions)}</div>
        </section>
        <section class="cockpit-grid">
            <section class="box cockpit-panel">
                <div class="section-head"><div><h2>Tagesaufgaben</h2><p class="muted">Zum Erledigen antippen.</p></div><span class="pill">{len(tasks)}</span></div>
                <div class="cockpit-list">{''.join(task_cards) if task_cards else '<p class="muted">Für diesen Standort sind aktuell keine Tagesaufgaben hinterlegt.</p>'}</div>
            </section>
            {order_section}
        </section>
        {production_section}
        """
        self.send_html(page("Cockpit", body, buyer_key=buyer_key))

    def show_time_form(self, error="", query=None):
        buyer_key = self.current_buyer_key()
        if not buyer_key:
            return self.redirect("/login")
        location = find_location(buyer_key)
        if not location or not location_can_time(location):
            return self.redirect("/")
        query = query or {}
        msg = (query.get("msg", [""])[0] or "").strip()
        limit_minutes = location_time_limit_minutes(location)
        end_options = '<option value="">Bitte auswählen</option>' + option_html(recent_end_time_options(limit_minutes))
        employee_names = get_time_employee_names(True)
        employee_options = '<option value="">Bitte auswählen</option>' + option_html(employee_names)
        employee_field = (
            f'<label>Mitarbeiter *<select name="employee_name" required>{employee_options}</select></label>'
            if employee_names
            else '<div class="error full">Es sind noch keine Personen für die Zeiterfassung angelegt. Bitte im Adminbereich unter Personen mindestens eine Person anlegen.</div>'
        )
        submit_disabled = "" if employee_names else " disabled"
        order_link = '<a class="button" href="/">Zum Shop</a>' if location_can_order(location) else ""
        body = f"""
        {f'<div class="error">{esc(error)}</div>' if error else ''}
        {f'<div class="success box narrow">{esc(msg)}</div>' if msg else ''}
        <section class="box">
            <div class="section-head">
                <div>
                    <h2>Zeiterfassung</h2>
                    <p class="muted">Standort: {esc(location['name'])} · Datum: {esc(today_iso())}</p>
                </div>
                {order_link}
            </div>
            <form method="post" action="/time" class="two time-form">
                {employee_field}
                <label>Einsatzort / Filiale *<input name="work_location" required value="{esc(location['name'])}" placeholder="z. B. Schwarzenbek"></label>
                <label>Anfangszeit *<input name="start_time" type="time" required></label>
                <label>Endzeit *<select name="end_time" required>{end_options}</select></label>
                <label class="full">Besondere Vorkommnisse<textarea name="note" rows="4" placeholder="Optional"></textarea></label>
                <button class="primary" type="submit"{submit_disabled}>Zeiterfassung speichern</button>
            </form>
            <p class="muted">Endzeiten zeigen nur das aktuelle 2-Stunden-Fenster und bleiben zusätzlich durch Uhrzeit und maximale Standort-Endzeit begrenzt.</p>
        </section>
        """
        self.send_html(page("Zeiterfassung", body, buyer_key=buyer_key))

    def show_admin_time(self, query=None):
        if not self.is_admin():
            return self.redirect("/admin/login")
        query = query or {}
        month = (query.get("month", [current_month()])[0] or current_month()).strip()
        if not re.match(r"^\d{4}-\d{2}$", month):
            month = current_month()
        employee_filter = (query.get("employee", [""])[0] or "").strip()
        location_filter = (query.get("work_location", [""])[0] or "").strip()
        msg = (query.get("msg", [""])[0] or "").strip()
        error = (query.get("error", [""])[0] or "").strip()
        entries = get_time_entries(month, employee_filter, location_filter)
        by_employee, by_location, matrix = summarize_time_entries(entries)
        full_options = option_html(time_options())
        active_employee_names = get_time_employee_names(active_only=True)
        add_employee_options = '<option value="">Bitte auswählen</option>' + option_html(active_employee_names)
        create_disabled = "" if active_employee_names else " disabled"
        location_options = '<option value="">Bitte auswählen</option>' + "".join(
            f'<option value="{esc(location["id"])}">{esc(location["name"])}</option>' for location in get_locations()
        )
        initial_shift_rows = []
        for index in range(3):
            initial_shift_rows.append(
                f"""
                <div class="bulk-time-row" data-bulk-time-row>
                    <label>Datum<input name="shift_date_{index}" type="date" value="{esc(today_iso())}"></label>
                    <label>Anfang<select name="shift_start_{index}"><option value="">Bitte auswählen</option>{full_options}</select></label>
                    <label>Ende<select name="shift_end_{index}"><option value="">Bitte auswählen</option>{full_options}</select></label>
                    <label class="bulk-time-note">Hinweis<textarea name="shift_note_{index}" rows="2" placeholder="Optional"></textarea></label>
                </div>
                """
            )
        bulk_shift_template = esc(
            """
            <div class="bulk-time-row" data-bulk-time-row>
                <label>Datum<input name="shift_date___INDEX__" type="date" value="__TODAY__"></label>
                <label>Anfang<select name="shift_start___INDEX__"><option value="">Bitte auswählen</option>__START_OPTIONS__</select></label>
                <label>Ende<select name="shift_end___INDEX__"><option value="">Bitte auswählen</option>__END_OPTIONS__</select></label>
                <label class="bulk-time-note">Hinweis<textarea name="shift_note___INDEX__" rows="2" placeholder="Optional"></textarea></label>
                <button class="button danger bulk-time-remove" type="button">Zeile entfernen</button>
            </div>
            """.replace("__TODAY__", today_iso()).replace("__START_OPTIONS__", full_options).replace("__END_OPTIONS__", full_options)
        )
        employees_by_key = employee_settings_map()
        summary_employee_rows = []
        for name, minutes in sorted(by_employee.items(), key=lambda item: item[0].lower()):
            hours, overtime, salary = employee_payroll_values(name, minutes, employees_by_key)
            summary_employee_rows.append(
                f"<tr><td>{esc(name)}</td><td>{esc(format_decimal(hours))}</td><td>{esc(format_decimal(overtime))}</td><td>{esc(format_money(salary))}</td></tr>"
            )
        summary_employee = "".join(summary_employee_rows)
        summary_location = "".join(
            f"<tr><td>{esc(name)}</td><td>{esc(format_duration(minutes))}</td></tr>"
            for name, minutes in sorted(by_location.items(), key=lambda item: item[0].lower())
        )
        matrix_rows = []
        for employee, locations in sorted(matrix.items(), key=lambda item: item[0].lower()):
            for work_location, minutes in sorted(locations.items(), key=lambda item: item[0].lower()):
                matrix_rows.append(f"<tr><td>{esc(employee)}</td><td>{esc(work_location)}</td><td>{esc(format_duration(minutes))}</td></tr>")
        entries_by_employee = {}
        employee_names_all = get_time_employee_names(active_only=False)
        for entry in entries:
            end_options = option_html(time_options(), entry["end_time"])
            entry_employee_names = list(employee_names_all)
            if entry["employee_name"] and entry["employee_name"] not in entry_employee_names:
                entry_employee_names.append(entry["employee_name"])
            employee_options = option_html(entry_employee_names, entry["employee_name"])
            status_text = "bearbeitet" if entry["edited"] else "original"
            if entry["updated_at"]:
                status_text += f" · {entry['updated_at']}"
            employee_name = entry["employee_name"] or "Ohne Namen"
            entries_by_employee.setdefault(employee_name, []).append(
                f"""
                <details class="time-entry-card compact-time-entry">
                    <summary class="time-entry-summary">
                        <span class="time-entry-main">
                            <span>{esc(entry['work_date'])} · {esc(entry['start_time'])}-{esc(entry['end_time'])}</span>
                            <strong>{esc(entry['work_location'])}</strong>
                        </span>
                        <span class="time-entry-meta">
                            <strong>{esc(format_duration(entry['duration_minutes']))}</strong>
                        </span>
                    </summary>
                    <div class="time-entry-edit">
                        <p class="entry-meta">{esc(status_text)}</p>
                    <form method="post" action="/admin/time/update" class="time-entry-form">
                        <input type="hidden" name="id" value="{entry['id']}">
                        <label>Datum<input name="work_date" type="date" value="{esc(entry['work_date'])}" required></label>
                        <label>Mitarbeiter<select name="employee_name" required>{employee_options}</select></label>
                        <label>Einsatzort<input name="work_location" value="{esc(entry['work_location'])}" required></label>
                        <label>Anfang<input name="start_time" type="time" value="{esc(entry['start_time'])}" required></label>
                        <label>Ende<select name="end_time">{end_options}</select></label>
                        <label>Dauer<input value="{esc(format_duration(entry['duration_minutes']))}" disabled></label>
                        <label class="full">Vorkommnisse<textarea name="note" rows="3">{esc(entry['note'])}</textarea></label>
                        <button class="primary" type="submit">Eintrag speichern</button>
                    </form>
                    <form method="post" action="/admin/time/delete" data-confirm="Diesen Zeiteintrag wirklich löschen? Diese Aktion kann nicht rückgängig gemacht werden.">
                        <input type="hidden" name="id" value="{entry['id']}">
                        <input type="hidden" name="month" value="{esc(month)}">
                        <button class="danger" type="submit">Zeiteintrag löschen</button>
                    </form>
                    </div>
                </details>
                """
            )
        employee_entry_panels = []
        for employee_name, employee_cards in sorted(entries_by_employee.items(), key=lambda item: item[0].lower()):
            total_minutes = sum(
                int(entry["duration_minutes"] or 0)
                for entry in entries
                if (entry["employee_name"] or "Ohne Namen") == employee_name
            )
            employee_entry_panels.append(
                f"""
                <details class="time-person-panel">
                    <summary>
                        <span><strong>{esc(employee_name)}</strong><small>{len(employee_cards)} Schicht(en)</small></span>
                        <strong>{esc(format_duration(total_minutes))}</strong>
                    </summary>
                    <div class="time-person-shifts">{''.join(employee_cards)}</div>
                </details>
                """
            )
        latest_log = latest_time_export_log()
        log_text = "Noch kein Export-Protokoll vorhanden."
        if latest_log:
            log_text = f"{latest_log['created_at']}: {latest_log['status']} ({latest_log['export_month']}) {latest_log['message'] or ''}"
        settings = load_settings()
        push_configured = push_is_configured()
        push_count = push_subscription_count(enabled_only=True)
        push_status = "eingerichtet" if push_configured else "nicht eingerichtet"
        body = f"""
        {admin_menu()}
        {f'<div class="success box narrow">{esc(msg)}</div>' if msg else ''}
        {f'<div class="error box narrow">{esc(error)}</div>' if error else ''}
        <section class="box">
            <div class="section-head">
                <div><h2>Zeiterfassung</h2><p class="muted">Monatsübersicht, Filter, Korrekturen und Export.</p></div>
                <a class="button primary" href="/admin/time/export?month={esc(month)}">Monat exportieren</a>
            </div>
            <form method="get" action="/admin/time" class="filters compact-form">
                <label>Monat<input type="month" name="month" value="{esc(month)}"></label>
                <label>Mitarbeiter<input name="employee" value="{esc(employee_filter)}" placeholder="Name filtern"></label>
                <label>Einsatzort<input name="work_location" value="{esc(location_filter)}" placeholder="Filiale filtern"></label>
                <button type="submit">Anzeigen</button>
                <a class="button" href="/admin/time">Zurücksetzen</a>
            </form>
        </section>
        <details class="category-panel admin-toggle-panel">
            <summary>Schichten manuell nachtragen <span>Admin</span></summary>
            <form method="post" action="/admin/time/create" class="admin-time-create-form">
                <label>Standort<select name="location_id" required>{location_options}</select></label>
                <label>Mitarbeiter<select name="employee_name" required>{add_employee_options}</select></label>
                <input type="hidden" name="shift_count" id="bulkTimeShiftCount" value="3">
                <div class="bulk-time-rows" id="bulkTimeRows" data-template="{bulk_shift_template}">
                    {''.join(initial_shift_rows)}
                </div>
                <div class="bulk-time-actions">
                    <button class="button" type="button" id="addBulkTimeRow">Weitere Schicht hinzufügen</button>
                    <button class="primary" type="submit"{create_disabled}>Schichten nachtragen</button>
                </div>
            </form>
            {'' if active_employee_names else '<p class="error">Bitte zuerst unter Personen mindestens eine aktive Person anlegen.</p>'}
        </details>
        <section class="stats">
            <div class="stat"><strong>{len(entries)}</strong><span>Einträge</span></div>
            <div class="stat"><strong>{esc(format_duration(sum(entry['duration_minutes'] for entry in entries)))}</strong><span>Gesamtstunden</span></div>
            <div class="stat"><strong>{len(by_employee)}</strong><span>Mitarbeiter</span></div>
        </section>
        <section class="box">
            <h2>Summen je Mitarbeiter</h2>
            <div class="table-wrap"><table><tr><th>Mitarbeiter</th><th>Gesamtstunden</th><th>Überstunden</th><th>Gehalt</th></tr>{summary_employee if summary_employee else '<tr><td colspan="4">Keine Daten.</td></tr>'}</table></div>
        </section>
        <section class="box">
            <h2>Summen je Einsatzort</h2>
            <div class="table-wrap"><table><tr><th>Einsatzort</th><th>Stunden</th></tr>{summary_location if summary_location else '<tr><td colspan="2">Keine Daten.</td></tr>'}</table></div>
        </section>
        <section class="box">
            <h2>Mitarbeiter in Filialen</h2>
            <div class="table-wrap"><table><tr><th>Mitarbeiter</th><th>Einsatzort</th><th>Stunden</th></tr>{''.join(matrix_rows) if matrix_rows else '<tr><td colspan="3">Keine Daten.</td></tr>'}</table></div>
        </section>
        <section class="box">
            <h2>Einträge bearbeiten</h2>
            <div class="time-entry-list person-entry-list">{''.join(employee_entry_panels) if employee_entry_panels else '<p>Keine Zeiteinträge für diesen Filter.</p>'}</div>
        </section>
        <section class="box narrow">
            <h2>Automatischer Monats-Export</h2>
            <p><strong>Empfänger:</strong> {esc(settings.get('time_export_email') or 'nicht gesetzt')}</p>
            <p><strong>Status:</strong> {esc(settings.get('last_auto_time_export_status') or 'Noch kein automatischer Export.')}</p>
            <p class="muted">{esc(log_text)}</p>
        </section>
        """
        self.send_html(page("Zeiterfassung Admin", body, admin=True, buyer_key=self.current_buyer_key()))

    def show_pdf_viewer(self, query=None):
        if not (self.is_admin() or self.current_buyer_key()):
            return self.redirect("/login")
        query = query or {}
        pdf_name = os.path.basename((query.get("file", [""])[0] or "").strip())
        if not pdf_name or not read_order_pdf_data(pdf_name):
            self.send_error(404, "PDF wurde nicht gefunden.")
            return
        close_href = "/admin/orders" if self.is_admin() else "/"
        body = f"""
        <section class="pdf-viewer-shell">
            <div class="pdf-viewer-bar">
                <strong>{esc(pdf_name)}</strong>
                <div class="pdf-viewer-actions">
                    <a class="button" href="{esc(pdf_direct_href(pdf_name))}" target="_blank" rel="noopener">PDF direkt öffnen / drucken</a>
                    <a class="button primary" href="{esc(pdf_download_href(pdf_name))}">PDF herunterladen</a>
                    <a class="pdf-close-button" href="{close_href}" aria-label="PDF schließen">×</a>
                </div>
            </div>
            <iframe class="pdf-frame" src="/orders/{esc(pdf_name)}" title="PDF Vorschau"></iframe>
        </section>
        """
        self.send_html(page("PDF", body, admin=self.is_admin(), buyer_key=self.current_buyer_key()))

    def show_employees(self, query=None):
        if not self.is_admin():
            return self.redirect("/admin/login")
        query = query or {}
        msg = (query.get("msg", [""])[0] or "").strip()
        error = (query.get("error", [""])[0] or "").strip()
        employees = get_time_employees(active_only=False)
        rows = []
        for index, employee in enumerate(employees):
            checked = "checked" if employee.get("active", True) else ""
            status = "aktiv" if employee.get("active", True) else "deaktiviert"
            rows.append(
                f"""
                <tr>
                    <td><input form="employeesForm" type="hidden" name="employee_original_{index}" value="{esc(employee['name'])}"><input form="employeesForm" name="employee_name_{index}" value="{esc(employee['name'])}" required></td>
                    <td><input form="employeesForm" name="employee_hourly_wage_{index}" value="{esc(format_decimal(employee.get('hourly_wage', 0)))}" inputmode="decimal" placeholder="z. B. 13,50"></td>
                    <td><input form="employeesForm" name="employee_monthly_hours_{index}" value="{esc(format_decimal(employee.get('monthly_hours', 0)))}" inputmode="decimal" placeholder="z. B. 80"></td>
                    <td><label class="check"><input form="employeesForm" type="checkbox" name="employee_active_{index}" value="1" {checked}> {status}</label></td>
                    <td><label class="check danger-check"><input form="employeesForm" type="checkbox" name="employee_remove_{index}" value="1"> Entfernen</label></td>
                </tr>
                """
            )
        body = f"""
        {admin_menu()}
        {f'<div class="success box narrow">{esc(msg)}</div>' if msg else ''}
        {f'<div class="error box narrow">{esc(error)}</div>' if error else ''}
        <section class="box">
            <div class="section-head">
                <div>
                    <h2>Personen für Zeiterfassung</h2>
                    <p class="muted">Diese Personen erscheinen im Zeiterfassungsformular als Auswahl. Ein freies Namensfeld gibt es dort nicht mehr.</p>
                </div>
                <a class="button" href="/admin/time">Zur Zeiterfassung</a>
            </div>
            <form id="employeesForm" method="post" action="/admin/employees" data-confirm="Personenliste speichern?">
                <input type="hidden" name="employee_count" value="{len(employees)}">
                <div class="table-wrap">
                    <table>
                        <tr><th>Name</th><th>Stundenlohn</th><th>Reguläre Monatsstunden</th><th>Status</th><th>Entfernen</th></tr>
                        {''.join(rows) if rows else '<tr><td colspan="5">Noch keine Personen angelegt.</td></tr>'}
                    </table>
                </div>
                <fieldset class="visibility-box">
                    <legend>Neue Person hinzufügen</legend>
                    <label>Name<input name="new_employee_name" placeholder="z. B. Lisa"></label>
                    <label>Stundenlohn<input name="new_employee_hourly_wage" inputmode="decimal" placeholder="z. B. 13,50"></label>
                    <label>Reguläre Monatsstunden<input name="new_employee_monthly_hours" inputmode="decimal" placeholder="z. B. 80"></label>
                </fieldset>
                <button class="primary" type="submit">Personen speichern</button>
            </form>
        </section>
        """
        self.send_html(page("Personen", body, admin=True, buyer_key=self.current_buyer_key()))

    def show_task_report(self, query=None):
        if not self.is_admin():
            return self.redirect("/admin/login")
        query = query or {}
        month = (query.get("month", [current_month()])[0] or current_month()).strip()
        if not re.match(r"^\d{4}-\d{2}$", month):
            month = current_month()
        con = db()
        rows = con.execute(
            """
            SELECT location_id, item_id, completed_by, updated_at
            FROM cockpit_item_states
            WHERE item_type='task'
              AND state='done'
              AND completed_by IS NOT NULL
              AND TRIM(completed_by) != ''
            ORDER BY updated_at DESC
            """
        ).fetchall()
        con.close()
        task_lookup = {}
        for location in get_locations():
            for task in normalize_cockpit_tasks(location.get("cockpit_tasks", [])):
                task_lookup[(location["id"], task["id"])] = {"location": location["name"], "title": task["title"]}
        report_rows = []
        counts = {}
        for row in rows:
            try:
                done_at = datetime.strptime(row["updated_at"], "%d.%m.%Y %H:%M")
            except ValueError:
                continue
            if done_at.strftime("%Y-%m") != month:
                continue
            person = row["completed_by"] or "Unbekannt"
            info = task_lookup.get((row["location_id"], row["item_id"]), {"location": row["location_id"], "title": "Aufgabe"})
            counts[person] = counts.get(person, 0) + 1
            report_rows.append(
                f"<tr><td>{esc(done_at.strftime('%d.%m.%Y %H:%M'))}</td><td>{esc(person)}</td><td>{esc(info['location'])}</td><td>{esc(info['title'])}</td></tr>"
            )
        summary_rows = "".join(
            f"<tr><td>{esc(person)}</td><td>{count}</td></tr>"
            for person, count in sorted(counts.items(), key=lambda item: (-item[1], item[0].lower()))
        )
        body = f"""
        {admin_menu()}
        <section class="box">
            <div class="section-head"><div><h2>Aufgabenreport</h2><p class="muted">Monatsübersicht, wer welche Tagesaufgaben erledigt hat.</p></div></div>
            <form method="get" action="/admin/task-report" class="filters compact-form">
                <label>Monat<input type="month" name="month" value="{esc(month)}"></label>
                <button class="primary" type="submit">Anzeigen</button>
            </form>
        </section>
        <section class="box">
            <h2>Summe je Person</h2>
            <div class="table-wrap"><table><tr><th>Person</th><th>Erledigte Aufgaben</th></tr>{summary_rows if summary_rows else '<tr><td colspan="2">Keine erledigten Aufgaben in diesem Monat.</td></tr>'}</table></div>
        </section>
        <section class="box">
            <h2>Einzelne Aufgaben</h2>
            <div class="table-wrap"><table><tr><th>Datum</th><th>Person</th><th>Standort</th><th>Aufgabe</th></tr>{''.join(report_rows) if report_rows else '<tr><td colspan="4">Keine Einträge vorhanden.</td></tr>'}</table></div>
        </section>
        """
        self.send_html(page("Aufgabenreport", body, admin=True, buyer_key=self.current_buyer_key()))

    def show_buyer_login(self, error=""):
        options = "".join(f'<option value="{esc(location["id"])}">{esc(location["name"])} · {esc(role_label(location.get("role")))}</option>' for location in get_locations())
        body = f"""
        {f'<div class="error">{esc(error)}</div>' if error else ''}
        <form class="box narrow" method="post" action="/login">
            <h2>{esc(time_greeting())}</h2>
            <p class="muted">Bitte wähle deinen Standort und melde dich an.</p>
            <label>Standort *<select name="buyer_key" required><option value="">Bitte auswählen</option>{options}</select></label>
            <label>Standort-Passwort
                <span class="password-wrap"><input class="numeric-pin password-field" type="password" name="pin" inputmode="numeric" placeholder="PIN eingeben"><button type="button" class="password-toggle" aria-label="Passwort anzeigen">Anzeigen</button></span>
            </label>
            <button class="primary" type="submit">Einloggen</button>
        </form>"""
        self.send_html(page("Besteller Login", body))

    def show_order_form(self, error="", query=None):
        buyer_key = self.current_buyer_key()
        if not buyer_key:
            return self.redirect("/login")
        location_data = find_location(buyer_key)
        if not location_can_order(location_data):
            return self.redirect("/time" if location_can_time(location_data) else "/choose")

        query = query or {}
        sort = (query.get("sort", ["name_az"])[0] or "name_az").strip()
        category_filter = (query.get("category", [""])[0] or "").strip()
        search_text = (query.get("q", [""])[0] or "").strip()
        all_order_products = get_products(True, buyer_key=buyer_key, sort=sort)
        products = get_products(True, buyer_key=buyer_key, sort=sort, category_filter=category_filter)
        products = filter_products_by_search(products, search_text, include_source=False)
        categories = get_categories_for_buyer(buyer_key)
        location_name = location_data["name"] if location_data else buyer_label(buyer_key)
        open_order_product_quantities = get_open_order_product_quantities(location_name)
        contact_name = (location_data or {}).get("contact_name", "").strip()
        role_badge = f"<p class='role-badge'>{esc(role_label((location_data or {}).get('role')))}</p>"
        personal_greeting = f"<p class='personal-greeting'>Moin Moin, {esc(contact_name)}</p>" if contact_name else ""
        time_button = '<a class="button primary" href="/time">Zur Zeiterfassung</a>' if location_can_time(location_data) else ""
        logout_button = '<a class="button logout-button" href="/logout">Logout</a>'
        min_stock_items = get_min_stock_items_for_location(location_data, buyer_key=buyer_key)
        min_stock_html = ""
        if min_stock_items:
            min_stock_cards = "".join(
                f"""
                <article class="min-stock-info-card">
                    <strong>{esc(item['product']['name'])}</strong>
                    <span>{esc(category_text(product_categories(item['product'])))}</span>
                    <small>{esc(item['product']['package_size'])} · {esc(item['product']['source'])}</small>
                    <b>{esc(item['quantity'])} mindestens bestellen</b>
                    <button class="button min-stock-cart-button" type="button" data-product-id="{item['product']['id']}" data-min-qty="{esc(item['quantity'])}">In den Warenkorb</button>
                </article>
                """
                for item in min_stock_items
            )
            min_stock_html = f"""
            <details class="category-panel min-stock-info">
                <summary>Mindestbestellmengen <span>{len(min_stock_items)} Produkt(e)</span></summary>
                <p class="muted">Orientierung für {esc(location_name)}.</p>
                <div class="min-stock-info-grid">{min_stock_cards}</div>
            </details>
            """
        cat_options = '<option value="">Alle Kategorien</option>' + "".join(
            f'<option value="{esc(c)}" {"selected" if c == category_filter else ""}>{esc(c)}</option>' for c in categories
        )
        sort_options = [
            ("name_az", "Name A–Z"),
            ("name_za", "Name Z–A"),
            ("category", "Kategorie + Name"),
            ("package", "Gebindegröße"),
            ("newest", "Neueste zuerst"),
        ]
        sort_options_html = "".join(
            f'<option value="{esc(value)}" {"selected" if value == sort else ""}>{esc(label)}</option>' for value, label in sort_options
        )
        def product_card(p):
            img = f'<img src="/uploads/{esc(p["image_filename"])}" alt="">' if p["image_filename"] else '<div class="noimg">kein Bild</div>'
            categories_label = category_text(product_categories(p))
            open_quantity = open_order_product_quantities.get(str(p["id"]), 0)
            already_ordered = open_quantity > 0
            status_badge = f'<div class="ordered-open-badge">Bereits offen bestellt: {open_quantity}</div>' if already_ordered else ""
            card_class = "card product-card ordered-open-product" if already_ordered else "card product-card"
            return f"""
            <article class="{card_class}" data-product-id="{p['id']}" data-product-name="{esc(p['name'])}" data-product-package="{esc(p['package_size'])}">
                {img}
                <div class="category-badge">{esc(categories_label)}</div>
                {status_badge}
                <h3>{esc(p['name'])}</h3>
                <p class="muted">Gebindegröße: {esc(p['package_size'])}</p>
                <div class="quantity-control" aria-label="Menge für {esc(p['name'])}">
                    <button type="button" class="qty-minus" data-product-id="{p['id']}">−</button>
                    <input class="qty-display" type="number" data-product-id="{p['id']}" min="0" max="{MAX_ORDER_QUANTITY}" step="1" value="0" inputmode="numeric" aria-label="Menge">
                    <button type="button" class="qty-plus" data-product-id="{p['id']}">+</button>
                </div>
            </article>"""
        hidden_products = {str(p["id"]): p for p in all_order_products}
        for item in min_stock_items:
            hidden_products[str(item["product"]["id"])] = item["product"]
        hidden_qty_inputs = "".join(
            f'<input type="hidden" name="qty_{p["id"]}" id="qty_{p["id"]}" value="0" data-product-id="{p["id"]}" data-product-name="{esc(p["name"])}" data-product-package="{esc(p["package_size"])}">'
            for p in hidden_products.values()
        )
        category_panels = []
        all_cards = "".join(product_card(p) for p in products)
        if products:
            category_panels.append(f"<details class='category-panel'><summary>Alle Produkte <span>{len(products)}</span></summary><section class='grid'>{all_cards}</section></details>")
            for category in categories:
                category_products = [p for p in products if product_has_category(p, category)]
                if category_products:
                    category_panels.append(
                        f"<details class='category-panel'><summary>{esc(category)} <span>{len(category_products)}</span></summary><section class='grid'>{''.join(product_card(p) for p in category_products)}</section></details>"
                    )
        body = f"""
        {f'<div class="error">{esc(error)}</div>' if error else ''}
        <section class="box">
            <div class="section-head"><div><h2>Angemeldet als: {esc(location_name)}</h2>{role_badge}{personal_greeting}</div><div class="table-actions order-actions">{time_button}<a class="button" href="/logout">Standort wechseln</a>{logout_button}</div></div>
            <form method="get" action="/" class="filters compact-form order-search">
                <input type="hidden" name="sort" value="name_az">
                <label class="search-main"><span>Produkt suchen</span><input name="q" value="{esc(search_text)}" placeholder="z. B. Becher, Servietten, Reinigung"></label>
                <label class="search-category"><span>Kategorie</span><select name="category">{cat_options}</select></label>
                <button class="primary" type="submit">Suchen</button>
                <a class="button" href="/">Reset</a>
            </form>
            {min_stock_html}
            <p class="muted">{len(products)} Produkt(e) gefunden.</p>
        </section>
        <form method="post" action="/order" enctype="multipart/form-data">
            {hidden_qty_inputs}
            <input type="hidden" name="location" value="{esc(buyer_key)}">
            <section class="category-sections">{''.join(category_panels) if category_panels else '<section class="box"><p>Keine passenden Produkte gefunden oder für diesen Standort sind noch keine Produkte sichtbar.</p></section>'}</section>
            <section id="cartReview" class="cart-review" hidden>
                <div class="cart-card">
                    <h2>Warenkorb & Bestelldaten</h2>
                    <p class="muted">Standort: {esc(location_name)}. Bitte prüfe deine Bestellung und ergänze die Bestelldaten.</p>
                    <div id="cartItems"></div>
                    <section class="cart-details">
                        <h3>Bestelldaten</h3>
                        <label>Name des Bestellers *<input name="ordered_by" required placeholder="Name eingeben"></label>
                        <label>Bemerkung / Freitext<textarea name="note" rows="3" placeholder="Optionale Hinweise"></textarea></label>
                        <fieldset class="visibility-box">
                            <legend>Bild zur Bestellung</legend>
                            <label>Bild zur Bestellung hinzufügen<input type="file" name="order_image" accept="image/*"></label>
                            <p class="muted">Optional, zum Beispiel für Notizen, Schäden oder Lagerbestand. Maximal 10 MB.</p>
                        </fieldset>
                    </section>
                    <div class="cart-actions">
                        <button type="button" id="cartCancel">Weiter bearbeiten</button>
                        <button class="primary" type="button" id="cartSubmit">Bestellung endgültig absenden</button>
                    </div>
                </div>
            </section>
            <button id="reviewOrder" class="primary" type="button"><span>Warenkorb prüfen</span><span id="orderCount" class="order-count">0 Positionen</span></button>
        </form>"""
        self.send_html(page("Interne Warenbestellung", body, buyer_key=buyer_key))

    def show_admin_login(self, error=""):
        body = f"""
        {f'<div class="error">{esc(error)}</div>' if error else ''}
        <form class="box narrow" method="post" action="/admin/login">
            <label>Admin-PIN
                <span class="password-wrap"><input class="numeric-pin password-field" type="password" name="pin" inputmode="numeric" required><button type="button" class="password-toggle" aria-label="Passwort anzeigen">Anzeigen</button></span>
            </label>
            <button class="primary" type="submit">Einloggen</button>
        </form>"""
        self.send_html(page("Admin Login", body, buyer_key=self.current_buyer_key()))

    def show_admin(self, query=None):
        if not self.is_admin():
            return self.redirect("/admin/login")
        query = query or {}
        search_text = (query.get("q", [""])[0] or "").strip()
        msg = (query.get("msg", [""])[0] or "").strip()
        error = (query.get("error", [""])[0] or "").strip()
        all_products = get_products(False)
        products = filter_products_by_search(all_products, search_text, include_source=True)
        source_options = "".join(f'<option value="{esc(source)}"></option>' for source in get_source_names())
        orders = get_orders()
        active_count = sum(1 for p in all_products if p["active"])
        inactive_count = len(all_products) - active_count
        rows = []
        rows_by_category = {}
        for p in products:
            img = f'<img class="thumb" src="/uploads/{esc(p["image_filename"])}">' if p["image_filename"] else "—"
            status = "aktiv" if p["active"] else "deaktiviert"
            visibility = visible_labels(p["visible_to"])
            actions = (
                f"<div class='table-actions'>"
                f"<a class='button' href='/admin/edit-product?id={p['id']}'>Bearbeiten</a>"
                f"<form method='post' action='/admin/delete-product' data-confirm='Produkt wirklich löschen? Diese Aktion entfernt es aus dem Bestellmenü.'>"
                f"<input type='hidden' name='id' value='{p['id']}'><button class='danger'>Löschen</button></form>"
                f"</div>"
            )
            categories_label = category_text(product_categories(p))
            category_keys = " ".join(normalize_text_key(category) for category in product_categories(p))
            row_html = f"<tr class='product-row' data-category-keys='{esc(category_keys)}'><td class='select-cell'><label class='bulk-select-label'><input form='bulkProductForm' class='bulk-product-check' type='checkbox' name='selected_{p['id']}' value='1' aria-label='Produkt auswählen'><span class='bulk-checkmark'>✓</span></label> {img}</td><td>{esc(p['name'])}</td><td>{esc(categories_label)}</td><td>{esc(p['package_size'])}</td><td>{esc(p['source'])}</td><td>{esc(visibility)}</td><td>{status}</td><td>{actions}</td></tr>"
            rows.append(row_html)
            for category in product_categories(p):
                rows_by_category.setdefault(category, []).append(row_html)
        order_rows = []
        for o in orders:
            o_dict = dict(o)
            o_dict["base_url"] = self.base_url()
            wa_link = whatsapp_order_link(o_dict)
            wa_cell = f"<a target='_blank' rel='noopener' href='{esc(wa_link)}'>WhatsApp</a>" if wa_link else "—"
            image_cell = f"<a href='/order-images/{esc(o['order_image_filename'])}' target='_blank' rel='noopener'>Bild öffnen</a>" if o["order_image_filename"] else "—"
            delete_cell = (
                f"<form method='post' action='/admin/orders/delete' data-confirm='Diese Bestellung wirklich löschen?'>"
                f"<input type='hidden' name='order_{o['id']}' value='1'>"
                f"<input type='hidden' name='return_to' value='/admin'>"
                f"<button class='danger' type='submit'>Löschen</button></form>"
            )
            order_rows.append(
                f"<tr><td>{esc(o['created_at'])}</td><td>{esc(o['order_number'])}</td><td>{esc(o['buyer_group'] or '')}</td><td>{esc(o['location'])}</td><td>{esc(o['ordered_by'])}</td><td><a href='{esc(pdf_viewer_href(o['pdf_filename']))}'>PDF öffnen</a></td><td>{image_cell}</td><td>{wa_cell}</td><td>{delete_cell}</td></tr>"
            )
        visibility_checks = "".join(
            f'<label class="visibility-option"><input type="checkbox" name="visible_{esc(location["id"])}" value="1" checked><span>{esc(location["name"])}</span></label>'
            for location in get_locations()
        )
        category_checks = category_checkboxes("category", ["Allgemein"])
        bulk_category_checks = category_checkboxes("bulk_category")
        bulk_category_select_buttons = "".join(
            f'<button class="button bulk-category-select" type="button" data-category-key="{esc(normalize_text_key(category))}">{esc(category)}</button>'
            for category in get_category_names(True)
        )
        bulk_visibility_checks = "".join(
            f'<label class="visibility-option"><input type="checkbox" name="bulk_visible_{esc(location["id"])}" value="1"><span>{esc(location["name"])}</span></label>'
            for location in get_locations()
        )
        product_table_header = "<tr><th>Auswahl / Bild</th><th>Name</th><th>Kategorie</th><th>Gebinde</th><th>Bezugsquelle</th><th>Sichtbar für</th><th>Status</th><th></th></tr>"
        def product_table(row_html):
            product_rows = "".join(row_html) if row_html else '<tr><td colspan="8">Keine passenden Produkte gefunden.</td></tr>'
            return f"<div class='table-wrap'><table>{product_table_header}{product_rows}</table></div>"
        admin_category_panels = []
        if rows:
            admin_category_panels.append(f"<details class='category-panel admin-product-panel'><summary>Alle Produkte <span>{len(rows)}</span></summary>{product_table(rows)}</details>")
            for category in sorted(rows_by_category.keys(), key=lambda item: item.lower()):
                category_rows = rows_by_category[category]
                admin_category_panels.append(f"<details class='category-panel admin-product-panel'><summary>{esc(category)} <span>{len(category_rows)}</span></summary>{product_table(category_rows)}</details>")
        body = f"""
        {admin_menu()}
        {f'<div class="success box narrow">{esc(msg)}</div>' if msg else ''}
        {f'<div class="error box narrow">{esc(error)}</div>' if error else ''}
        <section class="stats">
            <div class="stat"><strong>{active_count}</strong><span>aktive Produkte</span></div>
            <div class="stat"><strong>{inactive_count}</strong><span>deaktiviert</span></div>
            <div class="stat"><strong>{len(orders)}</strong><span>letzte Bestellungen</span></div>
        </section>
        <details class="category-panel admin-toggle-panel">
            <summary>Produkte anlegen <span>+</span></summary>
            <form method="post" action="/admin/add-product" enctype="multipart/form-data" class="two">
                <datalist id="sourceSuggestions">{source_options}</datalist>
                <p class="muted full">Kategorien legst du separat an und wählst sie hier aus.</p>
                <label>Produktname *<input name="name" required></label>
                <fieldset class="visibility-box full">
                    <legend>Kategorien *</legend>
                    <div class="visibility-grid">{category_checks}</div>
                </fieldset>
                <label>Gebindegröße *<input name="package_size" required placeholder="z. B. Karton à 12 Stück"></label>
                <label>Bezugsquelle *<input name="source" list="sourceSuggestions" required placeholder="Vorhandene wählen oder neue eintragen"></label>
                <label>Produktbild<input type="file" name="image" accept="image/*"></label>
                <fieldset class="visibility-box">
                    <legend>Sichtbar für Standorte *</legend>
                    <div class="visibility-grid">{visibility_checks}</div>
                </fieldset>
                <button class="primary" type="submit">Produkt speichern</button>
                <a class="button" href="/admin/categories">Kategorien verwalten</a>
            </form>
            <details class="category-panel bulk-add-panel">
                <summary>Mehrere Produkte gleichzeitig anlegen <span>+</span></summary>
                <form method="post" action="/admin/add-product" class="bulk-add-products">
                    <input type="hidden" id="bulkAddProductCount" name="bulk_add_count" value="3">
                    <p class="muted">Trage mehrere Produkte untereinander ein. Leere Zeilen werden automatisch übersprungen. Kategorien können mit Komma getrennt werden.</p>
                    <div id="bulkAddProductRows" class="bulk-add-product-rows">
                        <div class="bulk-add-product-row" data-bulk-add-row>
                            <label>Produktname<input name="bulk_name_0" placeholder="Produktname"></label>
                            <label>Kategorien<input name="bulk_category_0" placeholder="z. B. Getränke, Verpackung"></label>
                            <label>Gebinde<input name="bulk_package_size_0" placeholder="z. B. Karton à 12 Stück"></label>
                            <label>Bezugsquelle<input name="bulk_source_0" list="sourceSuggestions" placeholder="Vorhandene wählen oder neu eintragen"></label>
                        </div>
                        <div class="bulk-add-product-row" data-bulk-add-row>
                            <label>Produktname<input name="bulk_name_1" placeholder="Produktname"></label>
                            <label>Kategorien<input name="bulk_category_1" placeholder="z. B. Getränke, Verpackung"></label>
                            <label>Gebinde<input name="bulk_package_size_1" placeholder="z. B. Karton à 12 Stück"></label>
                            <label>Bezugsquelle<input name="bulk_source_1" list="sourceSuggestions" placeholder="Vorhandene wählen oder neu eintragen"></label>
                        </div>
                        <div class="bulk-add-product-row" data-bulk-add-row>
                            <label>Produktname<input name="bulk_name_2" placeholder="Produktname"></label>
                            <label>Kategorien<input name="bulk_category_2" placeholder="z. B. Getränke, Verpackung"></label>
                            <label>Gebinde<input name="bulk_package_size_2" placeholder="z. B. Karton à 12 Stück"></label>
                            <label>Bezugsquelle<input name="bulk_source_2" list="sourceSuggestions" placeholder="Vorhandene wählen oder neu eintragen"></label>
                        </div>
                    </div>
                    <template id="bulkAddProductTemplate">
                        <div class="bulk-add-product-row" data-bulk-add-row>
                            <label>Produktname<input name="bulk_name___INDEX__" placeholder="Produktname"></label>
                            <label>Kategorien<input name="bulk_category___INDEX__" placeholder="z. B. Getränke, Verpackung"></label>
                            <label>Gebinde<input name="bulk_package_size___INDEX__" placeholder="z. B. Karton à 12 Stück"></label>
                            <label>Bezugsquelle<input name="bulk_source___INDEX__" list="sourceSuggestions" placeholder="Vorhandene wählen oder neu eintragen"></label>
                            <button class="button bulk-add-remove" type="button">Zeile entfernen</button>
                        </div>
                    </template>
                    <div class="table-actions">
                        <button class="button" id="addBulkProductRow" type="button">Weitere Zeile hinzufügen</button>
                        <button class="primary" type="submit">Mehrere Produkte speichern</button>
                    </div>
                </form>
            </details>
        </details>
        <section class="box">
            <h2>Produkte</h2>
            <details class="category-panel admin-toggle-panel">
                <summary>Mehrere Produkte bearbeiten <span>+</span></summary>
            <form id="bulkProductForm" method="post" action="/admin/bulk-products" class="bulk-editor" data-confirm="Mehrfachbearbeitung auf ausgewählte Produkte anwenden?">
                <div class="bulk-editor-head">
                    <strong>Mehrere Produkte bearbeiten</strong>
                    <span class="muted">Produkte unten anhaken, gewünschte Felder setzen und anwenden.</span>
                </div>
                <div class="bulk-editor-grid">
                    <fieldset class="visibility-box bulk-visibility bulk-category-tools">
                        <legend>Produkte nach Kategorie auswählen</legend>
                        <label>Kategorie suchen<input id="bulkCategorySearch" type="search" placeholder="Kategorie suchen"></label>
                        <div class="bulk-category-buttons">{bulk_category_select_buttons}</div>
                        <p class="muted">Ein Klick wählt alle sichtbaren Produkte dieser Kategorie aus.</p>
                    </fieldset>
                    <label>Aktion
                        <select name="bulk_action">
                            <option value="update">Ausgewählte bearbeiten</option>
                            <option value="delete">Ausgewählte löschen</option>
                        </select>
                    </label>
                    <fieldset class="visibility-box bulk-visibility">
                        <legend>Kategorien setzen</legend>
                        <div class="visibility-grid">{bulk_category_checks}</div>
                        <p class="muted">Leer lassen = Kategorien behalten.</p>
                    </fieldset>
                    <label>Bezugsquelle
                        <input name="bulk_source" list="sourceSuggestions" placeholder="leer lassen = behalten">
                    </label>
                    <label>Status
                        <select name="bulk_active">
                            <option value="keep">Status behalten</option>
                            <option value="1">aktiv setzen</option>
                            <option value="0">deaktivieren</option>
                        </select>
                    </label>
                    <label>Sichtbarkeit
                        <select name="bulk_visibility_mode">
                            <option value="keep">Sichtbarkeit behalten</option>
                            <option value="all">für alle Standorte sichtbar</option>
                            <option value="custom">unten gewählte Standorte setzen</option>
                        </select>
                    </label>
                    <fieldset class="visibility-box bulk-visibility">
                        <legend>Standorte für Sichtbarkeit</legend>
                        <div class="visibility-grid">{bulk_visibility_checks}</div>
                    </fieldset>
                </div>
                <div class="bulk-editor-actions">
                    <label class="check"><input id="selectAllProducts" type="checkbox"> Alle sichtbaren Produkte auswählen</label>
                    <span class="bulk-selected-count" id="bulkSelectedCount">0 Produkte ausgewählt</span>
                    <button class="primary" type="submit">Auf Auswahl anwenden</button>
                </div>
            </form>
            </details>
            <form method="get" action="/admin" class="filters compact-form product-search">
                <label>Produktsuche<input name="q" value="{esc(search_text)}" placeholder="Name, Kategorie, Gebinde, Bezugsquelle oder Standort suchen"></label>
                <button type="submit">Suchen</button>
                <a class="button" href="/admin">Zurücksetzen</a>
            </form>
            <p class="muted">{len(products)} von {len(all_products)} Produkt(en) angezeigt.</p>
            <section class="category-sections">{''.join(admin_category_panels) if admin_category_panels else '<section class="box"><p>Keine passenden Produkte gefunden.</p></section>'}</section>
        </section>
        <section class="box"><h2>Letzte Bestellungen</h2><div class="table-wrap"><table><tr><th>Datum</th><th>Nr.</th><th>Zugang</th><th>Standort</th><th>Besteller</th><th>PDF</th><th>Bild</th><th>WhatsApp</th><th></th></tr>{''.join(order_rows) if order_rows else '<tr><td colspan="9">Noch keine Bestellungen.</td></tr>'}</table></div></section>
        """
        self.send_html(page("Adminbereich", body, admin=True, buyer_key=self.current_buyer_key()))

    def show_edit_product(self, query=None):
        if not self.is_admin():
            return self.redirect("/admin/login")
        query = query or {}
        product_id = (query.get("id", [""])[0] or "").strip()
        msg = (query.get("msg", [""])[0] or "").strip()
        error = (query.get("error", [""])[0] or "").strip()
        product = get_product(product_id)
        if not product:
            return self.redirect("/admin?error=" + quote_plus("Produkt wurde nicht gefunden."))
        selected_locations = set(product_visible_location_keys(product["visible_to"]))
        visibility_checks = "".join(
            f'<label class="visibility-option"><input type="checkbox" name="visible_{esc(location["id"])}" value="1" {"checked" if location["id"] in selected_locations else ""}><span>{esc(location["name"])}</span></label>'
            for location in get_locations()
        )
        category_checks = category_checkboxes("category", product_categories(product))
        source_options = "".join(f'<option value="{esc(source)}"></option>' for source in get_source_names())
        current_image = f'<p><img class="edit-preview" src="/uploads/{esc(product["image_filename"])}" alt=""></p>' if product["image_filename"] else "<p class='muted'>Für dieses Produkt ist noch kein Bild hinterlegt.</p>"
        body = f"""
        {admin_menu()}
        {f'<div class="success box narrow">{esc(msg)}</div>' if msg else ''}
        {f'<div class="error box narrow">{esc(error)}</div>' if error else ''}
        <section class="box">
            <div class="section-head"><div><h2>Produkt bearbeiten</h2><p class="muted">Änderungen sind nach dem Speichern sofort im Bestellmenü sichtbar.</p></div><a class="button" href="/admin">Zurück zu Produkten</a></div>
            <form method="post" action="/admin/update-product" enctype="multipart/form-data" class="two">
                <datalist id="sourceSuggestions">{source_options}</datalist>
                <input type="hidden" name="id" value="{esc(product['id'])}">
                <label>Produktname *<input name="name" required value="{esc(product['name'])}"></label>
                <fieldset class="visibility-box full">
                    <legend>Kategorien *</legend>
                    <div class="visibility-grid">{category_checks}</div>
                </fieldset>
                <label>Gebindegröße *<input name="package_size" required value="{esc(product['package_size'])}"></label>
                <label>Bezugsquelle *<input name="source" list="sourceSuggestions" required value="{esc(product['source'])}"></label>
                <label>Status<select name="active"><option value="1" {"selected" if product["active"] else ""}>aktiv</option><option value="0" {"selected" if not product["active"] else ""}>deaktiviert</option></select></label>
                <label>Neues Produktbild<input type="file" name="image" accept="image/*"></label>
                <div>{current_image}</div>
                <fieldset class="visibility-box">
                    <legend>Sichtbar für Standorte *</legend>
                    <div class="visibility-grid">{visibility_checks}</div>
                </fieldset>
                <button class="primary" type="submit">Änderungen speichern</button>
            </form>
        </section>
        """
        self.send_html(page("Produkt bearbeiten", body, admin=True, buyer_key=self.current_buyer_key()))

    def show_admin_orders(self, query=None):
        if not self.is_admin():
            return self.redirect("/admin/login")
        query = query or {}
        msg = (query.get("msg", [""])[0] or "").strip()
        error = (query.get("error", [""])[0] or "").strip()
        combined_pdf = (query.get("pdf", [""])[0] or "").strip()
        open_location = (query.get("open_location", [""])[0] or "").strip()
        orders = get_orders()
        grouped_orders = {}
        for order in orders:
            grouped_orders.setdefault(order["location"] or "Ohne Standort", []).append(order)
        location_panels = []
        for location_name in sorted(grouped_orders.keys(), key=lambda value: value.lower()):
            location_orders = sorted(grouped_orders[location_name], key=order_created_sort_key, reverse=True)
            day_groups = {}
            for order in location_orders:
                day_groups.setdefault(order_day_key(order), []).append(order)
            cards = []
            for day_key in sorted(day_groups.keys(), key=order_day_sort_key, reverse=True):
                day_orders = sorted(day_groups[day_key], key=order_created_sort_key, reverse=True)
                aggregated = {}
                order_rows = []
                note_rows = []
                hidden_inputs = []
                for order in day_orders:
                    hidden_inputs.append(f'<input type="hidden" name="order_{order["id"]}" value="1">')
                    is_completed = bool(order["completed_at"])
                    status_label = f"Erledigt am {esc(order['completed_at'])}" if is_completed else "Offen"
                    status_class = "order-status done" if is_completed else "order-status open"
                    image_link = f"<a class='button tiny' href='/order-images/{esc(order['order_image_filename'])}' target='_blank' rel='noopener'>Bild</a>" if order["order_image_filename"] else ""
                    order_rows.append(
                        f"""
                        <tr class="{'order-completed-row' if is_completed else ''}">
                            <td><label class="check order-select"><input form="combineOrdersForm" class="combine-order-check" type="checkbox" name="order_{order['id']}" value="1"> <span>{esc(order['order_number'])}</span></label></td>
                            <td>{esc(order['created_at'])}</td>
                            <td>{esc(order['ordered_by'])}</td>
                            <td><span class="{status_class}">{status_label}</span></td>
                            <td><div class="table-actions"><a class="button tiny" href="{esc(pdf_viewer_href(order['pdf_filename']))}">PDF</a>{image_link}</div></td>
                        </tr>
                        """
                    )
                    if order["note"]:
                        note_rows.append(f"<li><strong>{esc(order['order_number'])}:</strong> {esc(order['note'])}</li>")
                    for item in get_order_items(order["id"]):
                        key = (
                            item["product_id"] or "",
                            item["product_name"] or "",
                            item["category"] or "",
                            item["package_size"] or "",
                            item["source"] or "",
                        )
                        entry = aggregated.setdefault(
                            key,
                            {
                                "product_name": item["product_name"] or "",
                                "category": item["category"] or "",
                                "package_size": item["package_size"] or "",
                                "source": item["source"] or "",
                                "quantity": 0,
                                "completed": 0,
                                "open": 0,
                                "item_ids": [],
                            },
                        )
                        quantity = int(item["quantity"] or 0)
                        entry["quantity"] += quantity
                        entry["completed" if item["completed_at"] else "open"] += quantity
                        entry["item_ids"].append(str(item["id"]))
                item_rows = []
                for entry in sorted(aggregated.values(), key=lambda value: (value["source"].lower(), value["category"].lower(), value["product_name"].lower())):
                    item_rows.append(
                        f"""
                        <tr class="{'order-item-completed' if entry['open'] <= 0 and entry['quantity'] > 0 else ''}">
                            <td>{esc(entry['product_name'])}</td>
                            <td>{esc(entry['category'])}</td>
                            <td>{esc(entry['package_size'])}</td>
                            <td>{esc(entry['source'])}</td>
                            <td>{esc(entry['quantity'])}</td>
                            <td>{esc(entry['completed'])} erledigt · {esc(entry['open'])} offen</td>
                        </tr>
                        """
                    )
                position_count = len(aggregated)
                quantity_sum = sum(entry["quantity"] for entry in aggregated.values())
                day_label = order_day_label(day_key)
                day_completed = all(bool(order["completed_at"]) for order in day_orders)
                card_class = "box order-card order-day-card order-completed" if day_completed else "box order-card order-day-card"
                cards.append(
                    f"""
                    <article class="{card_class}">
                        <div class="section-head">
                            <div>
                                <h3>{esc(day_label)}</h3>
                                <p class="muted">{len(day_orders)} Einzelbestellung(en) · {position_count} zusammengefasste Position(en) · {quantity_sum} Gebinde</p>
                            </div>
                            <div class="table-actions">
                                <form method="post" action="/admin/orders/combined-pdf">
                                    {''.join(hidden_inputs)}
                                    <button class="primary" type="submit">Tagesbestellung als PDF</button>
                                </form>
                                <form method="post" action="/admin/orders/status">
                                    {''.join(hidden_inputs)}
                                    <input type="hidden" name="completed" value="{0 if day_completed else 1}">
                                    <button type="submit" class="{'button' if day_completed else 'primary'}">{'Tag wieder öffnen' if day_completed else 'Tag erledigt'}</button>
                                </form>
                                <form method="post" action="/admin/orders/delete" data-confirm="Alle Einzelbestellungen dieses Tages wirklich löschen?">
                                    {''.join(hidden_inputs)}
                                    <button type="submit" class="danger">Tag löschen</button>
                                </form>
                            </div>
                        </div>
                        <details class="category-panel order-day-panel">
                            <summary>Zusammengefasste Produkte <span>{position_count}</span></summary>
                            <div class="table-wrap"><table><tr><th>Produkt</th><th>Kategorie</th><th>Gebinde</th><th>Bezugsquelle</th><th>Menge</th><th>Status</th></tr>{''.join(item_rows) if item_rows else '<tr><td colspan="6">Keine Positionen gespeichert.</td></tr>'}</table></div>
                        </details>
                        <details class="category-panel order-day-panel">
                            <summary>Einzelbestellungen <span>{len(day_orders)}</span></summary>
                            <div class="table-wrap"><table><tr><th>Bestellnummer</th><th>Zeitpunkt</th><th>Von</th><th>Status</th><th>Dateien</th></tr>{''.join(order_rows)}</table></div>
                        </details>
                        {f'<div class="order-note-list"><strong>Bemerkungen:</strong><ul>{"".join(note_rows)}</ul></div>' if note_rows else ''}
                    </article>
                    """
                )
            location_panels.append(
                f"""
                <details class="category-panel order-location-panel">
                    <summary>{esc(location_name)} <span>{len(day_groups)} Tag(e), {len(location_orders)} Bestellung(en)</span></summary>
                    <div class="order-location-content">{''.join(cards)}</div>
                </details>
                """
            )
        combined_button = (
            f'<span class="inline-actions">'
            f'<a class="button primary" href="{esc(pdf_download_href(combined_pdf))}">Gesamtbestellung als A4-PDF herunterladen</a>'
            f'<a class="button" href="{esc(pdf_direct_href(combined_pdf))}" target="_blank" rel="noopener">Direkt öffnen / drucken</a>'
            f'<a class="button" href="{esc(pdf_viewer_href(combined_pdf))}">Vorschau öffnen</a>'
            f'</span>'
        ) if combined_pdf else ""
        body = f"""
        {admin_menu()}
        {f'<div class="success box narrow">{esc(msg)} {combined_button}</div>' if msg else ''}
        {f'<div class="error box narrow">{esc(error)}</div>' if error else ''}
        <section class="box">
            <h2>Getätigte Bestellungen</h2>
            <p class="muted">Hier siehst du Datum, Standort, Produkte, Mengen und die Gesamtübersicht jeder Bestellung.</p>
            <form id="combineOrdersForm" method="post" action="/admin/orders/combined-pdf" class="bulk-editor order-combine-bar" data-confirm="Aktion für die ausgewählten Bestellungen ausführen?">
                <div class="bulk-editor-actions">
                    <label class="check"><input id="selectAllOrders" type="checkbox"> Alle angezeigten Bestellungen auswählen</label>
                    <span class="bulk-selected-count" id="ordersSelectedCount">0 Bestellungen ausgewählt</span>
                    <button class="primary" type="submit">Gesamtbestellung als PDF erstellen</button>
                    <button type="submit" class="primary" formaction="/admin/orders/status" name="completed" value="1">Ausgewählte als erledigt markieren</button>
                    <button type="submit" class="danger" formaction="/admin/orders/delete">Ausgewählte löschen</button>
                </div>
            </form>
        </section>
        <section class="order-location-sections">{''.join(location_panels) if location_panels else '<section class="box"><p>Noch keine Bestellungen vorhanden.</p></section>'}</section>
        """
        self.send_html(page("Bestellungen", body, admin=True, buyer_key=self.current_buyer_key()))

    def show_categories(self, query=None):
        if not self.is_admin():
            return self.redirect("/admin/login")
        query = query or {}
        msg = (query.get("msg", [""])[0] or "").strip()
        error = (query.get("error", [""])[0] or "").strip()
        rows = []
        con = db()
        categories = get_all_categories(active_only=False)
        product_rows = con.execute("SELECT category FROM products WHERE deleted_at IS NULL").fetchall()
        for c in categories:
            status = "aktiv" if c["active"] else "deaktiviert"
            product_count = sum(1 for row in product_rows if product_has_category(row["category"], c["name"]))
            action = "—" if category_key(c["name"]) == category_key("Allgemein") else f"<form method='post' action='/admin/delete-category' data-confirm='Kategorie wirklich löschen? Das klappt nur, wenn keine Produkte mehr zugeordnet sind.'><input type='hidden' name='name' value='{esc(c['name'])}'><button class='danger'>Löschen</button></form>"
            rows.append(
                f"<tr><td><span class='pill'>{esc(c['name'])}</span></td><td>{product_count}</td><td>{status}</td><td>{action}</td></tr>"
            )
        con.close()
        body = f"""
        {admin_menu()}
        {f'<div class="success box narrow">{esc(msg)}</div>' if msg else ''}
        {f'<div class="error box narrow">{esc(error)}</div>' if error else ''}
        <section class="box narrow">
            <div class="section-head"><div><h2>Kategorie erstellen</h2><p class="muted">Neu angelegte Kategorien erscheinen beim Produkt-Neuanlegen als Auswahl.</p></div></div>
            <form method="post" action="/admin/add-category" class="inline-form">
                <label>Kategoriename *<input name="name" required placeholder="z. B. Getränke, Verpackung, TK-Ware"></label>
                <button class="primary" type="submit">Kategorie speichern</button>
            </form>
        </section>
        <section class="box">
            <h2>Alle Kategorien</h2>
            <p class="muted">Kategorien können gelöscht werden, sobald keine Produkte mehr zugeordnet sind. So gehen keine Produktzuordnungen unbemerkt verloren.</p>
            <div class="table-wrap"><table><tr><th>Name</th><th>Produkte</th><th>Status</th><th></th></tr>{''.join(rows) if rows else '<tr><td colspan="4">Noch keine Kategorien.</td></tr>'}</table></div>
        </section>
        """
        self.send_html(page("Kategorien", body, admin=True, buyer_key=self.current_buyer_key()))

    def show_import(self, query=None):
        if not self.is_admin():
            return self.redirect("/admin/login")
        query = query or {}
        msg = (query.get("msg", [""])[0] or "").strip()
        error = (query.get("error", [""])[0] or "").strip()
        body = f"""
        {admin_menu()}
        {f'<div class="success box narrow">{esc(msg)}</div>' if msg else ''}
        {f'<div class="error box narrow">{esc(error)}</div>' if error else ''}
        <section class="box">
            <div class="section-head"><div><h2>Produkte importieren</h2><p class="muted">CSV funktioniert direkt. Excel-Dateien funktionieren mit installierter openpyxl-Erweiterung.</p></div></div>
            <form method="post" action="/admin/import-products" enctype="multipart/form-data" class="two">
                <label>CSV- oder Excel-Datei *<input type="file" name="import_file" accept=".csv,.xlsx" required></label>
                <button class="primary" type="submit">Import starten</button>
            </form>
            <p class="muted">Importierte Produkte sind automatisch für alle Standorte sichtbar. Die Sichtbarkeit kann danach im Produktbereich angepasst werden.</p>
        </section>
        <section class="box">
            <h2>Dateiaufbau</h2>
            <p>Die Datei braucht eine Kopfzeile. Diese Spalten werden erkannt:</p>
            <table>
                <tr><th>Pflicht?</th><th>Spaltenname</th><th>Beispiel</th></tr>
                <tr><td>Ja</td><td>Produktname oder name</td><td>Kaffeebecher</td></tr>
                <tr><td>Ja</td><td>Gebindegröße oder package_size</td><td>Karton à 100 Stück</td></tr>
                <tr><td>Ja</td><td>Bezugsquelle oder source</td><td>Großhandel A</td></tr>
                <tr><td>Nein</td><td>Kategorie</td><td>Verbrauchsmaterial</td></tr>
            </table>
            <p class="muted">Neue Kategorien aus der Importdatei werden automatisch angelegt. Produktbilder können beim Massenimport nicht automatisch hochgeladen werden.</p>
            <p><a class="button" href="/admin/import-template.csv">CSV-Vorlage herunterladen</a></p>
        </section>
        """
        self.send_html(page("Produktimport", body, admin=True, buyer_key=self.current_buyer_key()))

    def show_settings(self, query=None):
        if not self.is_admin():
            return self.redirect("/admin/login")
        query = query or {}
        msg = (query.get("msg", [""])[0] or "").strip()
        error = (query.get("error", [""])[0] or "").strip()
        settings = load_settings()
        body = f"""
        {admin_menu()}
        {f'<div class="success box narrow">{esc(msg)}</div>' if msg else ''}
        {f'<div class="error box narrow">{esc(error)}</div>' if error else ''}
        <section class="box narrow">
            <h2>Versand-Einstellungen</h2>
            <form method="post" action="/admin/settings">
                <label>E-Mail-Adresse für Bestellungen<input type="email" name="order_email_to" value="{esc(settings.get('order_email_to', ''))}" placeholder="zentrale@example.com"></label>
                <label>WhatsApp-Kontakt / Nummer<input name="whatsapp_number" value="{esc(settings.get('whatsapp_number', ''))}" placeholder="z. B. +491701234567"></label>
                <label>E-Mail-Adresse für Zeiterfassungs-Export<input name="time_export_email" value="{esc(settings.get('time_export_email', ''))}" placeholder="info@opapeters"></label>
                <button class="primary" type="submit">Einstellungen speichern</button>
            </form>
            <p class="muted">E-Mail-Versand funktioniert nur, wenn zusätzlich SMTP-Daten beim Start gesetzt sind. WhatsApp wird als klickbarer Link vorbereitet; ein vollautomatischer WhatsApp-Versand benötigt die WhatsApp Business API.</p>
        </section>
        <section class="box narrow">
            <h2>Admin-PIN ändern</h2>
            <form method="post" action="/admin/settings" class="pin-change-form">
                <input type="hidden" name="change_admin_pin" value="1">
                <label>Aktuelle Admin-PIN<span class="password-wrap"><input class="password-field" type="password" name="current_admin_pin" inputmode="numeric" required><button type="button" class="password-toggle">Anzeigen</button></span></label>
                <label>Neue Admin-PIN<span class="password-wrap"><input class="password-field" type="password" name="new_admin_pin" inputmode="numeric" required><button type="button" class="password-toggle">Anzeigen</button></span></label>
                <label>Neue Admin-PIN wiederholen<span class="password-wrap"><input class="password-field" type="password" name="new_admin_pin_confirm" inputmode="numeric" required><button type="button" class="password-toggle">Anzeigen</button></span></label>
                <button class="primary" type="submit">Admin-PIN speichern</button>
            </form>
        </section>
        <section class="box narrow">
            <h2>Aktueller Status</h2>
            <p><strong>E-Mail-Empfänger:</strong> {esc(settings.get('order_email_to') or 'nicht gesetzt')}</p>
            <p><strong>WhatsApp-Kontakt:</strong> {esc(settings.get('whatsapp_number') or 'nicht gesetzt')}</p>
            <p><strong>Zeiterfassungs-Export:</strong> {esc(settings.get('time_export_email') or 'nicht gesetzt')}</p>
            <p><strong>Push:</strong> {esc(push_status)} · {push_count} aktive(s) Admin-Gerät(e)</p>
            <form method="post" action="/admin/push/test">
                <button class="button primary" type="submit" {"disabled" if not push_configured or push_count == 0 else ""}>Push-Test senden</button>
            </form>
        </section>"""
        self.send_html(page("Admin Einstellungen", body, admin=True, buyer_key=self.current_buyer_key()))


    def show_locations(self, query=None):
        if not self.is_admin():
            return self.redirect("/admin/login")
        query = query or {}
        msg = (query.get("msg", [""])[0] or "").strip()
        locations = get_locations()
        time_limit_options = '<option value="">Keine feste maximale Endzeit</option>' + option_html(time_options())
        categories = get_category_names(True)
        products_for_visibility = get_products(False, sort="category")
        products_for_min_stock = get_products(False, sort="category")
        def min_stock_product_options(selected_product_id=""):
            selected_product_id = str(selected_product_id or "")
            options = ['<option value="">Produkt auswählen</option>']
            for product in products_for_min_stock:
                label = f"{product['name']} · {product['package_size']} · {category_text(product_categories(product))}"
                selected = "selected" if str(product["id"]) == selected_product_id else ""
                options.append(f'<option value="{product["id"]}" {selected}>{esc(label)}</option>')
            return "".join(options)
        def min_stock_rows(prefix, items, product_options_blank):
            normalized_items = normalize_min_stock_items(items)
            row_items = normalized_items if normalized_items else [{"product_id": "", "quantity": ""}]
            rows_html = []
            for row_index, item in enumerate(row_items):
                rows_html.append(
                    f"""
                    <div class="min-stock-row" data-min-stock-row>
                        <label>Produkt<select name="{prefix}_product_{row_index}">{min_stock_product_options(item.get('product_id'))}</select></label>
                        <label>Mindestmenge<input type="number" name="{prefix}_qty_{row_index}" min="1" max="{MAX_ORDER_QUANTITY}" step="1" inputmode="numeric" value="{esc(item.get('quantity', ''))}"></label>
                        <button class="button min-stock-remove" type="button">Zeile entfernen</button>
                    </div>
                    """
                )
            template = f"""
            <template id="{prefix}_template">
                <div class="min-stock-row" data-min-stock-row>
                    <label>Produkt<select name="{prefix}_product___INDEX__">{product_options_blank}</select></label>
                    <label>Mindestmenge<input type="number" name="{prefix}_qty___INDEX__" min="1" max="{MAX_ORDER_QUANTITY}" step="1" inputmode="numeric"></label>
                    <button class="button min-stock-remove" type="button">Zeile entfernen</button>
                </div>
            </template>
            """
            return "".join(rows_html), len(row_items), template
        def cockpit_task_rows(prefix, tasks):
            row_items = normalize_cockpit_tasks(tasks) or [{"id": "", "title": "", "template": False, "active": True, "image_filename": ""}]
            rows_html = []
            for row_index, item in enumerate(row_items):
                image_filename = item.get("image_filename", "")
                image_preview = f'<a class="cockpit-image-preview" href="/uploads/{esc(image_filename)}" target="_blank" rel="noopener"><img src="/uploads/{esc(image_filename)}" alt="">Bild öffnen</a>' if image_filename else '<span class="muted">Kein Bild hinterlegt.</span>'
                rows_html.append(
                    f"""
                    <div class="cockpit-admin-row" data-cockpit-task-row>
                        <input type="hidden" name="{prefix}_task_id_{row_index}" value="{esc(item.get('id', ''))}">
                        <input type="hidden" name="{prefix}_task_image_current_{row_index}" value="{esc(image_filename)}">
                        <label>Aufgabe<input name="{prefix}_task_title_{row_index}" value="{esc(item.get('title', ''))}" placeholder="z. B. Kühlschrank prüfen"></label>
                        <label>Bild<input type="file" name="{prefix}_task_image_{row_index}" accept="image/*">{image_preview}</label>
                        <label class="check"><input id="{prefix}_task_active_{row_index}" type="checkbox" name="{prefix}_task_active_{row_index}" value="1" {"checked" if item.get('active', True) else ""}> Im Cockpit anzeigen</label>
                        <label class="check"><input type="checkbox" name="{prefix}_task_template_{row_index}" value="1" {"checked" if item.get('template') else ""}> Als Muster speichern</label>
                        <button class="button cockpit-row-remove" type="button">Zeile entfernen</button>
                        <button class="primary cockpit-row-save" type="submit">Diese Aufgabe speichern</button>
                    </div>
                    """
                )
            template = f"""
            <template id="{prefix}_task_template">
                <div class="cockpit-admin-row" data-cockpit-task-row>
                    <input type="hidden" name="{prefix}_task_id___INDEX__" value="">
                    <input type="hidden" name="{prefix}_task_image_current___INDEX__" value="">
                    <label>Aufgabe<input name="{prefix}_task_title___INDEX__" placeholder="z. B. Kühlschrank prüfen"></label>
                    <label>Bild<input type="file" name="{prefix}_task_image___INDEX__" accept="image/*"><span class="muted">Optional</span></label>
                    <label class="check"><input id="{prefix}_task_active___INDEX__" type="checkbox" name="{prefix}_task_active___INDEX__" value="1" checked> Im Cockpit anzeigen</label>
                    <label class="check"><input type="checkbox" name="{prefix}_task_template___INDEX__" value="1"> Als Muster speichern</label>
                    <button class="button cockpit-row-remove" type="button">Zeile entfernen</button>
                    <button class="primary cockpit-row-save" type="submit">Diese Aufgabe speichern</button>
                </div>
            </template>
            """
            template_cards = []
            for row_index, item in enumerate(row_items):
                if not item.get("template"):
                    continue
                template_cards.append(
                    f"""
                    <div class="task-template-card">
                        <strong>{esc(item.get('title', ''))}</strong>
                        <button class="button activate-template-task" type="button" data-target="{prefix}_task_active_{row_index}">Im Cockpit anzeigen</button>
                    </div>
                    """
                )
            template_collection = f"""
            <div class="task-template-box">
                <h3>Musteraufgaben</h3>
                <p class="muted">Gespeicherte Muster kannst du hier schnell aktivieren.</p>
                <div class="task-template-list">{''.join(template_cards) if template_cards else '<p class="muted">Noch keine Musteraufgaben gespeichert.</p>'}</div>
            </div>
            """
            return "".join(rows_html), len(row_items), template, template_collection
        def cockpit_order_rows(prefix, orders):
            row_items = normalize_cockpit_orders(orders) or [{"id": "", "title": "", "note": "", "active": True, "image_filename": ""}]
            rows_html = []
            for row_index, item in enumerate(row_items):
                image_filename = item.get("image_filename", "")
                image_preview = f'<a class="cockpit-image-preview" href="/uploads/{esc(image_filename)}" target="_blank" rel="noopener"><img src="/uploads/{esc(image_filename)}" alt="">Bild öffnen</a>' if image_filename else '<span class="muted">Kein Bild hinterlegt.</span>'
                rows_html.append(
                    f"""
                    <div class="cockpit-admin-row cockpit-admin-order-row" data-cockpit-order-row>
                        <input type="hidden" name="{prefix}_order_id_{row_index}" value="{esc(item.get('id', ''))}">
                        <input type="hidden" name="{prefix}_order_image_current_{row_index}" value="{esc(image_filename)}">
                        <label>Bestellung<input name="{prefix}_order_title_{row_index}" value="{esc(item.get('title', ''))}" placeholder="z. B. Tortenbestellung Müller"></label>
                        <label>Hinweis<textarea name="{prefix}_order_note_{row_index}" rows="2" placeholder="Optional">{esc(item.get('note', ''))}</textarea></label>
                        <label>Bild<input type="file" name="{prefix}_order_image_{row_index}" accept="image/*">{image_preview}</label>
                        <label class="check"><input type="checkbox" name="{prefix}_order_active_{row_index}" value="1" {"checked" if item.get('active', True) else ""}> Im Cockpit anzeigen</label>
                        <button class="button cockpit-row-remove" type="button">Zeile entfernen</button>
                        <button class="primary cockpit-row-save" type="submit">Diese Bestellung speichern</button>
                    </div>
                    """
                )
            template = f"""
            <template id="{prefix}_order_template">
                <div class="cockpit-admin-row cockpit-admin-order-row" data-cockpit-order-row>
                    <input type="hidden" name="{prefix}_order_id___INDEX__" value="">
                    <input type="hidden" name="{prefix}_order_image_current___INDEX__" value="">
                    <label>Bestellung<input name="{prefix}_order_title___INDEX__" placeholder="z. B. Tortenbestellung Müller"></label>
                    <label>Hinweis<textarea name="{prefix}_order_note___INDEX__" rows="2" placeholder="Optional"></textarea></label>
                    <label>Bild<input type="file" name="{prefix}_order_image___INDEX__" accept="image/*"><span class="muted">Optional</span></label>
                    <label class="check"><input type="checkbox" name="{prefix}_order_active___INDEX__" value="1" checked> Im Cockpit anzeigen</label>
                    <button class="button cockpit-row-remove" type="button">Zeile entfernen</button>
                    <button class="primary cockpit-row-save" type="submit">Diese Bestellung speichern</button>
                </div>
            </template>
            """
            return "".join(rows_html), len(row_items), template
        def production_job_rows(prefix, jobs):
            section_options = lambda selected="": "".join(
                f'<option value="{esc(key)}" {"selected" if key == selected else ""}>{esc(label)}</option>'
                for key, label in PRODUCTION_SECTIONS.items()
            )
            row_items = normalize_production_jobs(jobs) or [{"id": "", "section": "backen", "customer_name": "", "order_text": "", "due_date": "", "note": "", "active": True, "image_filename": ""}]
            rows_html = []
            for row_index, item in enumerate(row_items):
                image_filename = item.get("image_filename", "")
                image_preview = f'<a class="cockpit-image-preview" href="/uploads/{esc(image_filename)}" target="_blank" rel="noopener"><img src="/uploads/{esc(image_filename)}" alt="">Bild öffnen</a>' if image_filename else '<span class="muted">Kein Bild hinterlegt.</span>'
                rows_html.append(
                    f"""
                    <div class="cockpit-admin-row cockpit-admin-production-row" data-cockpit-production-row>
                        <input type="hidden" name="{prefix}_production_id_{row_index}" value="{esc(item.get('id', ''))}">
                        <input type="hidden" name="{prefix}_production_image_current_{row_index}" value="{esc(image_filename)}">
                        <label>Bereich<select name="{prefix}_production_section_{row_index}">{section_options(item.get('section', 'backen'))}</select></label>
                        <label>Name<input name="{prefix}_production_customer_name_{row_index}" value="{esc(item.get('customer_name') or item.get('title', ''))}" placeholder="z. B. Müller"></label>
                        <label>Abholdatum<input type="date" name="{prefix}_production_due_date_{row_index}" value="{esc(item.get('due_date', ''))}"></label>
                        <label>Was wurde bestellt<textarea name="{prefix}_production_order_text_{row_index}" rows="2" placeholder="z. B. Eistorte Erdbeere">{esc(item.get('order_text', ''))}</textarea></label>
                        <label>Hinweis<textarea name="{prefix}_production_note_{row_index}" rows="2" placeholder="Optional">{esc(item.get('note', ''))}</textarea></label>
                        <label>Bild<input type="file" name="{prefix}_production_image_{row_index}" accept="image/*">{image_preview}</label>
                        <label class="check"><input type="checkbox" name="{prefix}_production_active_{row_index}" value="1" {"checked" if item.get('active', True) else ""}> Im Cockpit anzeigen</label>
                        <button class="button cockpit-row-remove" type="button">Zeile entfernen</button>
                        <button class="primary cockpit-row-save" type="submit">Diesen Auftrag speichern</button>
                    </div>
                    """
                )
            template = f"""
            <template id="{prefix}_production_template">
                <div class="cockpit-admin-row cockpit-admin-production-row" data-cockpit-production-row>
                    <input type="hidden" name="{prefix}_production_id___INDEX__" value="">
                    <input type="hidden" name="{prefix}_production_image_current___INDEX__" value="">
                    <label>Bereich<select name="{prefix}_production_section___INDEX__">{section_options('backen')}</select></label>
                    <label>Name<input name="{prefix}_production_customer_name___INDEX__" placeholder="z. B. Müller"></label>
                    <label>Abholdatum<input type="date" name="{prefix}_production_due_date___INDEX__"></label>
                    <label>Was wurde bestellt<textarea name="{prefix}_production_order_text___INDEX__" rows="2" placeholder="z. B. Eistorte Erdbeere"></textarea></label>
                    <label>Hinweis<textarea name="{prefix}_production_note___INDEX__" rows="2" placeholder="Optional"></textarea></label>
                    <label>Bild<input type="file" name="{prefix}_production_image___INDEX__" accept="image/*"><span class="muted">Optional</span></label>
                    <label class="check"><input type="checkbox" name="{prefix}_production_active___INDEX__" value="1" checked> Im Cockpit anzeigen</label>
                    <button class="button cockpit-row-remove" type="button">Zeile entfernen</button>
                    <button class="primary cockpit-row-save" type="submit">Diesen Auftrag speichern</button>
                </div>
            </template>
            """
            return "".join(rows_html), len(row_items), template
        category_visibility_checks = "".join(
            f'<label class="visibility-option"><input type="checkbox" name="new_location_category_{esc(category)}" value="1" checked><span>{esc(category)}</span></label>'
            for category in categories
        )
        rows = []
        for index, location in enumerate(locations):
            max_end_options = '<option value="">Keine feste maximale Endzeit</option>' + option_html(time_options(), location.get("time_tracking_max_end", ""))
            current_role_options = role_options(location.get("role", "standard"))
            location_category_checks = []
            for category in categories:
                category_products = [p for p in products_for_visibility if product_has_category(p, category)]
                all_visible = bool(category_products) and all(location["id"] in product_visible_location_keys(p["visible_to"]) for p in category_products)
                location_category_checks.append(
                    f'<label class="visibility-option"><input type="checkbox" name="location_category_{index}_{esc(category)}" value="1" {"checked" if all_visible else ""}><span>{esc(category)} ({len(category_products)})</span></label>'
                )
            min_stock_prefix = f"location_min_stock_{index}"
            min_stock_enabled = bool(location.get("min_stock_enabled"))
            min_stock_row_html, min_stock_count, min_stock_template = min_stock_rows(
                min_stock_prefix,
                location.get("min_stock_items", []),
                min_stock_product_options(""),
            )
            cockpit_prefix = f"location_cockpit_{index}"
            cockpit_task_html, cockpit_task_count, cockpit_task_template, cockpit_task_templates = cockpit_task_rows(cockpit_prefix, location.get("cockpit_tasks", []))
            cockpit_order_html, cockpit_order_count, cockpit_order_template = cockpit_order_rows(cockpit_prefix, location.get("cockpit_orders", []))
            production_job_html, production_job_count, production_job_template = production_job_rows(cockpit_prefix, location.get("production_jobs", []))
            cockpit_order_section = "" if is_schwarzenbek_location(location) else f"""
                            <details class="category-panel cockpit-admin-panel">
                                <summary>Bestellungen <span>{cockpit_order_count}</span></summary>
                                <p class="muted">Diese Bestellinformationen erscheinen im Cockpit mit Statusauswahl.</p>
                                <input type="hidden" id="{cockpit_prefix}_order_count" name="{cockpit_prefix}_order_count" value="{cockpit_order_count}">
                                <div id="{cockpit_prefix}_order_rows" class="cockpit-admin-rows" data-template-id="{cockpit_prefix}_order_template" data-count-id="{cockpit_prefix}_order_count">{cockpit_order_html}</div>
                                {cockpit_order_template}
                                <button class="button add-cockpit-row" type="button" data-target="{cockpit_prefix}_order_rows">Bestellung hinzufügen</button>
                            </details>
            """
            rows.append(
                f"""
                <details id="location-{esc(location['id'])}" class="category-panel location-panel" data-location-id="{esc(location['id'])}">
                    <summary>{esc(location['name'])} <span>{esc(role_label(location.get('role', 'standard')))}</span></summary>
                    <div class="location-row">
                        <input type="hidden" name="location_id_{index}" value="{esc(location['id'])}">
                        <label>Standortname<input name="location_name_{index}" value="{esc(location['name'])}" required></label>
                        <label>Rolle<select name="location_role_{index}">{current_role_options}</select></label>
                        <label>Name der bestellenden Person<input name="location_contact_name_{index}" value="{esc(location.get('contact_name', ''))}" placeholder="z. B. Lisa"></label>
                        <label>Passwort<span class="password-wrap"><input class="password-field" type="password" name="location_password_{index}" value="{esc(location.get('password', ''))}" autocomplete="new-password" placeholder="Leer lassen = kein Passwort"><button type="button" class="password-toggle">Anzeigen</button></span></label>
                        <label class="check feature-check"><input type="checkbox" name="location_time_tracking_{index}" value="1" {"checked" if location.get("time_tracking_enabled") else ""}> Zeiterfassung aktivieren</label>
                        <label>Maximale Endzeit<select name="location_time_tracking_max_end_{index}">{max_end_options}</select></label>
                        <label class="check remove-check danger-check"><input type="checkbox" name="location_remove_{index}" value="1"> Diesen Standort löschen</label>
                        <fieldset class="visibility-box location-category-visibility">
                            <legend>Produktkategorien für diesen Standort</legend>
                            <div class="visibility-grid">{''.join(location_category_checks)}</div>
                            <p class="muted">Hier steuerst du Kategorien. Einzelne Produkte kannst du zusätzlich unter <a href="/admin/visibility?location={esc(location['id'])}">Sichtbarkeit</a> bearbeiten.</p>
                        </fieldset>
                        <fieldset class="visibility-box min-stock-admin full">
                            <legend>Mindestbestand</legend>
                            <label class="check feature-check"><input class="min-stock-toggle" type="checkbox" name="location_min_stock_enabled_{index}" value="1" data-target="min_stock_content_{index}" {"checked" if min_stock_enabled else ""}> Mindestbestellmengen für diesen Standort aktivieren</label>
                            <div id="min_stock_content_{index}" class="min-stock-admin-content" {"hidden" if not min_stock_enabled else ""}>
                                <p class="muted">Wähle Produkte aus dem Sortiment und hinterlege pro Produkt eine Orientierungsmindestmenge.</p>
                                <input type="hidden" id="{min_stock_prefix}_count" name="{min_stock_prefix}_count" value="{min_stock_count}">
                                <div id="{min_stock_prefix}_rows" class="min-stock-rows" data-template-id="{min_stock_prefix}_template" data-count-id="{min_stock_prefix}_count">{min_stock_row_html}</div>
                                {min_stock_template}
                                <button class="button add-min-stock-row" type="button" data-target="{min_stock_prefix}_rows">Produkt hinzufügen</button>
                            </div>
                        </fieldset>
                        <fieldset class="visibility-box cockpit-admin full">
                            <legend>Cockpit</legend>
                            <details class="category-panel cockpit-admin-panel">
                                <summary>Tagesaufgaben <span>{cockpit_task_count}</span></summary>
                                <p class="muted">Aufgaben werden im Standort-Cockpit angezeigt und können dort abgehakt werden.</p>
                                <input type="hidden" id="{cockpit_prefix}_task_count" name="{cockpit_prefix}_task_count" value="{cockpit_task_count}">
                                <div id="{cockpit_prefix}_task_rows" class="cockpit-admin-rows" data-template-id="{cockpit_prefix}_task_template" data-count-id="{cockpit_prefix}_task_count">{cockpit_task_html}</div>
                                {cockpit_task_template}
                                <button class="button add-cockpit-row" type="button" data-target="{cockpit_prefix}_task_rows">Aufgabe hinzufügen</button>
                                {cockpit_task_templates}
                            </details>
                            {cockpit_order_section}
                            <details class="category-panel cockpit-admin-panel">
                                <summary>Produktion <span>{production_job_count}</span></summary>
                                <p class="muted">Diese Aufträge erscheinen in der Produktions-Wochenansicht und Eisvitrinen/Eistorten tagesaktuell im Cockpit Schwarzenbek.</p>
                                <input type="hidden" id="{cockpit_prefix}_production_count" name="{cockpit_prefix}_production_count" value="{production_job_count}">
                                <div id="{cockpit_prefix}_production_rows" class="cockpit-admin-rows" data-template-id="{cockpit_prefix}_production_template" data-count-id="{cockpit_prefix}_production_count">{production_job_html}</div>
                                {production_job_template}
                                <button class="button add-cockpit-row" type="button" data-target="{cockpit_prefix}_production_rows">Produktionsauftrag hinzufügen</button>
                            </details>
                        </fieldset>
                    </div>
                </details>
                """
            )
        new_min_stock_prefix = "new_location_min_stock"
        new_min_stock_row_html, new_min_stock_count, new_min_stock_template = min_stock_rows(
            new_min_stock_prefix,
            [],
            min_stock_product_options(""),
        )
        new_cockpit_prefix = "new_location_cockpit"
        new_cockpit_task_html, new_cockpit_task_count, new_cockpit_task_template, new_cockpit_task_templates = cockpit_task_rows(new_cockpit_prefix, [])
        new_cockpit_order_html, new_cockpit_order_count, new_cockpit_order_template = cockpit_order_rows(new_cockpit_prefix, [])
        new_production_job_html, new_production_job_count, new_production_job_template = production_job_rows(new_cockpit_prefix, [])
        body = f"""
        {admin_menu()}
        {f'<div class="success box narrow">{esc(msg)}</div>' if msg else ''}
        <section class="box">
            <h2>Standorte bearbeiten</h2>
            <form method="post" action="/admin/locations" enctype="multipart/form-data">
                <input type="hidden" name="location_count" value="{len(locations)}">
                <input type="hidden" id="locationsOpenLocation" name="open_location" value="{esc(open_location)}">
                <div class="location-list">{''.join(rows) if rows else '<p>Noch keine Standorte.</p>'}</div>
                <details class="category-panel location-panel new-location-panel">
                    <summary>Neuen Standort anlegen <span>+</span></summary>
                    <div class="location-row new-location">
                        <label>Neuer Standort<input name="new_location_name" placeholder="z. B. Filiale 4"></label>
                        <label>Rolle<select name="new_location_role">{role_options("standard")}</select></label>
                        <label>Name der bestellenden Person<input name="new_location_contact_name" placeholder="Optional"></label>
                        <label>Passwort<span class="password-wrap"><input class="password-field" type="password" name="new_location_password" autocomplete="new-password" placeholder="Optional"><button type="button" class="password-toggle">Anzeigen</button></span></label>
                        <label class="check feature-check"><input type="checkbox" name="new_location_time_tracking" value="1"> Zeiterfassung aktivieren</label>
                        <label>Maximale Endzeit<select name="new_location_time_tracking_max_end">{time_limit_options}</select></label>
                        <fieldset class="visibility-box location-category-visibility">
                            <legend>Produktkategorien für neuen Standort</legend>
                            <div class="visibility-grid">{category_visibility_checks}</div>
                            <p class="muted">Nur Produkte aus diesen Kategorien werden für den neuen Standort sichtbar.</p>
                        </fieldset>
                        <fieldset class="visibility-box min-stock-admin full">
                            <legend>Mindestbestand</legend>
                            <label class="check feature-check"><input class="min-stock-toggle" type="checkbox" name="new_location_min_stock_enabled" value="1" data-target="new_location_min_stock_content"> Mindestbestellmengen für diesen Standort aktivieren</label>
                            <div id="new_location_min_stock_content" class="min-stock-admin-content" hidden>
                                <p class="muted">Wähle Produkte aus dem Sortiment und hinterlege pro Produkt eine Orientierungsmindestmenge.</p>
                                <input type="hidden" id="{new_min_stock_prefix}_count" name="{new_min_stock_prefix}_count" value="{new_min_stock_count}">
                                <div id="{new_min_stock_prefix}_rows" class="min-stock-rows" data-template-id="{new_min_stock_prefix}_template" data-count-id="{new_min_stock_prefix}_count">{new_min_stock_row_html}</div>
                                {new_min_stock_template}
                                <button class="button add-min-stock-row" type="button" data-target="{new_min_stock_prefix}_rows">Produkt hinzufügen</button>
                            </div>
                        </fieldset>
                        <fieldset class="visibility-box cockpit-admin full">
                            <legend>Cockpit</legend>
                            <details class="category-panel cockpit-admin-panel">
                                <summary>Tagesaufgaben <span>{new_cockpit_task_count}</span></summary>
                                <input type="hidden" id="{new_cockpit_prefix}_task_count" name="{new_cockpit_prefix}_task_count" value="{new_cockpit_task_count}">
                                <div id="{new_cockpit_prefix}_task_rows" class="cockpit-admin-rows" data-template-id="{new_cockpit_prefix}_task_template" data-count-id="{new_cockpit_prefix}_task_count">{new_cockpit_task_html}</div>
                                {new_cockpit_task_template}
                                <button class="button add-cockpit-row" type="button" data-target="{new_cockpit_prefix}_task_rows">Aufgabe hinzufügen</button>
                                {new_cockpit_task_templates}
                            </details>
                            <details class="category-panel cockpit-admin-panel">
                                <summary>Bestellungen <span>{new_cockpit_order_count}</span></summary>
                                <input type="hidden" id="{new_cockpit_prefix}_order_count" name="{new_cockpit_prefix}_order_count" value="{new_cockpit_order_count}">
                                <div id="{new_cockpit_prefix}_order_rows" class="cockpit-admin-rows" data-template-id="{new_cockpit_prefix}_order_template" data-count-id="{new_cockpit_prefix}_order_count">{new_cockpit_order_html}</div>
                                {new_cockpit_order_template}
                                <button class="button add-cockpit-row" type="button" data-target="{new_cockpit_prefix}_order_rows">Bestellung hinzufügen</button>
                            </details>
                            <details class="category-panel cockpit-admin-panel">
                                <summary>Produktion <span>{new_production_job_count}</span></summary>
                                <p class="muted">Diese Aufträge erscheinen in der Produktions-Wochenansicht und Eisvitrinen/Eistorten tagesaktuell im Cockpit Schwarzenbek.</p>
                                <input type="hidden" id="{new_cockpit_prefix}_production_count" name="{new_cockpit_prefix}_production_count" value="{new_production_job_count}">
                                <div id="{new_cockpit_prefix}_production_rows" class="cockpit-admin-rows" data-template-id="{new_cockpit_prefix}_production_template" data-count-id="{new_cockpit_prefix}_production_count">{new_production_job_html}</div>
                                {new_production_job_template}
                                <button class="button add-cockpit-row" type="button" data-target="{new_cockpit_prefix}_production_rows">Produktionsauftrag hinzufügen</button>
                            </details>
                        </fieldset>
                    </div>
                </details>
                <button class="primary" type="submit">Standorte speichern</button>
            </form>
            <p class="muted">Das Passwort wird bei der Bestellung für den gewählten Standort abgefragt. Leeres Passwort bedeutet: keine Passwortprüfung für diesen Standort.</p>
        </section>"""
        self.send_html(page("Standorte", body, admin=True, buyer_key=self.current_buyer_key()))

    def show_visibility(self, query=None):
        if not self.is_admin():
            return self.redirect("/admin/login")
        query = query or {}
        msg = (query.get("msg", [""])[0] or "").strip()
        error = (query.get("error", [""])[0] or "").strip()
        locations = get_locations()
        selected_location_id = (query.get("location", [""])[0] or "").strip()
        if not selected_location_id and locations:
            selected_location_id = locations[0]["id"]
        selected_location = find_location(selected_location_id)
        products = get_products(False, sort="category")
        categories = sorted({category for p in products for category in product_categories(p)}, key=lambda item: item.lower())
        location_options = "".join(
            f'<option value="{esc(location["id"])}" {"selected" if location["id"] == selected_location_id else ""}>{esc(location["name"])}</option>'
            for location in locations
        )
        visible_product_ids = {
            str(product["id"]) for product in products
            if selected_location_id in product_visible_location_keys(product["visible_to"])
        }
        category_checks = []
        for category in categories:
            category_products = [p for p in products if product_has_category(p, category)]
            all_visible = bool(category_products) and all(str(p["id"]) in visible_product_ids for p in category_products)
            category_checks.append(
                f'<label class="visibility-option"><input class="visibility-category-toggle" type="checkbox" name="category_{esc(category)}" value="1" data-category="{esc(category)}" {"checked" if all_visible else ""}><span>{esc(category)} ({len(category_products)})</span></label>'
            )
        product_panels = []
        for category in categories:
            category_products = [p for p in products if product_has_category(p, category)]
            product_checks = "".join(
                f'<label class="visibility-option product-visibility-option"><input class="visibility-product-check" type="checkbox" name="product_{product["id"]}" value="1" data-category="{esc(category)}" {"checked" if str(product["id"]) in visible_product_ids else ""}><span>{esc(product["name"])}<small>{esc(product["package_size"])}</small></span></label>'
                for product in category_products
            )
            product_panels.append(
                f"<details class='category-panel admin-product-panel'><summary>{esc(category)} <span>{len(category_products)}</span></summary><div class='visibility-grid product-visibility-grid'>{product_checks}</div></details>"
            )
        body = f"""
        {admin_menu()}
        {f'<div class="success box narrow">{esc(msg)}</div>' if msg else ''}
        {f'<div class="error box narrow">{esc(error)}</div>' if error else ''}
        <section class="box">
            <div class="section-head">
                <div>
                    <h2>Sichtbarkeit je Standort</h2>
                    <p class="muted">Lege fest, welche Produktkategorien oder Einzelprodukte ein Standort im Bestellbereich sehen darf.</p>
                </div>
                <a class="button" href="/admin/locations">Standorte bearbeiten</a>
            </div>
            <form method="get" action="/admin/visibility" class="filters compact-form">
                <label>Standort auswählen<select name="location">{location_options}</select></label>
                <button class="primary" type="submit">Anzeigen</button>
            </form>
        </section>
        <section class="box">
            <h2>{esc(selected_location["name"] if selected_location else "Kein Standort")}</h2>
            <form method="post" action="/admin/visibility" data-confirm="Sichtbarkeit für diesen Standort speichern?">
                <input type="hidden" name="location_id" value="{esc(selected_location_id)}">
                <fieldset class="visibility-box">
                    <legend>Kategorien schnell auswählen</legend>
                    <div class="visibility-grid">{''.join(category_checks) if category_checks else '<p>Keine Kategorien vorhanden.</p>'}</div>
                    <p class="muted">Eine Kategorie setzt die darunterliegenden Produkte. Einzelprodukte kannst du danach zusätzlich an- oder abwählen.</p>
                </fieldset>
                <div class="category-sections">{''.join(product_panels) if product_panels else '<p>Keine Produkte vorhanden.</p>'}</div>
                <button class="primary visibility-save" type="submit">Sichtbarkeit speichern</button>
            </form>
        </section>
        """
        self.send_html(page("Sichtbarkeit", body, admin=True, buyer_key=self.current_buyer_key()))

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path
        maybe_run_auto_time_export()
        try:
            if path == "/login":
                return self.handle_buyer_login()
            if path == "/cart-draft":
                return self.handle_cart_draft()
            if path == "/cockpit/task":
                return self.handle_cockpit_task_state()
            if path == "/cockpit/order-state":
                return self.handle_cockpit_order_state()
            if path == "/cockpit/production-job":
                return self.handle_cockpit_production_job_state()
            if path == "/api/production-job":
                return self.handle_api_production_job()
            if path == "/push/subscribe":
                return self.handle_push_subscribe()
            if path == "/push/unsubscribe":
                return self.handle_push_unsubscribe()
            if path == "/order":
                return self.handle_order()
            if path == "/time":
                return self.handle_time_entry()
            if path == "/admin/login":
                return self.handle_admin_login()
            if path == "/admin/push/test":
                return self.handle_admin_push_test()
            if path == "/admin/add-product":
                return self.handle_add_product()
            if path == "/admin/update-product":
                return self.handle_update_product()
            if path == "/admin/bulk-products":
                return self.handle_bulk_products()
            if path == "/admin/orders/combined-pdf":
                return self.handle_combined_order_pdf()
            if path == "/admin/orders/status":
                return self.handle_order_status()
            if path == "/admin/orders/delete":
                return self.handle_delete_orders()
            if path == "/admin/order-items/status":
                return self.handle_order_item_status()
            if path == "/admin/order-items/delete":
                return self.handle_delete_order_items()
            if path == "/admin/delete-product":
                return self.handle_delete_product()
            if path == "/admin/add-category":
                return self.handle_add_category()
            if path == "/admin/delete-category":
                return self.handle_delete_category()
            if path == "/admin/import-products":
                return self.handle_import_products()
            if path == "/admin/settings":
                return self.handle_settings()
            if path == "/admin/employees":
                return self.handle_employees()
            if path == "/admin/locations":
                return self.handle_locations()
            if path == "/admin/visibility":
                return self.handle_visibility()
            if path == "/admin/time/create":
                return self.handle_create_time_entry()
            if path == "/admin/time/update":
                return self.handle_update_time_entry()
            if path == "/admin/time/delete":
                return self.handle_delete_time_entry()
            self.send_error(404)
        except RequestTooLarge as exc:
            self.send_html(page("Datei zu groß", f"<div class='error box narrow'>{esc(exc)}</div>", admin=self.is_admin(), buyer_key=self.current_buyer_key()), 413)

    def handle_buyer_login(self):
        form = self.read_form()
        requested_open_location = self.form_value(form, "open_location").strip()
        buyer_key = self.form_value(form, "buyer_key").strip()
        pin = self.form_value(form, "pin").strip()
        location = find_location(buyer_key)
        expected_password = (location or {}).get("password", "")
        if location and pin == expected_password:
            if not (location_can_order(location) or location_can_time(location)):
                return self.show_buyer_login("Für diese Rolle ist aktuell kein Bereich freigeschaltet.")
            target = "/cockpit"
            self.send_response(303)
            self.send_header("Set-Cookie", f"buyer={make_buyer_token(buyer_key)}; HttpOnly; SameSite=Lax; Path=/")
            self.send_header("Location", target)
            self.end_headers()
        else:
            self.show_buyer_login("Falscher Standort oder falsches Standort-Passwort.")

    def handle_cockpit_task_state(self):
        buyer_key = self.current_buyer_key()
        if not buyer_key:
            return self.redirect("/login")
        location = find_location(buyer_key)
        form = self.read_form()
        task_id = self.form_value(form, "task_id").strip()
        state = self.form_value(form, "state", "open").strip()
        completed_by = self.form_value(form, "completed_by").strip()
        valid_ids = {task["id"] for task in normalize_cockpit_tasks((location or {}).get("cockpit_tasks", []))}
        if task_id not in valid_ids:
            return self.redirect("/cockpit?msg=" + quote_plus("Aufgabe wurde nicht gefunden."))
        if state == "done" and completed_by not in get_time_employee_names(True):
            return self.redirect("/cockpit?msg=" + quote_plus("Bitte auswählen, wer die Aufgabe erledigt hat."))
        set_cockpit_state(buyer_key, "task", task_id, state, completed_by if state == "done" else "")
        return self.redirect("/cockpit")

    def handle_cockpit_order_state(self):
        buyer_key = self.current_buyer_key()
        if not buyer_key:
            return self.redirect("/login")
        location = find_location(buyer_key)
        form = self.read_form()
        order_id = self.form_value(form, "order_id").strip()
        state = self.form_value(form, "state", "open").strip()
        valid_ids = {item["id"] for item in normalize_cockpit_orders((location or {}).get("cockpit_orders", []))}
        if order_id not in valid_ids:
            return self.redirect("/cockpit?msg=" + quote_plus("Bestellinformation wurde nicht gefunden."))
        set_cockpit_state(buyer_key, "order", order_id, state)
        return self.redirect("/cockpit")

    def handle_cockpit_production_job_state(self):
        buyer_key = self.current_buyer_key()
        if not buyer_key:
            return self.redirect("/login")
        location = find_location(buyer_key)
        if not location or not (is_production_location(location) or is_schwarzenbek_location(location)):
            return self.redirect("/cockpit?msg=" + quote_plus("Produktionsauftrag wurde nicht gefunden."))
        form = self.read_form()
        job_id = self.form_value(form, "job_id").strip()
        state = self.form_value(form, "state", "open").strip()
        completed_by = self.form_value(form, "completed_by").strip()
        owner_location = location if is_production_location(location) else find_production_job_owner(job_id)
        valid_ids = {item["id"] for item in normalize_production_jobs((owner_location or {}).get("production_jobs", []))}
        if not owner_location or job_id not in valid_ids:
            return self.redirect("/cockpit?msg=" + quote_plus("Produktionsauftrag wurde nicht gefunden."))
        if state == "started" and not completed_by:
            return self.redirect("/cockpit?msg=" + quote_plus("Bitte eintragen, wer den Auftrag angefangen hat."))
        if state == "done" and not completed_by:
            previous = get_cockpit_state_details(owner_location["id"]).get(("production_job", job_id), {})
            completed_by = previous.get("completed_by", "")
        set_cockpit_state(owner_location["id"], "production_job", job_id, state, completed_by)
        return self.redirect("/cockpit")

    def handle_api_production_job(self):
        if not COCKPIT_API_TOKEN:
            return self.send_json({"ok": False, "error": "Schnittstelle ist nicht eingerichtet."}, 403)
        auth = self.headers.get("Authorization", "")
        api_key = self.headers.get("X-API-Key", "")
        expected = f"Bearer {COCKPIT_API_TOKEN}"
        if not (hmac.compare_digest(auth, expected) or hmac.compare_digest(api_key, COCKPIT_API_TOKEN)):
            return self.send_json({"ok": False, "error": "Nicht erlaubt."}, 401)
        payload = self.read_json_payload(max_bytes=64 * 1024)
        section = str(payload.get("section") or "").strip().lower()
        customer_name = str(payload.get("customer_name") or payload.get("name") or payload.get("title") or "").strip()
        order_text = str(payload.get("order_text") or payload.get("what") or payload.get("title") or "").strip()
        title = customer_name or order_text
        due_date = str(payload.get("due_date") or "").strip()
        note = str(payload.get("note") or "").strip()
        location_hint = str(payload.get("location") or "produktion").strip()
        if section not in PRODUCTION_SECTIONS:
            return self.send_json({"ok": False, "error": "Bereich muss backen, eisvitrinen oder eistorten sein."}, 400)
        if not title:
            return self.send_json({"ok": False, "error": "Name oder bestellter Inhalt fehlt."}, 400)
        if not is_valid_iso_date(due_date):
            return self.send_json({"ok": False, "error": "Datum muss im Format JJJJ-MM-TT angegeben werden."}, 400)
        locations = get_locations()
        target_location = None
        normalized_hint = normalize_text_key(location_hint)
        for location in locations:
            if normalize_text_key(location.get("id", "")) == normalized_hint or normalize_text_key(location.get("name", "")) == normalized_hint:
                target_location = location
                break
        if not target_location:
            for location in locations:
                if is_production_location(location):
                    target_location = location
                    break
        if not target_location:
            return self.send_json({"ok": False, "error": "Standort Produktion wurde nicht gefunden."}, 404)
        job = {
            "id": make_cockpit_item_id("production"),
            "section": section,
            "title": title,
            "customer_name": customer_name,
            "order_text": order_text,
            "due_date": due_date,
            "note": note,
            "active": True,
            "source": str(payload.get("source") or "api").strip()[:40] or "api",
        }
        updated_locations = []
        for location in locations:
            if location["id"] == target_location["id"]:
                jobs = normalize_production_jobs(location.get("production_jobs", []))
                jobs.append(job)
                location = dict(location)
                location["production_jobs"] = normalize_production_jobs(jobs)
            updated_locations.append(location)
        save_locations(updated_locations)
        return self.send_json({"ok": True, "id": job["id"], "location": target_location["name"]})

    def handle_admin_login(self):
        form = self.read_form()
        pin = self.form_value(form, "pin")
        settings = load_settings()
        if verify_pin(pin, settings.get("admin_pin_hash")):
            self.send_response(303)
            self.send_header("Set-Cookie", f"admin={make_admin_token()}; HttpOnly; SameSite=Lax; Path=/")
            self.send_header("Location", "/admin")
            self.end_headers()
        else:
            self.show_admin_login("Falsche PIN.")

    def handle_add_product(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        form = self.read_form()
        try:
            bulk_count = int(self.form_value(form, "bulk_add_count", "0") or "0")
        except ValueError:
            bulk_count = 0
        if bulk_count > 0:
            prepared = []
            for index in range(bulk_count):
                name = self.form_value(form, f"bulk_name_{index}").strip()
                category_value = self.form_value(form, f"bulk_category_{index}").strip() or "Allgemein"
                package_size = self.form_value(form, f"bulk_package_size_{index}").strip()
                source = self.form_value(form, f"bulk_source_{index}").strip()
                if not any([name, category_value and category_value != "Allgemein", package_size, source]):
                    continue
                if not (name and package_size and source):
                    continue
                categories = ensure_categories(split_categories(category_value))
                prepared.append((name, category_text(categories), package_size, source))
            if not prepared:
                return self.redirect("/admin?error=" + quote_plus("Bitte mindestens eine vollständige Produktzeile ausfüllen."))
            visible_to_text = ALL_LOCATIONS_KEY
            con = db()
            now = datetime.now().isoformat(timespec="seconds")
            for name, category, package_size, source in prepared:
                con.execute(
                    """
                    INSERT INTO products (name, package_size, category, source, visible_to, image_filename, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (name, package_size, category, source, visible_to_text, None, now),
                )
            con.commit()
            con.close()
            return self.redirect("/admin?msg=" + quote_plus(f"{len(prepared)} Produkt(e) angelegt."))
        name = self.form_value(form, "name").strip()
        categories = ensure_categories(form_categories(form, "category"))
        category = category_text(categories)
        package_size = self.form_value(form, "package_size").strip()
        source = self.form_value(form, "source").strip()
        visible_to = [location["id"] for location in get_locations() if self.form_value(form, f"visible_{location['id']}")]
        if not visible_to:
            visible_to = location_ids()
        visible_to_text = ALL_LOCATIONS_KEY if set(visible_to) == set(location_ids()) else ",".join(visible_to)
        image_filename = None
        if "image" in form and isinstance(form["image"], UploadedFile):
            item = form["image"]
            if item.filename and item.content:
                if len(item.content) > MAX_IMAGE_BYTES:
                    return self.redirect("/admin?error=" + quote_plus("Das Produktbild ist zu groß. Bitte maximal 6 MB hochladen."))
                image_filename = slug_filename(item.filename)
                with open(os.path.join(UPLOAD_DIR, image_filename), "wb") as f:
                    f.write(item.content)
        if name and package_size and source and categories:
            con = db()
            con.execute(
                """
                INSERT INTO products (name, package_size, category, source, visible_to, image_filename, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (name, package_size, category, source, visible_to_text, image_filename, datetime.now().isoformat(timespec="seconds")),
            )
            con.commit()
            con.close()
            self.redirect("/admin?msg=" + quote_plus("Produkt gespeichert."))
        else:
            self.redirect("/admin?error=" + quote_plus("Bitte Produktname, mindestens eine Kategorie, Gebindegröße und Bezugsquelle ausfüllen."))

    def handle_delete_product(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        form = self.read_form()
        pid = self.form_value(form, "id")
        con = db()
        con.execute(
            "UPDATE products SET active=0, visible_to=?, deleted_at=? WHERE id=? AND deleted_at IS NULL",
            (NO_LOCATIONS_KEY, berlin_now().isoformat(timespec="seconds"), pid),
        )
        con.commit()
        changed = con.total_changes
        con.close()
        if changed:
            self.redirect("/admin?msg=" + quote_plus("Produkt gelöscht. Bereits gespeicherte Bestellungen behalten ihre Produktdaten."))
        else:
            self.redirect("/admin?error=" + quote_plus("Produkt wurde nicht gefunden oder war bereits gelöscht."))

    def handle_update_product(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        form = self.read_form()
        product_id = self.form_value(form, "id").strip()
        product = get_product(product_id)
        if not product:
            return self.redirect("/admin?error=" + quote_plus("Produkt wurde nicht gefunden."))
        name = self.form_value(form, "name").strip()
        categories = ensure_categories(form_categories(form, "category", product_categories(product)))
        category = category_text(categories)
        package_size = self.form_value(form, "package_size").strip()
        source = self.form_value(form, "source").strip()
        active = 1 if self.form_value(form, "active", "1") == "1" else 0
        visible_to = [location["id"] for location in get_locations() if self.form_value(form, f"visible_{location['id']}")]
        if not visible_to:
            visible_to = location_ids()
        visible_to_text = ALL_LOCATIONS_KEY if set(visible_to) == set(location_ids()) else ",".join(visible_to)
        if not (name and package_size and source and categories):
            return self.redirect(f"/admin/edit-product?id={quote_plus(product_id)}&error=" + quote_plus("Bitte Produktname, mindestens eine Kategorie, Gebindegröße und Bezugsquelle ausfüllen."))
        image_filename = product["image_filename"]
        if "image" in form and isinstance(form["image"], UploadedFile):
            item = form["image"]
            if item.filename and item.content:
                if len(item.content) > MAX_IMAGE_BYTES:
                    return self.redirect(f"/admin/edit-product?id={quote_plus(product_id)}&error=" + quote_plus("Das Produktbild ist zu groß. Bitte maximal 6 MB hochladen."))
                image_filename = slug_filename(item.filename)
                with open(os.path.join(UPLOAD_DIR, image_filename), "wb") as f:
                    f.write(item.content)
        con = db()
        con.execute(
            """
            UPDATE products
            SET name=?, category=?, package_size=?, source=?, visible_to=?, image_filename=?, active=?
            WHERE id=?
            """,
            (name, category, package_size, source, visible_to_text, image_filename, active, product_id),
        )
        con.commit()
        con.close()
        self.redirect("/admin?msg=" + quote_plus("Produktänderungen gespeichert."))

    def handle_bulk_products(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        form = self.read_form()
        product_ids = [key.replace("selected_", "", 1) for key, value in form.items() if key.startswith("selected_") and value]
        product_ids = [pid for pid in product_ids if str(pid).isdigit()]
        if not product_ids:
            return self.redirect("/admin?error=" + quote_plus("Bitte zuerst mindestens ein Produkt auswählen."))

        placeholders = ",".join("?" for _ in product_ids)
        action = self.form_value(form, "bulk_action", "update")
        con = db()
        try:
            if action == "delete":
                con.execute(
                    f"UPDATE products SET active=0, visible_to=?, deleted_at=? WHERE id IN ({placeholders}) AND deleted_at IS NULL",
                    [NO_LOCATIONS_KEY, berlin_now().isoformat(timespec="seconds"), *product_ids],
                )
                con.commit()
                return self.redirect("/admin?msg=" + quote_plus(f"{con.total_changes} Produkt(e) gelöscht. Bereits gespeicherte Bestellungen behalten ihre Produktdaten."))

            updates = []
            params = []
            bulk_categories = [category for category in get_category_names(True) if self.form_value(form, category_input_name("bulk_category", category))]
            bulk_source = self.form_value(form, "bulk_source").strip()
            bulk_active = self.form_value(form, "bulk_active", "keep")
            visibility_mode = self.form_value(form, "bulk_visibility_mode", "keep")

            if bulk_categories:
                updates.append("category=?")
                params.append(category_text(ensure_categories(bulk_categories)))
            if bulk_source:
                updates.append("source=?")
                params.append(bulk_source)
            if bulk_active in ["0", "1"]:
                updates.append("active=?")
                params.append(1 if bulk_active == "1" else 0)
            if visibility_mode == "all":
                updates.append("visible_to=?")
                params.append(ALL_LOCATIONS_KEY)
            elif visibility_mode == "custom":
                visible_to = [location["id"] for location in get_locations() if self.form_value(form, f"bulk_visible_{location['id']}")]
                if not visible_to:
                    con.close()
                    return self.redirect("/admin?error=" + quote_plus("Bitte für die neue Sichtbarkeit mindestens einen Standort auswählen."))
                visible_to_text = ALL_LOCATIONS_KEY if set(visible_to) == set(location_ids()) else ",".join(visible_to)
                updates.append("visible_to=?")
                params.append(visible_to_text)

            if not updates:
                con.close()
                return self.redirect("/admin?error=" + quote_plus("Bitte mindestens eine Änderung für die ausgewählten Produkte auswählen."))

            params.extend(product_ids)
            con.execute(f"UPDATE products SET {', '.join(updates)} WHERE id IN ({placeholders}) AND deleted_at IS NULL", params)
            con.commit()
            return self.redirect("/admin?msg=" + quote_plus(f"{len(product_ids)} Produkt(e) aktualisiert."))
        finally:
            con.close()

    def handle_combined_order_pdf(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        form = self.read_form()
        order_ids = [key.replace("order_", "", 1) for key, value in form.items() if key.startswith("order_") and value]
        orders = get_orders_by_ids(order_ids)
        if len(orders) < 1:
            return self.redirect("/admin/orders?error=" + quote_plus("Bitte mindestens eine Bestellung auswählen."))
        try:
            pdf_filename = create_combined_order_pdf(orders)
            with open(os.path.join(ORDER_DIR, pdf_filename), "rb") as f:
                store_pdf_file(pdf_filename, f.read())
        except Exception as exc:
            print(f"Gesamtbestellung-PDF fehlgeschlagen: {exc}")
            return self.redirect("/admin/orders?error=" + quote_plus("Die Gesamtbestellung konnte nicht als PDF erstellt werden."))
        return self.redirect("/admin/orders?msg=" + quote_plus(f"Gesamtbestellung aus {len(orders)} Bestellung(en) wurde erstellt.") + "&pdf=" + quote_plus(pdf_filename))

    def handle_delete_orders(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        form = self.read_form()
        return_to = self.form_value(form, "return_to", "/admin/orders").strip()
        if return_to not in ["/admin", "/admin/orders"]:
            return_to = "/admin/orders"
        order_ids = [key.replace("order_", "", 1) for key, value in form.items() if key.startswith("order_") and value]
        if not order_ids:
            return self.redirect(return_to + "?error=" + quote_plus("Bitte zuerst mindestens eine Bestellung auswählen."))
        deleted_count = delete_orders_by_ids(order_ids)
        if not deleted_count:
            return self.redirect(return_to + "?error=" + quote_plus("Die ausgewählte Bestellung wurde nicht gefunden."))
        return self.redirect(return_to + "?msg=" + quote_plus(f"{deleted_count} Bestellung(en) gelöscht."))

    def handle_order_status(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        form = self.read_form()
        order_ids = [key.replace("order_", "", 1) for key, value in form.items() if key.startswith("order_") and value]
        if not order_ids:
            return self.redirect("/admin/orders?error=" + quote_plus("Bitte zuerst mindestens eine Bestellung auswählen."))
        completed = self.form_value(form, "completed", "1") == "1"
        changed = set_orders_completed(order_ids, completed=completed)
        if not changed:
            return self.redirect("/admin/orders?error=" + quote_plus("Die ausgewählte Bestellung wurde nicht gefunden."))
        message = "Bestellung als erledigt markiert." if completed else "Bestellung wieder geöffnet."
        return self.redirect("/admin/orders?msg=" + quote_plus(message))

    def handle_order_item_status(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        form = self.read_form()
        return_order = self.form_value(form, "return_order").strip()
        return_location = self.form_value(form, "return_location").strip()
        return_target = "/admin/orders"
        if return_location:
            return_target += "?open_location=" + quote_plus(return_location)
        if return_order.isdigit():
            return_target += "#order-" + return_order
        item_ids = [key.replace("item_", "", 1) for key, value in form.items() if key.startswith("item_") and value]
        if not item_ids:
            return self.redirect("/admin/orders?error=" + quote_plus("Bitte zuerst mindestens eine Position auswählen."))
        completed = self.form_value(form, "completed", "1") == "1"
        changed = set_order_items_completed(item_ids, completed=completed)
        if not changed:
            return self.redirect("/admin/orders?error=" + quote_plus("Die ausgewählte Position wurde nicht gefunden."))
        return self.redirect(return_target)

    def handle_delete_order_items(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        form = self.read_form()
        return_order = self.form_value(form, "return_order").strip()
        return_location = self.form_value(form, "return_location").strip()
        return_target = "/admin/orders"
        if return_location:
            return_target += "?open_location=" + quote_plus(return_location)
        if return_order.isdigit():
            return_target += "#order-" + return_order
        item_ids = [key.replace("item_", "", 1) for key, value in form.items() if key.startswith("item_") and value]
        if not item_ids:
            return self.redirect("/admin/orders?error=" + quote_plus("Bitte zuerst mindestens eine Position auswählen."))
        changed = delete_order_items_by_ids(item_ids)
        if not changed:
            return self.redirect("/admin/orders?error=" + quote_plus("Die ausgewählte Position wurde nicht gefunden."))
        return self.redirect(return_target)

    def handle_add_category(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        form = self.read_form()
        name = self.form_value(form, "name").strip()
        if name:
            ensure_category(name)
        self.redirect("/admin/categories?msg=" + quote_plus("Kategorie gespeichert."))

    def handle_delete_category(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        form = self.read_form()
        name = self.form_value(form, "name").strip()
        if name and category_key(name) != category_key("Allgemein"):
            con = db()
            product_rows = con.execute("SELECT category FROM products WHERE deleted_at IS NULL").fetchall()
            product_count = sum(1 for row in product_rows if product_has_category(row["category"], name))
            if product_count:
                con.close()
                return self.redirect("/admin/categories?error=" + quote_plus("Kategorie kann nicht gelöscht werden, solange noch Produkte zugeordnet sind. Bitte Produkte vorher bearbeiten."))
            cur = con.execute(
                "UPDATE categories SET active=0, deleted_at=? WHERE category_key=? AND deleted_at IS NULL",
                (berlin_now().isoformat(timespec="seconds"), category_key(name)),
            )
            con.commit()
            deleted = cur.rowcount
            con.close()
            if not deleted:
                return self.redirect("/admin/categories?error=" + quote_plus("Kategorie wurde nicht gefunden oder war bereits gelöscht."))
        self.redirect("/admin/categories?msg=" + quote_plus("Kategorie gelöscht."))

    def handle_import_products(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        form = self.read_form()
        uploaded = form.get("import_file")
        if not isinstance(uploaded, UploadedFile) or not uploaded.content:
            return self.redirect("/admin/import?error=" + quote_plus("Bitte eine CSV- oder Excel-Datei auswählen."))
        if len(uploaded.content) > MAX_IMPORT_BYTES:
            return self.redirect("/admin/import?error=" + quote_plus("Die Importdatei ist zu groß. Bitte maximal 5 MB hochladen."))
        try:
            rows = parse_import_file(uploaded)
        except Exception as exc:
            return self.redirect("/admin/import?error=" + quote_plus(str(exc)))
        inserted = 0
        skipped = 0
        prepared_rows = []
        categories_to_ensure = set()
        for row in rows:
            name = (row.get("name") or "").strip()
            package_size = (row.get("package_size") or "").strip()
            source = (row.get("source") or "").strip()
            if not (name and package_size and source):
                skipped += 1
                continue
            row_categories = split_categories(row.get("category") or "Allgemein")
            for category in row_categories:
                categories_to_ensure.add(category)
            prepared_rows.append(
                (name, package_size, category_text(row_categories), source, ALL_LOCATIONS_KEY, None, datetime.now().isoformat(timespec="seconds"))
            )
        con = db()
        category_map = {}
        for category in sorted(categories_to_ensure, key=lambda x: x.lower()):
            category_map[category] = ensure_category_in_connection(con, category, active=1)
        sync_category_metadata(con)
        for prepared in prepared_rows:
            prepared = list(prepared)
            prepared[2] = category_text([category_map.get(category, category) for category in split_categories(prepared[2])])
            con.execute(
                """
                INSERT INTO products (name, package_size, category, source, visible_to, image_filename, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                tuple(prepared),
            )
            inserted += 1
        con.commit()
        con.close()
        self.redirect("/admin/import?msg=" + quote_plus(f"Import abgeschlossen: {inserted} Produkt(e) angelegt, {skipped} Zeile(n) übersprungen."))

    def handle_settings(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        form = self.read_form()
        settings = load_settings()
        if self.form_value(form, "change_admin_pin"):
            current_pin = self.form_value(form, "current_admin_pin").strip()
            new_pin = self.form_value(form, "new_admin_pin").strip()
            confirm_pin = self.form_value(form, "new_admin_pin_confirm").strip()
            if not verify_pin(current_pin, settings.get("admin_pin_hash")):
                return self.redirect("/admin/settings?error=" + quote_plus("Die aktuelle Admin-PIN ist falsch."))
            if len(new_pin) < 4:
                return self.redirect("/admin/settings?error=" + quote_plus("Die neue Admin-PIN muss mindestens 4 Zeichen haben."))
            if new_pin != confirm_pin:
                return self.redirect("/admin/settings?error=" + quote_plus("Die neue Admin-PIN und die Wiederholung stimmen nicht überein."))
            settings["admin_pin_hash"] = hash_pin(new_pin)
            save_settings(settings)
            return self.redirect("/admin/settings?msg=" + quote_plus("Admin-PIN gespeichert. Bitte beim nächsten Login die neue PIN verwenden."))
        settings["order_email_to"] = self.form_value(form, "order_email_to").strip()
        settings["whatsapp_number"] = self.form_value(form, "whatsapp_number").strip()
        settings["time_export_email"] = self.form_value(form, "time_export_email").strip() or "info@opapeters"
        save_settings(settings)
        self.redirect("/admin/settings?msg=" + quote_plus("Einstellungen gespeichert."))

    def handle_employees(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        form = self.read_form()
        try:
            count = int(self.form_value(form, "employee_count", "0") or "0")
        except ValueError:
            count = 0
        employees = []
        for index in range(count):
            if self.form_value(form, f"employee_remove_{index}"):
                continue
            name = self.form_value(form, f"employee_name_{index}").strip()
            if not name:
                continue
            employees.append({
                "name": name,
                "active": bool(self.form_value(form, f"employee_active_{index}")),
                "hourly_wage": parse_decimal(self.form_value(form, f"employee_hourly_wage_{index}")),
                "monthly_hours": parse_decimal(self.form_value(form, f"employee_monthly_hours_{index}")),
            })
        new_name = self.form_value(form, "new_employee_name").strip()
        if new_name:
            employees.append({
                "name": new_name,
                "active": True,
                "hourly_wage": parse_decimal(self.form_value(form, "new_employee_hourly_wage")),
                "monthly_hours": parse_decimal(self.form_value(form, "new_employee_monthly_hours")),
            })
        employees = normalize_time_employees(employees)
        save_time_employees(employees)
        active_count = sum(1 for employee in employees if employee.get("active"))
        self.redirect("/admin/employees?msg=" + quote_plus(f"Personen gespeichert. {active_count} aktive Person(en) verfügbar."))


    def collect_min_stock_items(self, form, prefix):
        try:
            count = int(self.form_value(form, f"{prefix}_count", "0") or "0")
        except ValueError:
            count = 0
        items = []
        for row_index in range(count):
            product_id = self.form_value(form, f"{prefix}_product_{row_index}").strip()
            quantity = self.form_value(form, f"{prefix}_qty_{row_index}").strip()
            if not product_id and not quantity:
                continue
            items.append({"product_id": product_id, "quantity": quantity})
        return normalize_min_stock_items(items)

    def collect_cockpit_tasks(self, form, prefix):
        try:
            count = int(self.form_value(form, f"{prefix}_task_count", "0") or "0")
        except ValueError:
            count = 0
        items = []
        for row_index in range(count):
            title = self.form_value(form, f"{prefix}_task_title_{row_index}").strip()
            if not title:
                continue
            items.append({
                "id": self.form_value(form, f"{prefix}_task_id_{row_index}").strip(),
                "title": title,
                "template": bool(self.form_value(form, f"{prefix}_task_template_{row_index}")),
                "active": bool(self.form_value(form, f"{prefix}_task_active_{row_index}")),
                "image_filename": self.save_cockpit_image_upload(
                    form,
                    f"{prefix}_task_image_{row_index}",
                    self.form_value(form, f"{prefix}_task_image_current_{row_index}").strip(),
                ),
            })
        return normalize_cockpit_tasks(items)

    def collect_cockpit_orders(self, form, prefix):
        try:
            count = int(self.form_value(form, f"{prefix}_order_count", "0") or "0")
        except ValueError:
            count = 0
        items = []
        for row_index in range(count):
            title = self.form_value(form, f"{prefix}_order_title_{row_index}").strip()
            note = self.form_value(form, f"{prefix}_order_note_{row_index}").strip()
            if not title and not note:
                continue
            items.append({
                "id": self.form_value(form, f"{prefix}_order_id_{row_index}").strip(),
                "title": title,
                "note": note,
                "active": bool(self.form_value(form, f"{prefix}_order_active_{row_index}")),
                "image_filename": self.save_cockpit_image_upload(
                    form,
                    f"{prefix}_order_image_{row_index}",
                    self.form_value(form, f"{prefix}_order_image_current_{row_index}").strip(),
                ),
            })
        return normalize_cockpit_orders(items)


    def save_cockpit_image_upload(self, form, field_name, current_filename=""):
        current_filename = os.path.basename(str(current_filename or "").strip())
        uploaded = form.get(field_name)
        if not isinstance(uploaded, UploadedFile) or not uploaded.filename or not uploaded.content:
            return current_filename
        if len(uploaded.content) > MAX_IMAGE_BYTES:
            raise RequestTooLarge("Das Bild ist zu groß. Bitte maximal 6 MB hochladen.")
        image_filename = slug_filename(uploaded.filename)
        with open(os.path.join(UPLOAD_DIR, image_filename), "wb") as f:
            f.write(uploaded.content)
        return image_filename


    def collect_production_jobs(self, form, prefix):
        try:
            count = int(self.form_value(form, f"{prefix}_production_count", "0") or "0")
        except ValueError:
            count = 0
        items = []
        for row_index in range(count):
            customer_name = self.form_value(form, f"{prefix}_production_customer_name_{row_index}").strip()
            legacy_title = self.form_value(form, f"{prefix}_production_title_{row_index}").strip()
            order_text = self.form_value(form, f"{prefix}_production_order_text_{row_index}").strip()
            due_date = self.form_value(form, f"{prefix}_production_due_date_{row_index}").strip()
            section = self.form_value(form, f"{prefix}_production_section_{row_index}").strip()
            note = self.form_value(form, f"{prefix}_production_note_{row_index}").strip()
            if not customer_name and not legacy_title and not order_text and not due_date and not note:
                continue
            items.append({
                "id": self.form_value(form, f"{prefix}_production_id_{row_index}").strip(),
                "section": section,
                "title": customer_name or legacy_title or order_text,
                "customer_name": customer_name or legacy_title,
                "order_text": order_text,
                "due_date": due_date,
                "note": note,
                "active": bool(self.form_value(form, f"{prefix}_production_active_{row_index}")),
                "source": "admin",
                "image_filename": self.save_cockpit_image_upload(
                    form,
                    f"{prefix}_production_image_{row_index}",
                    self.form_value(form, f"{prefix}_production_image_current_{row_index}").strip(),
                ),
            })
        return normalize_production_jobs(items)


    def handle_locations(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        form = self.read_form()
        try:
            location_count = int(self.form_value(form, "location_count", "0") or "0")
        except ValueError:
            location_count = 0
        locations = []
        existing_category_updates = {}
        categories = get_category_names(True)
        for index in range(location_count):
            if self.form_value(form, f"location_remove_{index}"):
                continue
            name = self.form_value(form, f"location_name_{index}").strip()
            if not name:
                continue
            location_id = self.form_value(form, f"location_id_{index}").strip() or make_location_id(name)
            locations.append(
                {
                    "id": location_id,
                    "name": name,
                    "role": self.form_value(form, f"location_role_{index}", "standard").strip(),
                    "contact_name": self.form_value(form, f"location_contact_name_{index}").strip(),
                    "password": self.form_value(form, f"location_password_{index}").strip(),
                    "time_tracking_enabled": bool(self.form_value(form, f"location_time_tracking_{index}")),
                    "time_tracking_max_end": self.form_value(form, f"location_time_tracking_max_end_{index}").strip(),
                    "min_stock_enabled": bool(self.form_value(form, f"location_min_stock_enabled_{index}")),
                    "min_stock_items": self.collect_min_stock_items(form, f"location_min_stock_{index}"),
                    "cockpit_tasks": self.collect_cockpit_tasks(form, f"location_cockpit_{index}"),
                    "cockpit_orders": self.collect_cockpit_orders(form, f"location_cockpit_{index}"),
                    "production_jobs": self.collect_production_jobs(form, f"location_cockpit_{index}"),
                }
            )
            existing_category_updates[location_id] = [
                category for category in categories
                if self.form_value(form, f"location_category_{index}_{category}")
            ]
        new_name = self.form_value(form, "new_location_name").strip()
        new_location_categories = [
            category for category in categories
            if self.form_value(form, f"new_location_category_{category}")
        ]
        if new_name:
            locations.append(
                {
                    "id": make_location_id(new_name),
                    "name": new_name,
                    "role": self.form_value(form, "new_location_role", "standard").strip(),
                    "contact_name": self.form_value(form, "new_location_contact_name").strip(),
                    "password": self.form_value(form, "new_location_password").strip(),
                    "time_tracking_enabled": bool(self.form_value(form, "new_location_time_tracking")),
                    "time_tracking_max_end": self.form_value(form, "new_location_time_tracking_max_end").strip(),
                    "min_stock_enabled": bool(self.form_value(form, "new_location_min_stock_enabled")),
                    "min_stock_items": self.collect_min_stock_items(form, "new_location_min_stock"),
                    "cockpit_tasks": self.collect_cockpit_tasks(form, "new_location_cockpit"),
                    "cockpit_orders": self.collect_cockpit_orders(form, "new_location_cockpit"),
                    "production_jobs": self.collect_production_jobs(form, "new_location_cockpit"),
                }
            )
        if not locations:
            return self.redirect("/admin/locations?msg=" + quote_plus("Bitte mindestens einen Standort anlegen."))
        normalized_locations = normalize_locations(locations)
        new_location_id = normalized_locations[-1]["id"] if new_name and normalized_locations else ""
        save_locations(normalized_locations)
        valid_location_ids = {location["id"] for location in normalized_locations}
        if requested_open_location not in valid_location_ids:
            requested_open_location = new_location_id if new_location_id else ""
        for location_id, selected_categories in existing_category_updates.items():
            if location_id in valid_location_ids:
                apply_category_visibility_for_location(location_id, selected_categories)
        if new_location_id:
            apply_category_visibility_for_location(new_location_id, new_location_categories)
        target = "/admin/locations?msg=" + quote_plus("Standorte gespeichert.")
        if requested_open_location:
            target += "&open_location=" + quote_plus(requested_open_location)
            target += "#location-" + quote_plus(requested_open_location)
        self.redirect(target)

    def handle_visibility(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        form = self.read_form()
        location_id = self.form_value(form, "location_id").strip()
        if not find_location(location_id):
            return self.redirect("/admin/visibility?error=" + quote_plus("Bitte einen gültigen Standort auswählen."))
        selected_products = {
            key.replace("product_", "", 1)
            for key, value in form.items()
            if key.startswith("product_") and value
        }
        selected_categories = {
            key.replace("category_", "", 1)
            for key, value in form.items()
            if key.startswith("category_") and value
        }
        con = db()
        rows = con.execute("SELECT id, category, visible_to FROM products WHERE deleted_at IS NULL").fetchall()
        updated = 0
        for row in rows:
            visible = product_visible_location_keys(row["visible_to"])
            should_show = str(row["id"]) in selected_products or any(category in selected_categories for category in product_categories(row))
            if should_show and location_id not in visible:
                visible.append(location_id)
                updated += 1
            elif not should_show and location_id in visible:
                visible = [key for key in visible if key != location_id]
                updated += 1
            con.execute("UPDATE products SET visible_to=? WHERE id=?", (store_visible_locations(visible), row["id"]))
        con.commit()
        con.close()
        return self.redirect("/admin/visibility?location=" + quote_plus(location_id) + "&msg=" + quote_plus(f"Sichtbarkeit gespeichert. {updated} Produktzuordnung(en) geändert."))

    def handle_time_entry(self):
        buyer_key = self.current_buyer_key()
        if not buyer_key:
            return self.redirect("/login")
        location = find_location(buyer_key)
        if not location or not location_can_time(location):
            return self.redirect("/")
        form = self.read_form()
        employee_name = self.form_value(form, "employee_name").strip()
        work_location = self.form_value(form, "work_location").strip()
        start_time = self.form_value(form, "start_time").strip()
        end_time = self.form_value(form, "end_time").strip()
        note = self.form_value(form, "note").strip()
        allowed_employees = {normalize_text_key(name) for name in get_time_employee_names(True)}
        if not allowed_employees:
            return self.show_time_form("Bitte zuerst im Adminbereich Personen für die Zeiterfassung anlegen.")
        if normalize_text_key(employee_name) not in allowed_employees:
            return self.show_time_form("Bitte eine angelegte Person aus der Liste auswählen.")
        duration, error = validate_time_entry(location, employee_name, work_location, start_time, end_time, admin=False)
        if error:
            return self.show_time_form(error)
        if find_duplicate_time_entry(employee_name, today_iso(), start_time, end_time):
            return self.show_time_form("Schicht bereits registriert.")
        now = berlin_now().strftime("%d.%m.%Y %H:%M")
        con = db()
        con.execute(
            """
            INSERT INTO time_entries (location_id, location_name, employee_name, work_location, work_date, start_time, end_time, duration_minutes, note, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (buyer_key, location["name"], employee_name, work_location, today_iso(), start_time, end_time, duration, note, now),
        )
        con.commit()
        con.close()
        send_time_entry_push(employee_name, work_location, today_iso(), start_time, end_time, duration, note, created_by_admin=False)
        order_back_button = '<a class="button primary" href="/">Zurück zum Bestellsystem</a>' if location_can_order(location) else ""
        body = f"""
        <section class="box narrow success">
            <h2>Zeiterfassung gespeichert</h2>
            <p>Deine Zeiterfassung wurde erfolgreich gespeichert.</p>
            <p>Wir wünschen dir einen schönen Feierabend.</p>
            <p>{esc(farewell_message())}</p>
            <p>{order_back_button} <a class="button" href="/time">Weitere Zeit erfassen</a></p>
        </section>
        """
        self.send_html(page("Zeiterfassung gespeichert", body, buyer_key=buyer_key))

    def handle_create_time_entry(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        form = self.read_form()
        location_id = self.form_value(form, "location_id").strip()
        location = find_location(location_id)
        if not location:
            return self.redirect("/admin/time?error=" + quote_plus("Bitte einen gültigen Standort auswählen."))
        employee_name = self.form_value(form, "employee_name").strip()
        active_employees = {normalize_text_key(name) for name in get_time_employee_names(active_only=True)}
        if not active_employees:
            return self.redirect("/admin/time?error=" + quote_plus("Bitte zuerst unter Personen mindestens eine aktive Person anlegen."))
        if normalize_text_key(employee_name) not in active_employees:
            return self.redirect("/admin/time?error=" + quote_plus("Bitte eine aktive Person aus der Liste auswählen."))
        try:
            shift_count = min(max(int(self.form_value(form, "shift_count", "0") or 0), 1), 50)
        except ValueError:
            shift_count = 0
        work_location = location["name"]
        shift_rows = []
        first_month = current_month()
        for index in range(shift_count):
            work_date = self.form_value(form, f"shift_date_{index}").strip()
            start_time = self.form_value(form, f"shift_start_{index}").strip()
            end_time = self.form_value(form, f"shift_end_{index}").strip()
            note = self.form_value(form, f"shift_note_{index}").strip()
            if not any([start_time, end_time, note]):
                continue
            row_number = index + 1
            if not work_date or not start_time or not end_time:
                return self.redirect("/admin/time?error=" + quote_plus(f"Zeile {row_number}: Bitte Datum, Anfang und Ende ausfüllen oder die Zeile leer lassen."))
            try:
                datetime.strptime(work_date, "%Y-%m-%d")
            except ValueError:
                return self.redirect("/admin/time?error=" + quote_plus(f"Zeile {row_number}: Bitte ein gültiges Datum verwenden."))
            if not is_valid_hhmm(start_time):
                return self.redirect(f"/admin/time?month={quote_plus(work_date[:7])}&error=" + quote_plus(f"Zeile {row_number}: Bitte die Anfangszeit in 15-Minuten-Schritten auswählen."))
            duration, error = validate_time_entry(location, employee_name, work_location, start_time, end_time, admin=True)
            first_month = work_date[:7] if not shift_rows else first_month
            if error:
                return self.redirect(f"/admin/time?month={quote_plus(work_date[:7])}&error=" + quote_plus(f"Zeile {row_number}: {error}"))
            for existing_date, existing_start, existing_end, _existing_duration, _existing_note in shift_rows:
                if existing_date == work_date:
                    existing_start_minutes = minutes_from_clock_time(existing_start)
                    existing_end_minutes = minutes_from_hhmm(existing_end)
                    start_minutes = minutes_from_clock_time(start_time)
                    end_minutes = minutes_from_hhmm(end_time)
                    if existing_start_minutes is not None and existing_end_minutes is not None and start_minutes is not None and end_minutes is not None:
                        if existing_start_minutes < end_minutes and start_minutes < existing_end_minutes:
                            return self.redirect(f"/admin/time?month={quote_plus(work_date[:7])}&error=" + quote_plus(f"Zeile {row_number}: Schicht bereits registriert."))
            if find_duplicate_time_entry(employee_name, work_date, start_time, end_time):
                return self.redirect(f"/admin/time?month={quote_plus(work_date[:7])}&error=" + quote_plus(f"Zeile {row_number}: Schicht bereits registriert."))
            shift_rows.append((work_date, start_time, end_time, duration, note))
        if not shift_rows:
            return self.redirect("/admin/time?error=" + quote_plus("Bitte mindestens eine Schicht vollständig ausfüllen."))
        now = berlin_now().strftime("%d.%m.%Y %H:%M")
        con = db()
        for work_date, start_time, end_time, duration, note in shift_rows:
            con.execute(
                """
                INSERT INTO time_entries (location_id, location_name, employee_name, work_location, work_date, start_time, end_time, duration_minutes, note, created_at, updated_at, edited)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
                """,
                (location_id, location["name"], employee_name, work_location, work_date, start_time, end_time, duration, note, now, now),
            )
        con.commit()
        con.close()
        for work_date, start_time, end_time, duration, note in shift_rows:
            send_time_entry_push(employee_name, work_location, work_date, start_time, end_time, duration, note, created_by_admin=True)
        message = "Eine Schicht wurde manuell nachgetragen." if len(shift_rows) == 1 else f"{len(shift_rows)} Schichten wurden manuell nachgetragen."
        self.redirect(f"/admin/time?month={quote_plus(first_month)}&msg=" + quote_plus(message))

    def handle_update_time_entry(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        form = self.read_form()
        entry_id = self.form_value(form, "id").strip()
        entry = time_entry_by_id(entry_id)
        if not entry:
            return self.redirect("/admin/time?error=" + quote_plus("Zeiteintrag wurde nicht gefunden."))
        work_date = self.form_value(form, "work_date").strip()
        try:
            datetime.strptime(work_date, "%Y-%m-%d")
        except ValueError:
            return self.redirect("/admin/time?error=" + quote_plus("Bitte ein gültiges Datum verwenden."))
        employee_name = self.form_value(form, "employee_name").strip()
        work_location = self.form_value(form, "work_location").strip()
        start_time = self.form_value(form, "start_time").strip()
        end_time = self.form_value(form, "end_time").strip()
        note = self.form_value(form, "note").strip()
        duration, error = validate_time_entry(find_location(entry["location_id"]), employee_name, work_location, start_time, end_time, admin=True)
        month = work_date[:7]
        if error:
            return self.redirect(f"/admin/time?month={quote_plus(month)}&error=" + quote_plus(error))
        if find_duplicate_time_entry(employee_name, work_date, start_time, end_time, exclude_id=entry_id):
            return self.redirect(f"/admin/time?month={quote_plus(month)}&error=" + quote_plus("Schicht bereits registriert."))
        updated_at = berlin_now().strftime("%d.%m.%Y %H:%M")
        con = db()
        con.execute(
            """
            UPDATE time_entries
            SET employee_name=?, work_location=?, work_date=?, start_time=?, end_time=?, duration_minutes=?, note=?, updated_at=?, edited=1
            WHERE id=?
            """,
            (employee_name, work_location, work_date, start_time, end_time, duration, note, updated_at, entry_id),
        )
        con.commit()
        con.close()
        self.redirect(f"/admin/time?month={quote_plus(month)}&msg=" + quote_plus("Zeiteintrag gespeichert. Die Stunden wurden neu berechnet."))

    def handle_delete_time_entry(self):
        if not self.is_admin():
            return self.redirect("/admin/login")
        form = self.read_form()
        entry_id = self.form_value(form, "id").strip()
        month = self.form_value(form, "month", current_month()).strip()
        if not re.match(r"^\d{4}-\d{2}$", month):
            month = current_month()
        if not entry_id.isdigit():
            return self.redirect(f"/admin/time?month={quote_plus(month)}&error=" + quote_plus("Ungültiger Zeiteintrag."))
        con = db()
        cur = con.execute("DELETE FROM time_entries WHERE id=?", (entry_id,))
        con.commit()
        deleted = cur.rowcount
        con.close()
        if deleted:
            return self.redirect(f"/admin/time?month={quote_plus(month)}&msg=" + quote_plus("Zeiteintrag gelöscht."))
        return self.redirect(f"/admin/time?month={quote_plus(month)}&error=" + quote_plus("Zeiteintrag wurde nicht gefunden."))

    def handle_order(self):
        buyer_key = self.current_buyer_key()
        if not buyer_key:
            return self.redirect("/login")
        form = self.read_form()
        location_id = buyer_key
        location_data = find_location(location_id)
        if not location_can_order(location_data):
            return self.redirect("/time" if location_can_time(location_data) else "/choose")
        location = location_data["name"] if location_data else ""
        ordered_by = self.form_value(form, "ordered_by").strip()
        note = self.form_value(form, "note").strip()
        if not location or not ordered_by:
            return self.show_order_form("Bitte Standort und Namen des Bestellers ausfüllen.")

        products = get_products(True, buyer_key=buyer_key)
        items = []
        for p in products:
            raw = self.form_value(form, f"qty_{p['id']}", "0").strip() or "0"
            try:
                qty = int(raw)
            except ValueError:
                qty = 0
            qty = min(qty, MAX_ORDER_QUANTITY)
            if qty > 0:
                items.append(
                    {
                        "product_id": p["id"],
                        "product_name": p["name"],
                        "package_size": p["package_size"],
                        "category": category_text(product_categories(p)),
                        "source": p["source"],
                        "quantity": qty,
                    }
                )
        if not items:
            return self.show_order_form("Bitte bei mindestens einem Produkt eine Menge eintragen.")

        order_image_filename = None
        for image_field in ["order_image", "order_image_camera", "order_image_gallery"]:
            uploaded_image = form.get(image_field)
            if isinstance(uploaded_image, UploadedFile) and uploaded_image.filename and uploaded_image.content:
                if len(uploaded_image.content) > MAX_ORDER_IMAGE_BYTES:
                    return self.show_order_form("Das Bild zur Bestellung ist zu groß. Bitte maximal 10 MB hochladen.")
                order_image_filename = slug_filename(uploaded_image.filename)
                with open(os.path.join(ORDER_IMAGE_DIR, order_image_filename), "wb") as f:
                    f.write(uploaded_image.content)
                break

        order_number = berlin_now().strftime("%Y%m%d-%H%M%S") + "-" + uuid.uuid4().hex[:4].upper()
        created_at = berlin_now().strftime("%d.%m.%Y %H:%M")
        buyer_group = location
        order = {
            "order_number": order_number,
            "location": location,
            "ordered_by": ordered_by,
            "buyer_group": buyer_group,
            "note": note,
            "order_image_filename": order_image_filename,
            "created_at": created_at,
        }
        try:
            pdf_filename = create_pdf(order, items)
            with open(os.path.join(ORDER_DIR, pdf_filename), "rb") as f:
                pdf_data = f.read()
            store_pdf_file(pdf_filename, pdf_data)
        except Exception as exc:
            print(f"PDF-Erstellung fehlgeschlagen: {exc}")
            return self.show_order_form("Die PDF zur Bestellung konnte nicht erstellt werden. Bitte prüfe die Eingaben und versuche es erneut.")
        try:
            send_pdf_email_if_configured(order, pdf_filename)
        except Exception as exc:
            print(f"E-Mail-Versand fehlgeschlagen: {exc}")
        con = db()
        cur = con.cursor()
        cur.execute(
            "INSERT INTO orders (order_number, location, ordered_by, buyer_group, note, pdf_filename, pdf_data, order_image_filename, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (order_number, location, ordered_by, buyer_group, note, pdf_filename, pdf_data, order_image_filename, created_at),
        )
        order_id = cur.lastrowid
        for item in items:
            cur.execute(
                "INSERT INTO order_items (order_id, product_id, product_name, package_size, category, source, quantity) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (order_id, item["product_id"], item["product_name"], item["package_size"], item["category"], item["source"], item["quantity"]),
            )
        con.commit()
        con.close()
        delete_cart_draft(location_id)
        order_for_links = dict(order)
        order_for_links["pdf_filename"] = pdf_filename
        order_for_links["base_url"] = self.base_url()
        wa_link = whatsapp_order_link(order_for_links)
        whatsapp_button = f'<a class="button" target="_blank" rel="noopener" href="{esc(wa_link)}">Per WhatsApp senden</a>' if wa_link else ''
        cart_cleanup_script = f"""
        <script>
        (function () {{
          var locationKey = {json.dumps(location_id)};
          try {{
            window.localStorage.removeItem('opaPetersCart:' + locationKey);
            window.localStorage.removeItem('opaPetersCartBackup:' + locationKey);
          }} catch (error) {{}}
          try {{
            var xhr = new XMLHttpRequest();
            xhr.open('POST', '/cart-draft', true);
            xhr.setRequestHeader('Content-Type', 'application/json; charset=utf-8');
            xhr.send(JSON.stringify({{ action: 'clear' }}));
          }} catch (error) {{}}
        }})();
        </script>
        """
        body = f"""
        <section class="box narrow success">
            <h2>Bestellung gesendet</h2>
            <p>Die Bestellung <strong>{esc(order_number)}</strong> wurde erstellt.</p>
            <p>Deine Bestellung wurde erfolgreich übermittelt. Vielen Dank.</p>
            {f'<p>Das Bild zur Bestellung wurde mitgespeichert.</p>' if order_image_filename else ''}
            <p>Die PDF wurde erstellt und kann direkt geöffnet oder gedruckt werden.</p>
            <p><a class="button" href="/">Neue Bestellung erfassen</a> <a class="button primary" href="{esc(pdf_viewer_href(pdf_filename))}">PDF öffnen / drucken</a> {whatsapp_button}</p>
        </section>{cart_cleanup_script}"""
        self.send_html(page("Bestellung gesendet", body, buyer_key=buyer_key))


def main():
    init_db()
    seed_demo_products()
    print(f"Internes Bestellsystem läuft auf http://localhost:{PORT}")
    print(f"Admin-PIN: {ADMIN_PIN} (vor Onlinebetrieb ändern)")
    print("Standort-Zugänge werden im Adminbereich unter Standorte gepflegt.")
    HTTPServer((HOST, PORT), App).serve_forever()


if __name__ == "__main__":
    main()
